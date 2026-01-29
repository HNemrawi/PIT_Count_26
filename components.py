"""
UI Components module for PIT Count Application
Contains all interface components for upload, validation, reports, and download
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Tuple, Any


from config import VALID_AGE_RANGES, VALID_GENDERS, VALID_RACES
from processor import detect_duplicates, validate_data, map_name_columns_for_duplication
from utils import get_timezone_for_region, create_download_filename, get_current_timestamp, safe_dataframe_display, clean_dataframe_for_export

def show_upload_interface():
    """Show the data upload interface"""
    region = st.session_state.get('region')

    # Store files in session state if not already there
    if 'temp_files' not in st.session_state:
        st.session_state['temp_files'] = {}

    # Cache Excel file objects to avoid re-reading (performance optimization)
    if 'excel_files' not in st.session_state:
        st.session_state['excel_files'] = {}
    
    uploaded_data = {}
    
    st.write(f"**Selected Region:** {region}")
    st.write("---")
    
    # Emergency Shelter
    st.subheader("📁 Emergency Shelter (ES) Data")
    es_file = st.file_uploader(
        "Choose Emergency Shelter file",
        type=['csv', 'xlsx'],
        key="es_upload"
    )
    
    if es_file:
        if es_file.name.endswith('.xlsx'):
            # Store file in session state
            st.session_state['temp_files']['es'] = es_file

            # Read Excel file to get sheet names (cache the ExcelFile object)
            try:
                # Cache Excel file object if not already cached or if different file
                cache_key = f'es_{es_file.name}'
                if cache_key not in st.session_state['excel_files']:
                    excel_file = pd.ExcelFile(es_file, engine='calamine')
                    st.session_state['excel_files']['es'] = excel_file
                else:
                    excel_file = st.session_state['excel_files']['es']

                sheet_names = excel_file.sheet_names

                if len(sheet_names) > 1:
                    selected_sheet = st.selectbox(
                        "Select sheet for Emergency Shelter data:",
                        sheet_names,
                        key="es_sheet_select"
                    )
                else:
                    selected_sheet = sheet_names[0]

                # Show preview
                if st.button("Preview ES Data", key="preview_es"):
                    preview_df = excel_file.parse(selected_sheet, nrows=10)
                    preview_df_safe = safe_dataframe_display(preview_df)
                    st.dataframe(preview_df_safe, width='stretch')

            except Exception as e:
                st.error(f"Error reading Excel file: {str(e)}")
        else:
            # CSV file - no sheet selection needed
            st.session_state['temp_files']['es'] = es_file
    
    st.write("---")
    
    # Transitional Housing
    st.subheader("📁 Transitional Housing (TH) Data")
    th_file = st.file_uploader(
        "Choose Transitional Housing file",
        type=['csv', 'xlsx'],
        key="th_upload"
    )
    
    if th_file:
        if th_file.name.endswith('.xlsx'):
            st.session_state['temp_files']['th'] = th_file

            try:
                # Cache Excel file object if not already cached or if different file
                cache_key = f'th_{th_file.name}'
                if cache_key not in st.session_state['excel_files']:
                    excel_file = pd.ExcelFile(th_file, engine='calamine')
                    st.session_state['excel_files']['th'] = excel_file
                else:
                    excel_file = st.session_state['excel_files']['th']

                sheet_names = excel_file.sheet_names

                if len(sheet_names) > 1:
                    selected_sheet = st.selectbox(
                        "Select sheet for Transitional Housing data:",
                        sheet_names,
                        key="th_sheet_select"
                    )
                else:
                    selected_sheet = sheet_names[0]

                if st.button("Preview TH Data", key="preview_th"):
                    preview_df = excel_file.parse(selected_sheet, nrows=10)
                    preview_df_safe = safe_dataframe_display(preview_df)
                    st.dataframe(preview_df_safe, width='stretch')

            except Exception as e:
                st.error(f"Error reading Excel file: {str(e)}")
        else:
            st.session_state['temp_files']['th'] = th_file
    
    st.write("---")
    
    # Unsheltered
    st.subheader("📁 Unsheltered Data")
    unsheltered_file = st.file_uploader(
        "Choose Unsheltered file",
        type=['csv', 'xlsx'],
        key="unsheltered_upload"
    )
    
    if unsheltered_file:
        if unsheltered_file.name.endswith('.xlsx'):
            st.session_state['temp_files']['unsheltered'] = unsheltered_file

            try:
                # Cache Excel file object if not already cached or if different file
                cache_key = f'unsheltered_{unsheltered_file.name}'
                if cache_key not in st.session_state['excel_files']:
                    excel_file = pd.ExcelFile(unsheltered_file, engine='calamine')
                    st.session_state['excel_files']['unsheltered'] = excel_file
                else:
                    excel_file = st.session_state['excel_files']['unsheltered']

                sheet_names = excel_file.sheet_names

                if len(sheet_names) > 1:
                    selected_sheet = st.selectbox(
                        "Select sheet for Unsheltered data:",
                        sheet_names,
                        key="unsheltered_sheet_select"
                    )
                else:
                    selected_sheet = sheet_names[0]

                if st.button("Preview Unsheltered Data", key="preview_unsheltered"):
                    preview_df = excel_file.parse(selected_sheet, nrows=10)
                    preview_df_safe = safe_dataframe_display(preview_df)
                    st.dataframe(preview_df_safe, width='stretch')
                    
            except Exception as e:
                st.error(f"Error reading Excel file: {str(e)}")
        else:
            st.session_state['temp_files']['unsheltered'] = unsheltered_file

    st.write("---")

    # Process button
    if st.button("🔄 Process Uploaded Files", type="primary", width='stretch'):
        # Process files with selected sheets
        files_to_process = st.session_state.get('temp_files', {})
        
        if files_to_process:
            with st.spinner("Processing files..."):
                # Process ES file
                if 'es' in files_to_process:
                    es_file = files_to_process['es']
                    if es_file.name.endswith('.xlsx') and 'es_sheet_select' in st.session_state:
                        # Use cached Excel file object for faster reading
                        if 'es' in st.session_state['excel_files']:
                            es_df = st.session_state['excel_files']['es'].parse(st.session_state['es_sheet_select'])
                        else:
                            es_df = pd.read_excel(es_file, sheet_name=st.session_state['es_sheet_select'], engine='calamine')
                    else:
                        es_df = load_file_direct(es_file)

                    if es_df is not None:
                        uploaded_data["Sheltered_ES"] = es_df
                        st.success(f"✅ ES data: {len(es_df)} rows")

                # Process TH file
                if 'th' in files_to_process:
                    th_file = files_to_process['th']
                    if th_file.name.endswith('.xlsx') and 'th_sheet_select' in st.session_state:
                        # Use cached Excel file object for faster reading
                        if 'th' in st.session_state['excel_files']:
                            th_df = st.session_state['excel_files']['th'].parse(st.session_state['th_sheet_select'])
                        else:
                            th_df = pd.read_excel(th_file, sheet_name=st.session_state['th_sheet_select'], engine='calamine')
                    else:
                        th_df = load_file_direct(th_file)

                    if th_df is not None:
                        uploaded_data["Sheltered_TH"] = th_df
                        st.success(f"✅ TH data: {len(th_df)} rows")

                # Process Unsheltered file
                if 'unsheltered' in files_to_process:
                    unsheltered_file = files_to_process['unsheltered']
                    if unsheltered_file.name.endswith('.xlsx') and 'unsheltered_sheet_select' in st.session_state:
                        # Use cached Excel file object for faster reading
                        if 'unsheltered' in st.session_state['excel_files']:
                            unsheltered_df = st.session_state['excel_files']['unsheltered'].parse(st.session_state['unsheltered_sheet_select'])
                        else:
                            unsheltered_df = pd.read_excel(unsheltered_file, sheet_name=st.session_state['unsheltered_sheet_select'], engine='calamine')
                    else:
                        unsheltered_df = load_file_direct(unsheltered_file)

                    if unsheltered_df is not None:
                        uploaded_data["Unsheltered"] = unsheltered_df
                        st.success(f"✅ Unsheltered data: {len(unsheltered_df)} rows")

            if uploaded_data:
                # Validate data
                valid_data = {}
                for source_name, df in uploaded_data.items():
                    # Clean data
                    df.columns = df.columns.str.strip()
                    df = df.loc[:, ~df.columns.duplicated(keep='first')]
                    
                    # Remove rows without timestamps
                    if 'Timestamp' in df.columns:
                        initial_count = len(df)
                        df = df.dropna(subset=['Timestamp'])
                        dropped = initial_count - len(df)
                        if dropped > 0:
                            st.info(f"{source_name}: Removed {dropped} rows with missing timestamps")
                    
                    # Check essential columns (Gender is optional as not all regions collect it)
                    essential = ['Timestamp', 'Race/Ethnicity', 'Age Range']
                    optional = ['Gender']
                    missing = set(essential) - set(df.columns)
                    missing_optional = set(optional) - set(df.columns)

                    if missing:
                        st.error(f"❌ {source_name}: Missing columns: {', '.join(missing)}")
                    else:
                        if missing_optional:
                            st.info(f"ℹ️ {source_name}: Optional column(s) not found: {', '.join(missing_optional)} - Gender statistics will be unavailable")
                        valid_data[source_name] = df
                
                if valid_data:
                    # Clear temp files and Excel cache (data is now in uploaded_data)
                    st.session_state['temp_files'] = {}
                    st.session_state['excel_files'] = {}
                    st.session_state['uploaded_data'] = valid_data
                    return valid_data
                else:
                    st.error("Please fix the issues with your data and try again.")
        else:
            st.info("Please upload at least one data file to continue.")
    
    return {}

def load_file_direct(uploaded_file):
    """Load a file directly without sheet selection (optimized for performance)"""
    try:
        if uploaded_file.name.endswith('.csv'):
            # Use faster CSV reading with low_memory=False for mixed types
            return pd.read_csv(uploaded_file, low_memory=False)
        elif uploaded_file.name.endswith('.xlsx'):
            # Use calamine engine for faster Excel reading (fallback to openpyxl if needed)
            try:
                return pd.read_excel(uploaded_file, engine='calamine')
            except Exception:
                return pd.read_excel(uploaded_file, engine='openpyxl')
    except Exception as e:
        st.error(f"Error loading file: {str(e)}")
        return None

def load_file(uploaded_file):
    """Load a file into a dataframe (optimized for performance)"""
    try:
        if uploaded_file.name.endswith('.csv'):
            return pd.read_csv(uploaded_file, low_memory=False)
        elif uploaded_file.name.endswith('.xlsx'):
            # Handle multiple sheets with optimized engine
            excel_file = pd.ExcelFile(uploaded_file, engine='openpyxl')
            if len(excel_file.sheet_names) > 1:
                sheet_name = st.selectbox(
                    f"Select sheet from {uploaded_file.name}",
                    excel_file.sheet_names
                )
                return pd.read_excel(uploaded_file, sheet_name=sheet_name, engine='openpyxl')
            else:
                return pd.read_excel(uploaded_file, engine='openpyxl')
    except Exception as e:
        st.error(f"Error loading file: {str(e)}")
        return None

def show_validation_interface(uploaded_data, processed_data):
    """Show the validation and duplication detection interface"""
    region = st.session_state.get('region')
    
    # Create tabs
    tab1, tab2 = st.tabs(["🔍 Duplication Detection", "✅ Data Validation"])
    
    with tab1:
        show_duplication_interface(uploaded_data, region)
    
    with tab2:
        show_data_validation_interface(uploaded_data, region)

def show_duplication_interface(uploaded_data, region):
    """Show duplication detection interface"""
    st.subheader("🔍 Duplication Detection")
    
    with st.expander("ℹ️ How Duplication Detection Works", expanded=False):
        st.markdown(f"""
        ### Detection Logic (Current Region: {region})

        The system uses **region-specific hierarchical matching** to identify potential duplicates.
        Higher priority rules override lower ones - if a pair matches a "Likely" condition,
        it won't be downgraded even if it also matches "Somewhat Likely" or "Possible" conditions.

        ---

        ### New England Rules
        *Uses 3-letter name codes: 1st letter of first name + 1st & 3rd letters of last name*

        | Color | Confidence Level | Matching Conditions |
        |-------|-----------------|---------------------|
        | 🔴 | **Likely Duplicate** | Full Name (3-letter code) + DOB match |
        | 🟠 | **Somewhat Likely** | Full Name (3-letter code) + Age match |
        | 🟡 | **Possible Duplicate** | Full Name (3-letter code) + Age Range match |
        | 🟣 | **No Name Info** | Record has no name information (not matched) |

        **Name Format Example:** "John Smith" → "JSI" (J + S + I, where I is 3rd letter of Smith)

        ---

        ### Great Lakes Rules
        *Uses full names (First Name + Last Name) and initials*

        | Color | Confidence Level | Matching Conditions |
        |-------|-----------------|---------------------|
        | 🔴 | **Likely Duplicate** | Full Name + DOB match |
        | 🔴 | **Likely Duplicate** | Full Name + Age match |
        | 🔴 | **Likely Duplicate** | Initials + DOB match |
        | 🟠 | **Somewhat Likely** | Full Name + Age Range match |
        | 🟠 | **Somewhat Likely** | Initials + Age match |
        | 🟡 | **Possible Duplicate** | Initials + Age Range match |
        | 🟣 | **No Name Info** | Record has no name information (not matched) |

        **Name Format Example:** "John Smith" → Full name: "JOHN SMITH", Initials: "JS"

        ---

        ### Key Differences Between Regions

        | Feature | New England | Great Lakes |
        |---------|-------------|-------------|
        | Name Format | 3-letter code | Full names |
        | Initials Matching | Not used | Used for all confidence levels |
        | Full Name + Age | Somewhat Likely | Likely |
        | Matching Strictness | More strict (fewer rules) | More comprehensive (more rules) |

        ---

        ### Important Notes

        - **Hierarchy Enforcement**: Higher confidence matches prevent downgrading to lower confidence
        - **No Name Records**: Records without name information are flagged with 🟣 but not compared
        - **Heads of Household**: Comparison focuses on heads of household information
        - **Data Priority**: DOB is preferred over Age; Age is preferred over Age Range

        ### Excel Export

        - Color-coded rows (red/orange/yellow/purple)
        - "Duplicates_With" column shows Excel row numbers (starting at row 2)
        - "Duplication_Reason" explains why records matched
        """)
    
    if st.button("🚀 Run Duplication Detection", type="primary"):
        with st.spinner("Analyzing records..."):
            results = {}

            for source_name, df in uploaded_data.items():
                # Map name columns before detection to ensure standardized format
                df_with_mapped_names = map_name_columns_for_duplication(df, region)
                annotated = detect_duplicates(df_with_mapped_names, source_name, region)
                results[source_name] = annotated

            st.session_state['dup_results'] = results
            st.success("✅ Detection complete!")
    
    # Display results
    if 'dup_results' in st.session_state:
        for source_name, annotated in st.session_state['dup_results'].items():
            st.subheader(f"📊 Results: {source_name}")
            
            # Summary stats
            total = len(annotated)
            likely = (annotated['Duplication_Score'].str.contains('Likely Duplicate 🔴', na=False)).sum()
            somewhat = (annotated['Duplication_Score'].str.contains('Somewhat Likely', na=False)).sum()
            possible = (annotated['Duplication_Score'].str.contains('Possible', na=False)).sum()
            no_name = (annotated['Duplication_Score'].str.contains('No name', na=False)).sum()
            not_duplicate = (annotated['Duplication_Score'] == 'Not Duplicate').sum()
            
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric("Total", f"{total:,}")
            with col2:
                st.metric("🔴 Likely", likely)
            with col3:
                st.metric("🟠 Somewhat", somewhat)
            with col4:
                st.metric("🟡 Possible", possible)
            with col5:
                st.metric("🟣 No Name", no_name)
            
            # Summary message
            total_flagged = total - not_duplicate
            flagged_pct = (total_flagged / total * 100) if total > 0 else 0
            st.info(f"""
            **Summary**: {total_flagged} records ({flagged_pct:.1f}%) 
            flagged for review out of {total:,} total records.
            """)
            
            # Important note about indices
            st.warning("""
            **📌 Index Reference:**
            - **In this table**: `Duplicates_With` shows zero-based indices (0, 1, 2...)
            - **In Excel download**: Indices are adjusted to match Excel row numbers (2, 3, 4...)
            """)
            
            # Show data
            from utils import safe_dataframe_display
            st.dataframe(safe_dataframe_display(annotated), width='stretch', height=400)
            
            # Download buttons
            col1, col2 = st.columns(2)
            
            with col1:
                # CSV download
                csv_data = annotated.to_csv(index=False)
                st.download_button(
                    "📥 Download CSV",
                    data=csv_data,
                    file_name=f"{source_name}_duplicates.csv",
                    mime="text/csv",
                    key=f"dup_csv_{source_name}"
                )
            
            with col2:
                # Excel download with highlights
                from processor import DuplicationDetector
                # Create a new detector instance to generate Excel
                detector = DuplicationDetector(uploaded_data[source_name], source_name, region)
                excel_buffer = detector.create_excel_with_highlights(annotated)
                
                st.download_button(
                    "📥 Download Excel (with highlights)",
                    data=excel_buffer.getvalue(),
                    file_name=f"{source_name}_duplicates.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dup_excel_{source_name}",
                    help="Download color-coded Excel file with adjusted row numbers"
                )

def show_data_validation_interface(uploaded_data, region):
    """Show data validation interface"""
    st.subheader("✅ Data Validation")

    with st.expander("ℹ️ About Data Validation", expanded=False):
        st.markdown("""
        ### Validation Rules

        The system validates the following data fields:

        #### 🔢 Age Range (Single-Select)
        - **Valid Values**: Under 18, 18-24, 25-34, 35-44, 45-54, 55-64, 65+
        - **Applies To**: All age range columns (HOH + additional adults)
        - **Empty Values**: Allowed (skipped in validation)

        #### 👤 Sex (Single-Select, Required)
        - **Valid Values**: Male, Female
        - **Applies To**: All sex columns (HOH + additional adults + children)
        - **Empty Values**: Flagged as invalid

        #### 🎭 Gender (Multi-Select, Optional)
        - **Valid Values**: Man, Woman, Non-Binary, Transgender, Questioning, Different Identity
        - **Applies To**: All gender columns (HOH + additional adults + children)
        - **Format**: Comma-separated for multiple selections (e.g., "Woman, Transgender")
        - **Empty Values**: Allowed (optional field)
        - **Note**: Gender columns may not exist in all datasets - this is acceptable

        #### 🌍 Race/Ethnicity (Multi-Select)
        - **Valid Values**:
          - American Indian, Alaska Native, or Indigenous
          - Asian or Asian American
          - Black, African American, or African
          - Hispanic/Latina/e/o
          - Middle Eastern or North African
          - Native Hawaiian or Pacific Islander
          - White
          - Client Doesn't Know
          - Client Prefers Not to Answer
          - Data Not Collected
        - **Format**: Comma-separated for multiple selections
        - **Empty Values**: Allowed (skipped in validation)

        ### How Validation Works

        1. **Run Validation** button scans all uploaded data sources
        2. Invalid entries are flagged with their Excel row numbers
        3. Results show the invalid value and list of valid options
        4. Download options available for each issue type and complete reports

        ### What Gets Flagged

        - **Typos**: "Mal" instead of "Male"
        - **Extra Spaces**: " Male " with leading/trailing spaces
        - **Wrong Format**: "Male,Female" in a single-select field
        - **Invalid Options**: Values not in the approved list
        - **Required Fields**: Missing sex values (when column exists)
        """)

    st.info("Click **Run Validation** to check all uploaded data for invalid entries.")

    
    if st.button("🔍 Run Validation", type="primary", key="validate_btn"):
        all_results = {}
        
        for source_name, df in uploaded_data.items():
            with st.spinner(f"Validating {source_name}..."):
                validation_results = validate_data(df, source_name, region)
                if validation_results:
                    all_results[source_name] = validation_results
        
        st.session_state['validation_results'] = all_results
        
        if all_results:
            st.warning("⚠️ Validation issues found!")
        else:
            st.success("✅ All data is valid!")
    
    # Display results
    if 'validation_results' in st.session_state:
        results = st.session_state['validation_results']
        
        if not results:
            st.success("✅ No validation issues found!")
            return
        
        # Create comprehensive summary metrics
        total_issues = sum(sum(len(df) for df in source_results.values()) for source_results in results.values())

        # Count by type across all sources
        all_age_issues = sum(len(df) for source_results in results.values()
                            for k, df in source_results.items() if k.startswith('age_'))
        all_sex_issues = sum(len(df) for source_results in results.values()
                            for k, df in source_results.items() if k.startswith('sex_'))
        all_gender_issues = sum(len(df) for source_results in results.values()
                               for k, df in source_results.items() if k.startswith('gender_'))
        all_race_issues = sum(len(df) for source_results in results.values()
                             for k, df in source_results.items() if k.startswith('race_'))

        # Top row metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Issues", f"{total_issues:,}")
        with col2:
            st.metric("Sources with Issues", len(results))
        with col3:
            st.metric("Most Common",
                     "🔢 Age" if all_age_issues >= max(all_sex_issues, all_gender_issues, all_race_issues)
                     else "👤 Sex" if all_sex_issues >= max(all_age_issues, all_gender_issues, all_race_issues)
                     else "🎭 Gender" if all_gender_issues >= max(all_age_issues, all_sex_issues, all_race_issues)
                     else "🌍 Race")
        with col4:
            st.metric("Status", "⚠️ Needs Review")

        # Second row - breakdown by type
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("🔢 Age Issues", all_age_issues)
        with col2:
            st.metric("👤 Sex Issues", all_sex_issues)
        with col3:
            st.metric("🎭 Gender Issues", all_gender_issues)
        with col4:
            st.metric("🌍 Race Issues", all_race_issues)

        st.write("---")

        # Download all issues across all sources as Excel
        if total_issues > 0:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            # Create workbook with separate sheets per source
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                for source_name, validation_results in results.items():
                    if validation_results:
                        all_issues_df = pd.concat(validation_results.values(), ignore_index=True)
                        # Sort by Row number for easier navigation
                        all_issues_df = all_issues_df.sort_values('Row')
                        all_issues_df.to_excel(writer, sheet_name=source_name[:31], index=False)

            excel_buffer.seek(0)
            st.download_button(
                "📥 Download All Validation Issues (Excel)",
                data=excel_buffer.getvalue(),
                file_name=f"validation_issues_{region}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_all_validation_excel",
                help="Download all validation issues across all sources as an Excel file"
            )

            st.write("---")
        
        for source_name, validation_results in results.items():
            source_issues = sum(len(df) for df in validation_results.values())

            # Group by type
            age_issues = {k: v for k, v in validation_results.items() if k.startswith('age_')}
            sex_issues = {k: v for k, v in validation_results.items() if k.startswith('sex_')}
            gender_issues = {k: v for k, v in validation_results.items() if k.startswith('gender_')}
            race_issues = {k: v for k, v in validation_results.items() if k.startswith('race_')}

            # Count issues by type for this source
            age_count = sum(len(df) for df in age_issues.values())
            sex_count = sum(len(df) for df in sex_issues.values())
            gender_count = sum(len(df) for df in gender_issues.values())
            race_count = sum(len(df) for df in race_issues.values())

            with st.expander(f"📋 {source_name} - {source_issues} issues", expanded=True):
                # Summary for this source - only show categories that exist in data
                breakdown_lines = [
                    f"- 🔢 Age Range: {age_count} issues",
                    f"- 👤 Sex: {sex_count} issues"
                ]

                # Only show Gender if any Gender columns exist in the dataset
                has_gender_columns = any('Gender' in col for col in uploaded_data[source_name].columns)
                if has_gender_columns:
                    breakdown_lines.append(f"- 🎭 Gender: {gender_count} issues")

                breakdown_lines.append(f"- 🌍 Race/Ethnicity: {race_count} issues")

                st.markdown("**Issue Breakdown:**\n" + "\n".join(breakdown_lines))

                st.write("---")

                # Age Range Issues
                if age_issues:
                    st.write("### 🔢 Age Range Issues")
                    for key, df in age_issues.items():
                        if not df.empty:
                            column_name = df['Column'].iloc[0]
                            st.write(f"**{column_name}** ({len(df)} invalid entries)")

                            # Show detailed table
                            display_df = df[['Row', 'Value', 'Valid_Options']].copy()
                            display_df['Row'] = display_df['Row'].apply(lambda x: f"Row {x}")
                            st.dataframe(display_df, width='stretch', height=min(300, len(df) * 35 + 50))

                            # Download option for this specific issue
                            csv = df.to_csv(index=False)
                            st.download_button(
                                f"📥 Download CSV",
                                data=csv,
                                file_name=f"{source_name}_{column_name.replace(' ', '_')}_issues.csv",
                                mime="text/csv",
                                key=f"dl_{source_name}_{key}"
                            )
                            st.write("---")

                # Sex Issues
                if sex_issues:
                    st.write("### 👤 Sex Issues (Required Field)")
                    for key, df in sex_issues.items():
                        if not df.empty:
                            column_name = df['Column'].iloc[0]
                            st.write(f"**{column_name}** ({len(df)} invalid entries)")

                            # Show detailed table
                            display_df = df[['Row', 'Value', 'Invalid_Parts', 'Valid_Options']].copy()
                            display_df['Row'] = display_df['Row'].apply(lambda x: f"Row {x}")
                            st.dataframe(display_df, width='stretch', height=min(300, len(df) * 35 + 50))

                            # Download option
                            csv = df.to_csv(index=False)
                            st.download_button(
                                f"📥 Download CSV",
                                data=csv,
                                file_name=f"{source_name}_{column_name.replace(' ', '_')}_issues.csv",
                                mime="text/csv",
                                key=f"dl_{source_name}_{key}"
                            )
                            st.write("---")

                # Gender Issues
                if gender_issues:
                    st.write("### 🎭 Gender Issues (Optional Field)")
                    for key, df in gender_issues.items():
                        if not df.empty:
                            column_name = df['Column'].iloc[0]
                            st.write(f"**{column_name}** ({len(df)} invalid entries)")

                            # Show detailed table
                            display_df = df[['Row', 'Value', 'Invalid_Parts', 'Valid_Options']].copy()
                            display_df['Row'] = display_df['Row'].apply(lambda x: f"Row {x}")
                            st.dataframe(display_df, width='stretch', height=min(300, len(df) * 35 + 50))

                            # Download option
                            csv = df.to_csv(index=False)
                            st.download_button(
                                f"📥 Download CSV",
                                data=csv,
                                file_name=f"{source_name}_{column_name.replace(' ', '_')}_issues.csv",
                                mime="text/csv",
                                key=f"dl_{source_name}_{key}"
                            )
                            st.write("---")
                
                # Race/Ethnicity Issues
                if race_issues:
                    st.write("### 🌍 Race/Ethnicity Issues")
                    for key, df in race_issues.items():
                        if not df.empty:
                            column_name = df['Column'].iloc[0]
                            st.write(f"**{column_name}** ({len(df)} invalid entries)")

                            # Show detailed table
                            display_df = df[['Row', 'Value', 'Invalid_Parts', 'Valid_Options']].copy()
                            display_df['Row'] = display_df['Row'].apply(lambda x: f"Row {x}")
                            st.dataframe(display_df, width='stretch', height=min(300, len(df) * 35 + 50))

                            # Download option
                            csv = df.to_csv(index=False)
                            st.download_button(
                                f"📥 Download CSV",
                                data=csv,
                                file_name=f"{source_name}_{column_name.replace(' ', '_')}_issues.csv",
                                mime="text/csv",
                                key=f"dl_{source_name}_{key}"
                            )
                            st.write("---")

                # Download all issues for this source
                if source_issues > 0:
                    st.write("### 📦 Download All Issues for This Source")
                    all_issues_df = pd.concat(validation_results.values(), ignore_index=True)
                    # Sort by Row for easier navigation
                    all_issues_df = all_issues_df.sort_values('Row')
                    csv_all = all_issues_df.to_csv(index=False)
                    st.download_button(
                        f"📥 Download ALL {source_name} Issues (CSV)",
                        data=csv_all,
                        file_name=f"{source_name}_all_validation_issues.csv",
                        mime="text/csv",
                        key=f"dl_all_{source_name}",
                        type="primary",
                        width='stretch'
                    )

def show_report_filters():
    """Show filter interface for reports"""
    uploaded_data = st.session_state.get('uploaded_data', {})

    if not uploaded_data:
        return

    # Collect all unique values from all data sources for filter columns
    filter_columns = ['Project Name on HIC', 'County', 'AHS District', 'Location: General']
    available_filters = {}

    for col in filter_columns:
        all_values = set()
        col_exists = False

        for source_name, df in uploaded_data.items():
            if col in df.columns:
                col_exists = True
                # Get unique non-null values
                unique_vals = df[col].dropna().unique()
                all_values.update(unique_vals)

        if col_exists and all_values:
            available_filters[col] = sorted(list(all_values))

    if not available_filters:
        return

    # Show filter UI
    with st.expander("🔍 Filter Reports", expanded=False):
        st.info("Filter the data before viewing reports. Select values to include in the reports.")

        # Initialize filter state if not exists
        if 'report_filters' not in st.session_state:
            st.session_state['report_filters'] = {}

        # Create columns for filters
        num_filters = len(available_filters)
        if num_filters == 1:
            cols = [st.container()]
        elif num_filters == 2:
            cols = st.columns(2)
        else:
            cols = st.columns(3)

        filter_changed = False

        for idx, (col_name, values) in enumerate(available_filters.items()):
            with cols[idx % len(cols)]:
                st.write(f"**{col_name}**")

                # Multiselect for each filter column
                selected = st.multiselect(
                    f"Select {col_name}",
                    options=values,
                    default=st.session_state['report_filters'].get(col_name, []),
                    key=f"filter_{col_name}",
                    label_visibility="collapsed"
                )

                # Check if filter changed
                if selected != st.session_state['report_filters'].get(col_name, []):
                    filter_changed = True

                st.session_state['report_filters'][col_name] = selected

        # Apply filters button
        col1, col2, col3 = st.columns([1, 1, 1])

        with col1:
            if st.button("🔄 Apply Filters", type="primary", width='stretch'):
                apply_report_filters()

        with col2:
            if st.button("🔁 Reset Filters", width='stretch'):
                st.session_state['report_filters'] = {}
                apply_report_filters()

        with col3:
            # Show active filters count
            active_count = sum(1 for v in st.session_state['report_filters'].values() if v)
            if active_count > 0:
                st.info(f"**{active_count}** filter(s) active")

def apply_report_filters():
    """Apply filters and regenerate reports"""
    from reports import generate_all_reports

    filters = st.session_state.get('report_filters', {})
    uploaded_data = st.session_state.get('uploaded_data', {})

    # Check if any filters are active
    has_active_filters = any(v for v in filters.values())

    if not has_active_filters:
        # No filters, regenerate from original data
        with st.spinner("Regenerating reports without filters..."):
            processed_data = st.session_state.get('processed_data', {})
            calculated_reports = generate_all_reports(processed_data)
            st.session_state['calculated_reports'] = calculated_reports
            st.success("✅ Filters removed, reports regenerated!")
            st.rerun()
        return

    # Apply filters to uploaded data
    filtered_data = {}

    with st.spinner("Applying filters and regenerating reports..."):
        for source_name, df in uploaded_data.items():
            filtered_df = df.copy()

            # Apply each filter
            for col_name, selected_values in filters.items():
                if selected_values and col_name in filtered_df.columns:
                    filtered_df = filtered_df[filtered_df[col_name].isin(selected_values)]

            # Only include if data remains after filtering
            if not filtered_df.empty:
                filtered_data[source_name] = filtered_df

        if not filtered_data:
            st.error("No data matches the selected filters. Please adjust your filter criteria.")
            return

        # Process filtered data
        from processor import process_pit_data
        region = st.session_state.get('region')

        processed_filtered = {}
        for source_name, df in filtered_data.items():
            source_data = process_pit_data(df, source_name, region)
            processed_filtered[source_name] = source_data

        # Combine data for reporting
        all_persons = []
        all_households = []

        for source_name, source_data in processed_filtered.items():
            persons_df = source_data.get('persons_df')
            households_df = source_data.get('households_df')

            if not persons_df.empty:
                all_persons.append(persons_df)
            if not households_df.empty:
                all_households.append(households_df)

        # Create combined dataset
        if all_persons:
            combined_persons = pd.concat(all_persons, ignore_index=True)
            combined_households = pd.concat(all_households, ignore_index=True)
        else:
            combined_persons = pd.DataFrame()
            combined_households = pd.DataFrame()

        processed_filtered['combined'] = {
            'persons_df': combined_persons,
            'households_df': combined_households
        }

        # Generate reports from filtered data
        calculated_reports = generate_all_reports(processed_filtered)
        st.session_state['calculated_reports'] = calculated_reports

        # Show summary
        total_records = sum(len(df) for df in filtered_data.values())
        st.success(f"✅ Filters applied! Reports regenerated with {total_records} filtered records.")
        st.rerun()

def show_reports_interface():
    """Show the reports interface with all tabs preserved"""
    calculated_reports = st.session_state.get('calculated_reports', {})

    if not calculated_reports:
        st.info("No reports available.")
        return

    # Show filter interface
    show_report_filters()

    st.write("---")

    # Check if 'Project Name on HIC' column exists in ES or TH data
    uploaded_data = st.session_state.get('uploaded_data', {})
    has_project_name_col = False
    for source_name in ['Sheltered_ES', 'Sheltered_TH']:
        if source_name in uploaded_data:
            if 'Project Name on HIC' in uploaded_data[source_name].columns:
                has_project_name_col = True
                break

    # Create tabs for each report type - PRESERVE EXACT STRUCTURE
    report_types = ['HDX_Totals', 'HDX_Veterans', 'HDX_Youth', 'HDX_Subpopulations', 'PIT Summary']
    tab_labels = [f"📊 {rt}" for rt in report_types]

    # Add Project Breakdown tab if column exists
    if has_project_name_col:
        tab_labels.append("📋 Breakdown by Project")

    tabs = st.tabs(tab_labels)
    
    for idx, tab in enumerate(tabs):
        with tab:
            # Check if this is the Project Breakdown tab (last tab when it exists)
            if has_project_name_col and idx == len(tabs) - 1:
                show_project_breakdown_tab(uploaded_data)
            else:
                report_type = report_types[idx]
                reports = calculated_reports.get(report_type, {})

                if not reports:
                    st.info(f"No {report_type} reports available.")
                    continue

                # Display each report in this category
                for report_name, report_df in reports.items():
                    st.subheader(f"📋 {report_name}")

                    if report_df.empty:
                        st.warning("No data available for this report.")
                        continue

                    # Prepare display
                    display_df = report_df.copy()

                    # Handle MultiIndex
                    if isinstance(display_df.index, pd.MultiIndex):
                        display_df = display_df.reset_index()
                        display_df.columns = ['Category', 'Subcategory'] + list(display_df.columns[2:])
                    else:
                        display_df = display_df.reset_index()
                        if 'index' in display_df.columns:
                            display_df.rename(columns={'index': 'Category'}, inplace=True)
                        display_df['Subcategory'] = ''

                    # Show the table
                    from utils import safe_dataframe_display
                    st.dataframe(
                        safe_dataframe_display(display_df),
                        width='stretch',
                        height=min(600, len(display_df) * 35 + 100),
                        hide_index=True
                    )

                    # Download button
                    csv_data = display_df.to_csv(index=False)
                    st.download_button(
                        "📥 Download CSV",
                        data=csv_data,
                        file_name=f"{report_name.replace(' ', '_')}.csv",
                        mime="text/csv",
                        key=f"dl_{report_type}_{report_name}"
                    )

                    st.write("---")

def show_project_breakdown_tab(uploaded_data: Dict):
    """Show breakdown by Project Name on HIC for ES and TH data"""
    st.subheader("Breakdown by Project Name on HIC")

    st.info("This report shows household and client counts grouped by Project Name on HIC for Emergency Shelter (ES) and Transitional Housing (TH) data.")

    processed_data = st.session_state.get('processed_data', {})

    # Process ES and TH data
    for source_name, display_name in [('Sheltered_ES', 'Sheltered_ES Data'), ('Sheltered_TH', 'Sheltered_TH Data')]:
        if source_name not in uploaded_data:
            continue

        raw_df = uploaded_data[source_name]

        if 'Project Name on HIC' not in raw_df.columns:
            continue

        st.write(f"### {display_name}")

        # Get processed data to calculate household types and person counts
        source_processed = processed_data.get(source_name, {})
        persons_df = source_processed.get('persons_df', pd.DataFrame())
        households_df = source_processed.get('households_df', pd.DataFrame())

        if persons_df.empty or households_df.empty:
            st.warning(f"No processed data available for {source_name}.")
            continue

        # Add Project Name on HIC to households_df based on Household_ID
        # The raw data has one row per household, so we can map by index
        raw_df_indexed = raw_df.reset_index(drop=True)
        raw_df_indexed['Household_ID'] = range(1, len(raw_df_indexed) + 1)

        # Create mapping of Household_ID to Project Name on HIC
        project_mapping = raw_df_indexed.set_index('Household_ID')['Project Name on HIC'].to_dict()

        # Add project name to households_df
        households_with_project = households_df.copy()
        households_with_project['Project Name on HIC'] = households_with_project['household_id'].map(project_mapping)

        # Add project name to persons_df
        persons_with_project = persons_df.copy()
        persons_with_project['Project Name on HIC'] = persons_with_project['Household_ID'].map(project_mapping)

        # Calculate breakdown statistics
        breakdown_data = []

        # Group by Project Name on HIC and Household Type
        for project_name in sorted(households_with_project['Project Name on HIC'].dropna().unique()):
            project_households = households_with_project[households_with_project['Project Name on HIC'] == project_name]
            project_persons = persons_with_project[persons_with_project['Project Name on HIC'] == project_name]

            # Get household types in this project
            for hh_type in project_households['household_type'].unique():
                hh_subset = project_households[project_households['household_type'] == hh_type]
                hh_count = len(hh_subset)

                # Count persons in these households
                hh_ids = hh_subset['household_id'].tolist()
                persons_count = project_persons[project_persons['Household_ID'].isin(hh_ids)].shape[0]

                breakdown_data.append({
                    'Project Name on HIC': project_name,
                    'Household Type': hh_type,
                    'Count Households': hh_count,
                    'Number of Clients': persons_count
                })

        if breakdown_data:
            breakdown_df = pd.DataFrame(breakdown_data)

            # Display the breakdown table
            st.dataframe(
                breakdown_df,
                width='stretch',
                height=min(600, len(breakdown_df) * 35 + 100),
                hide_index=True
            )

            # Download button
            csv_data = breakdown_df.to_csv(index=False)
            st.download_button(
                f"📥 Download {source_name} Project Breakdown CSV",
                data=csv_data,
                file_name=f"{source_name}_Project_Breakdown.csv",
                mime="text/csv",
                key=f"dl_project_breakdown_{source_name}"
            )
        else:
            st.warning(f"No project breakdown data available for {source_name}.")

        st.write("---")


# ============================================================================
# DOWNLOAD INTERFACE - TAB-BASED REDESIGN
# ============================================================================

def show_download_interface():
    """Redesigned download interface with tabbed layout for better organization"""
    calculated_reports = st.session_state.get('calculated_reports', {})
    processed_data = st.session_state.get('processed_data', {})
    uploaded_data = st.session_state.get('uploaded_data', {})
    region = st.session_state.get('region', 'Unknown')

    if not calculated_reports:
        st.info("No reports available for download.")
        return

    # Quick summary at top
    _show_download_summary(calculated_reports, processed_data)

    # Create tabs for organized navigation
    tab_reports, tab_data = st.tabs([
        "Reports", "Data Exports"
    ])

    with tab_reports:
        _show_reports_tab(calculated_reports, uploaded_data, processed_data, region)

    with tab_data:
        _show_data_exports_tab(uploaded_data, processed_data, region)


def _show_download_summary(calculated_reports, processed_data):
    """Show quick summary of available data"""
    col1, col2, col3 = st.columns(3)

    with col1:
        total_reports = sum(len(reports) for reports in calculated_reports.values())
        st.metric("Total Reports", total_reports)

    with col2:
        total_persons = 0
        for source_data in processed_data.values():
            if isinstance(source_data, dict):
                persons_df = source_data.get('persons_df', pd.DataFrame())
                if not persons_df.empty:
                    total_persons += len(persons_df)
        st.metric("Total Persons", total_persons)

    with col3:
        sources_with_data = sum(1 for source_data in processed_data.values()
                               if isinstance(source_data, dict) and
                               not source_data.get('persons_df', pd.DataFrame()).empty)
        st.metric("Data Sources", sources_with_data)


def _show_reports_tab(calculated_reports, uploaded_data, processed_data, region):
    """Reports download tab with combined and individual options"""
    st.markdown("#### Combined Reports Download")
    st.caption("Download all HUD reports in a single Excel file, optionally with source data included.")

    # Download type selection
    download_type = st.radio(
        "Select contents",
        ["Reports Only",
         "Reports + Raw Data with IDs",
         "Reports + Processed Data with IDs",
         "All Data (Reports + Raw + Processed)"],
        horizontal=True,
        key="reports_download_type"
    )

    # Generate button and download
    timezone = get_timezone_for_region(region)
    timestamp = get_current_timestamp(timezone)
    filename = create_download_filename(region, "Reports", timestamp)

    if st.button("Generate Download", type="primary", key="generate_reports"):
        generate_download_file(
            filename,
            download_type,
            calculated_reports,
            uploaded_data,
            processed_data,
            region
        )

    st.markdown("---")
    st.markdown("#### Individual Report Downloads")
    st.caption("Download a specific report as CSV or Excel.")

    # Individual downloads using existing function logic
    report_options = []
    for report_type, reports in calculated_reports.items():
        for report_name in reports.keys():
            report_options.append(f"{report_type} - {report_name}")

    if report_options:
        selected_report = st.selectbox(
            "Select report",
            report_options,
            key="individual_report_select"
        )

        if selected_report:
            report_type, report_name = selected_report.split(" - ", 1)
            report_df = calculated_reports[report_type][report_name]

            with st.expander("Preview"):
                st.dataframe(report_df, width='stretch')

            col1, col2 = st.columns(2)
            with col1:
                csv_data = report_df.to_csv()
                st.download_button(
                    label="Download CSV",
                    data=csv_data,
                    file_name=f"{region}_{report_type}_{report_name.replace(' ', '_')}.csv",
                    mime='text/csv',
                    key="individual_csv"
                )
            with col2:
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    report_df.to_excel(writer, sheet_name=report_name[:30], index=True)
                st.download_button(
                    label="Download Excel",
                    data=excel_buffer.getvalue(),
                    file_name=f"{region}_{report_type}_{report_name.replace(' ', '_')}.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key="individual_xlsx"
                )


def _show_data_exports_tab(uploaded_data, processed_data, region):
    """Data exports tab with raw and processed data options"""

    # ============================================================================
    # OBSERVATION DATA EXPORT SECTION
    # ============================================================================
    st.markdown("### Observation Data Export")
    st.caption("Export observation records (consent != 'Yes' AND observation tool completed)")

    # Check if any data available
    has_uploaded_data = any(
        uploaded_data.get(source) is not None and not uploaded_data.get(source).empty
        for source in ['Sheltered_ES', 'Sheltered_TH', 'Unsheltered']
    )

    if not has_uploaded_data:
        st.info("No uploaded data available for observation export.")
    else:
        # Preview button to show counts
        col1, col2 = st.columns([1, 3])

        with col1:
            if st.button("Preview Counts", key="preview_observation_counts"):
                with st.spinner("Analyzing observation data..."):
                    try:
                        # Calculate counts for each source
                        counts = {}
                        total_count = 0

                        for source_name in ['Sheltered_ES', 'Sheltered_TH', 'Unsheltered']:
                            raw_df = uploaded_data.get(source_name)
                            if raw_df is None or raw_df.empty:
                                counts[source_name] = 0
                                continue

                            filtered_df = filter_observation_data(raw_df, source_name)
                            selected_df = select_observation_columns(filtered_df, region)
                            counts[source_name] = len(selected_df)
                            total_count += len(selected_df)

                        # Display counts
                        if total_count > 0:
                            st.success(f"Found {total_count} observation records:")
                            st.write(f"- Emergency Shelter: {counts.get('Sheltered_ES', 0)} records")
                            st.write(f"- Transitional Housing: {counts.get('Sheltered_TH', 0)} records")
                            st.write(f"- Unsheltered: {counts.get('Unsheltered', 0)} records")
                        else:
                            st.warning("No observation records found matching filter criteria.")

                    except Exception as e:
                        st.error(f"Error analyzing observation data: {str(e)}")

        with col2:
            # Generate and download button
            if st.button("Generate Observation Export", type="primary", key="generate_observation_export"):
                with st.spinner("Generating Excel file..."):
                    try:
                        from utils import get_timezone_for_region, get_current_timestamp

                        excel_buffer = generate_observation_data_export(uploaded_data, region)

                        # Create filename
                        timezone = get_timezone_for_region(region)
                        timestamp = get_current_timestamp(timezone)
                        filename = f"{region.replace(' ', '_')}_Observation_Data_{timestamp}.xlsx"

                        # Calculate file size
                        file_size_kb = len(excel_buffer.getvalue()) / 1024

                        st.download_button(
                            label=f"📥 Download Observation Data ({file_size_kb:.1f} KB)",
                            data=excel_buffer.getvalue(),
                            file_name=filename,
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            type="primary",
                            key="download_observation_data"
                        )

                        st.success("Observation data export generated successfully!")

                    except ValueError as ve:
                        st.warning(str(ve))
                    except Exception as e:
                        st.error(f"Error generating observation export: {str(e)}")
                        with st.expander("Error Details"):
                            st.exception(e)

    st.markdown("---")

    # ============================================================================
    # EXISTING CODE CONTINUES HERE
    # ============================================================================

    # Source selection
    available_sources = []
    for source_name in ['Sheltered_ES', 'Sheltered_TH', 'Unsheltered']:
        if uploaded_data.get(source_name) is not None or processed_data.get(source_name):
            available_sources.append(source_name)

    if not available_sources:
        st.info("No data available for export.")
        return

    selected_source = st.selectbox(
        "Select data source",
        available_sources,
        key="data_source_select"
    )

    st.markdown("---")

    # Create sub-tabs for Raw vs Processed
    raw_tab, processed_tab = st.tabs(["Raw Data", "Processed Data"])

    with raw_tab:
        st.markdown("#### Raw Data with IDs")
        st.caption("Original uploaded data with Household_ID and Person_IDs added for traceability.")

        raw_data_with_ids = prepare_raw_data_with_ids(uploaded_data, processed_data)
        if selected_source in raw_data_with_ids:
            enhanced_df = raw_data_with_ids[selected_source]

            col1, col2 = st.columns(2)
            with col1:
                st.metric("Households", len(enhanced_df))
            with col2:
                total_persons = sum(len(pid.split(', ')) for pid in enhanced_df['Person_IDs'])
                st.metric("Total Persons", total_persons)

            with st.expander("Preview first 10 rows"):
                preview_cols = ['Household_ID', 'Person_IDs', 'Timestamp', 'Gender', 'Age Range', 'Race/Ethnicity']
                display_cols = [col for col in preview_cols if col in enhanced_df.columns]
                st.dataframe(safe_dataframe_display(enhanced_df[display_cols].head(10)), width='stretch')

            csv_data = enhanced_df.to_csv(index=False)
            st.download_button(
                label="Download Raw Data CSV",
                data=csv_data,
                file_name=f"{region}_{selected_source}_Raw_with_IDs.csv",
                mime='text/csv',
                key=f"raw_download_{selected_source}"
            )
        else:
            st.info("No raw data available for this source.")

    with processed_tab:
        st.markdown("#### Processed Data")
        st.caption("Fully transformed data with all calculated fields (Person_ID, CH_Reason, etc.).")

        processed_data_with_ids = prepare_processed_data_with_ids(processed_data)
        if selected_source in processed_data_with_ids:
            data_dict = processed_data_with_ids[selected_source]

            # Persons section
            st.markdown("**Persons Data**")
            if 'persons' in data_dict and not data_dict['persons'].empty:
                persons_df = data_dict['persons']

                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Total Persons", len(persons_df))
                with col2:
                    st.metric("Unique Households", persons_df['Household_ID'].nunique())

                with st.expander("Preview first 20 rows"):
                    st.dataframe(persons_df.head(20), width='stretch')

                csv_data = persons_df.to_csv(index=False)
                st.download_button(
                    label="Download Persons CSV",
                    data=csv_data,
                    file_name=f"{region}_{selected_source}_Processed_Persons.csv",
                    mime='text/csv',
                    key=f"processed_persons_{selected_source}"
                )
            else:
                st.info("No persons data available.")

            st.markdown("---")

            # Households section
            st.markdown("**Households Data**")
            if 'households' in data_dict and not data_dict['households'].empty:
                households_df = data_dict['households']

                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Total Households", len(households_df))
                with col2:
                    if 'total_persons' in households_df.columns:
                        avg_size = households_df['total_persons'].mean()
                        st.metric("Avg Household Size", f"{avg_size:.1f}")

                with st.expander("Preview first 20 rows"):
                    st.dataframe(households_df.head(20), width='stretch')

                csv_data = households_df.to_csv(index=False)
                st.download_button(
                    label="Download Households CSV",
                    data=csv_data,
                    file_name=f"{region}_{selected_source}_Processed_Households.csv",
                    mime='text/csv',
                    key=f"processed_households_{selected_source}"
                )
            else:
                st.info("No households data available.")
        else:
            st.info("No processed data available for this source.")


def generate_download_file(filename, download_type, calculated_reports, uploaded_data, processed_data, region):
    """Generate download file with selected options"""
    with st.spinner("Generating Excel file..."):
        try:
            # Prepare data based on download type
            include_raw = download_type in ["Reports + Raw Data with IDs", "All Data (Reports + Raw + Processed)"]
            include_processed = download_type in ["Reports + Processed Data with IDs", "All Data (Reports + Raw + Processed)"]
            
            raw_data_with_ids = None
            processed_data_with_ids = None
            
            if include_raw:
                raw_data_with_ids = prepare_raw_data_with_ids(uploaded_data, processed_data)
            
            if include_processed:
                processed_data_with_ids = prepare_processed_data_with_ids(processed_data)
            
            # Create Excel file
            excel_buffer = create_comprehensive_excel_export(
                calculated_reports,
                raw_data_with_ids,
                processed_data_with_ids,
                include_raw,
                include_processed
            )
            
            # Calculate file size
            file_size_mb = len(excel_buffer.getvalue()) / 1024 / 1024
            
            # Provide download
            st.download_button(
                label=f"📥 Download Excel File ({file_size_mb:.1f} MB)",
                data=excel_buffer.getvalue(),
                file_name=filename,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                type="primary",
                width='stretch'
            )
            
            st.success("Excel file generated successfully!")
            
            # Show file details
            with st.expander("File Details"):
                st.write(f"**Filename:** {filename}")
                st.write(f"**File Size:** {file_size_mb:.2f} MB")
                st.write(f"**Report Sheets:** {len(calculated_reports)}")
                
                for sheet_name, sheet_data in calculated_reports.items():
                    st.write(f"- {sheet_name}: {len(sheet_data)} reports")
                
                if include_raw and raw_data_with_ids:
                    st.write(f"**Raw Data Sheets:** {len(raw_data_with_ids)}")
                    for source_name, raw_df in raw_data_with_ids.items():
                        st.write(f"- Raw_{source_name}: {len(raw_df)} households")
                
                if include_processed and processed_data_with_ids:
                    st.write(f"**Processed Data Sheets:** {len(processed_data_with_ids) * 2}")
                    for source_name, data_dict in processed_data_with_ids.items():
                        if 'persons' in data_dict:
                            st.write(f"- Processed_Persons_{source_name}: {len(data_dict['persons'])} persons")
                        if 'households' in data_dict:
                            st.write(f"- Processed_Households_{source_name}: {len(data_dict['households'])} households")
        
        except Exception as e:
            st.error(f"Error generating Excel file: {str(e)}")

def prepare_raw_data_with_ids(uploaded_data, processed_data):
    """Prepare raw data with Household ID, Excel Row Number, and Person IDs for traceability"""
    raw_data_with_ids = {}

    for source_name in ['Sheltered_ES', 'Sheltered_TH', 'Unsheltered']:
        raw_df = uploaded_data.get(source_name, pd.DataFrame())
        if raw_df.empty:
            continue

        source_data = processed_data.get(source_name, {})
        persons_df = source_data.get('persons_df', pd.DataFrame())

        if persons_df.empty:
            continue

        # Create enhanced raw data
        enhanced_raw = raw_df.copy()

        # Add Household ID
        enhanced_raw.insert(0, 'Household_ID', range(1, len(enhanced_raw) + 1))

        # Add Excel Row Number for traceability (Excel row = index + 2 for header)
        enhanced_raw.insert(1, 'Excel_Row', range(2, len(enhanced_raw) + 2))

        # Create Person IDs list for each household
        person_ids_list = []

        for idx in range(len(enhanced_raw)):
            household_id = idx + 1

            # Get all persons in this household
            household_persons = persons_df[persons_df['Household_ID'] == household_id]

            if not household_persons.empty:
                # Use existing Person_ID if available
                if 'Person_ID' in household_persons.columns:
                    person_ids = household_persons['Person_ID'].tolist()
                    person_ids_str = ', '.join(sorted([str(p) for p in person_ids]))
                else:
                    person_ids = []
                    for _, person in household_persons.iterrows():
                        member_type = person.get('Member_Type', 'Unknown')
                        member_number = person.get('Member_Number', 1)
                        person_id = f"HH{household_id}_{member_type[0]}{member_number}"
                        person_ids.append(person_id)
                    person_ids_str = ', '.join(sorted(person_ids))
            else:
                person_ids_str = f"HH{household_id}_A1"

            person_ids_list.append(person_ids_str)

        # Add Person IDs column
        enhanced_raw.insert(2, 'Person_IDs', person_ids_list)

        raw_data_with_ids[source_name] = enhanced_raw

    return raw_data_with_ids

def prepare_processed_data_with_ids(processed_data):
    """Prepare processed persons and households data with traceability columns"""
    processed_data_with_ids = {}

    for source_name in ['Sheltered_ES', 'Sheltered_TH', 'Unsheltered']:
        source_data = processed_data.get(source_name, {})
        persons_df = source_data.get('persons_df', pd.DataFrame())
        households_df = source_data.get('households_df', pd.DataFrame())

        if persons_df.empty and households_df.empty:
            continue

        # Prepare persons data
        if not persons_df.empty:
            persons_enhanced = persons_df.copy()

            # Use existing Person_ID if available, otherwise generate one
            if 'Person_ID' not in persons_enhanced.columns:
                person_ids = []
                for _, person in persons_enhanced.iterrows():
                    household_id = person.get('Household_ID', 0)
                    member_type = person.get('Member_Type', 'Unknown')
                    member_number = person.get('Member_Number', 1)
                    person_id = f"HH{household_id}_{member_type[0]}{member_number}"
                    person_ids.append(person_id)
                persons_enhanced.insert(0, 'Person_ID', person_ids)

            # Reorder columns with traceability columns first
            important_cols = ['Person_ID', 'Source_Row_Number', 'Household_ID', 'Member_Type', 'Member_Number',
                            'Sex', 'Gender', 'race', 'age_range', 'age_group',
                            'household_type', 'DV', 'vet', 'CH', 'CH_Reason', 'disability']
            available_important = [col for col in important_cols if col in persons_enhanced.columns]
            other_cols = [col for col in persons_enhanced.columns if col not in available_important]
            persons_enhanced = persons_enhanced[available_important + other_cols]
        else:
            persons_enhanced = pd.DataFrame()

        # Prepare households data
        if not households_df.empty:
            households_enhanced = households_df.copy()

            # Add person count and IDs using existing Person_ID from persons_df
            person_counts = []
            person_ids_lists = []

            for _, household in households_enhanced.iterrows():
                household_id = household.get('household_id', 0)

                # Get all persons in this household
                household_persons = persons_df[persons_df['Household_ID'] == household_id]

                person_counts.append(len(household_persons))

                if not household_persons.empty:
                    # Use existing Person_ID if available
                    if 'Person_ID' in household_persons.columns:
                        person_ids = household_persons['Person_ID'].tolist()
                    else:
                        person_ids = []
                        for _, person in household_persons.iterrows():
                            member_type = person.get('Member_Type', 'Unknown')
                            member_number = person.get('Member_Number', 1)
                            person_id = f"HH{household_id}_{member_type[0]}{member_number}"
                            person_ids.append(person_id)
                    person_ids_lists.append(', '.join(sorted([str(p) for p in person_ids])))
                else:
                    person_ids_lists.append('')

            households_enhanced.insert(1, 'Person_Count', person_counts)
            households_enhanced.insert(2, 'Person_IDs', person_ids_lists)

            # Reorder columns
            important_cols = ['household_id', 'Person_Count', 'Person_IDs', 'household_type',
                            'total_persons', 'count_adult', 'count_youth', 'count_child_hoh',
                            'count_child_hh', 'youth']
            available_important = [col for col in important_cols if col in households_enhanced.columns]
            other_cols = [col for col in households_enhanced.columns if col not in available_important]
            households_enhanced = households_enhanced[available_important + other_cols]
        else:
            households_enhanced = pd.DataFrame()

        processed_data_with_ids[source_name] = {
            'persons': persons_enhanced,
            'households': households_enhanced
        }

    return processed_data_with_ids

# ============================================================================
# OBSERVATION DATA EXPORT FUNCTIONS
# ============================================================================

def filter_observation_data(raw_df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    """
    Filter raw data for observation records based on consent and observation criteria.

    Criteria:
    - "Does the individual consent to survey?" != "Yes"
    - AND "Are you able to complete an observation tool..." == "Yes"

    Args:
        raw_df: Raw uploaded DataFrame
        source_name: Source identifier (for logging)

    Returns:
        Filtered DataFrame containing only observation records
    """
    from config import OBSERVATION_FILTER_COLUMNS

    # Check column presence
    consent_col = OBSERVATION_FILTER_COLUMNS['consent']
    observation_col = OBSERVATION_FILTER_COLUMNS['observation']

    has_consent = consent_col in raw_df.columns
    has_observation = observation_col in raw_df.columns

    if not has_consent and not has_observation:
        st.warning(f"Source '{source_name}': No consent/observation columns found. No observation data to export.")
        return pd.DataFrame()

    # Build filter mask
    mask = pd.Series([True] * len(raw_df), index=raw_df.index)

    if has_consent:
        # Consent != "Yes" (case insensitive, handle NaN)
        consent_series = raw_df[consent_col].fillna('').astype(str).str.strip().str.lower()
        mask = mask & (consent_series != 'yes')

    if has_observation:
        # Observation == "Yes" (case insensitive, handle NaN)
        obs_series = raw_df[observation_col].fillna('').astype(str).str.strip().str.lower()
        mask = mask & (obs_series == 'yes')

    filtered_df = raw_df[mask].copy()

    return filtered_df


def select_observation_columns(filtered_df: pd.DataFrame, region: str) -> pd.DataFrame:
    """
    Select only observation-relevant columns based on region.

    Args:
        filtered_df: Filtered observation data
        region: Region identifier ('New England' or 'Great Lakes')

    Returns:
        DataFrame with only observation-relevant columns present in source data
    """
    from config import OBSERVATION_COLUMNS_BY_REGION

    # Get expected columns for region
    expected_cols = OBSERVATION_COLUMNS_BY_REGION.get(region, [])

    if not expected_cols:
        st.warning(f"No observation column configuration for region: {region}")
        return filtered_df

    # Find columns that exist in the DataFrame
    available_cols = [col for col in expected_cols if col in filtered_df.columns]

    if not available_cols:
        st.info(f"None of the expected observation columns were found in the data.")
        return pd.DataFrame()

    # Select only available columns
    result_df = filtered_df[available_cols].copy()

    # Log missing columns for debugging (optional, can be expanded by user)
    missing_cols = [col for col in expected_cols if col not in filtered_df.columns]
    if missing_cols and len(missing_cols) < 10:  # Only show if not too many missing
        with st.expander("ℹ️ Missing observation columns (optional)", expanded=False):
            st.caption(f"The following {len(missing_cols)} columns were not found in the uploaded data:")
            for col in missing_cols:
                st.text(f"• {col}")

    return result_df


# ============================================================================
# OBSERVATION SUMMARY CALCULATION FUNCTIONS
# ============================================================================

def calculate_observation_stats(df: pd.DataFrame, region: str) -> Dict[str, Any]:
    """
    Calculate summary statistics from observation data.

    Args:
        df: Filtered observation DataFrame
        region: Region identifier for region-specific calculations

    Returns:
        Dict of calculated statistics
    """
    from config import (
        OBSERVATION_PERSON_COLUMNS,
        OBSERVATION_VALID_AGE_RANGES,
        OBSERVATION_VALID_SEX,
        OBSERVATION_VALID_GENDERS,
        OBSERVATION_RACE_CATEGORIES,
        OBSERVATION_LOCATION_COLUMNS
    )

    stats = {}

    # Overview stats
    stats['total_observations'] = len(df)
    stats['total_persons'] = _count_total_persons(df)
    stats['total_adults'] = _safe_numeric_sum(df, 'Number of adults')
    stats['total_children'] = _safe_numeric_sum(df, 'Number of children')
    stats['total_unknown_age'] = _safe_numeric_sum(df, 'Number of persons of unknown age (not sure if adult or child)')

    # Homeless status
    homeless_col = 'Is this person/family homeless?'
    if homeless_col in df.columns:
        homeless_series = df[homeless_col].fillna('').astype(str).str.strip().str.lower()
        stats['homeless_yes'] = (homeless_series == 'yes').sum()
        stats['homeless_other'] = len(df) - stats['homeless_yes']
    else:
        stats['homeless_yes'] = 0
        stats['homeless_other'] = len(df)

    # Average household size
    hh_size_col = 'Total persons staying together as a household'
    if hh_size_col in df.columns:
        numeric_sizes = pd.to_numeric(df[hh_size_col], errors='coerce')
        stats['avg_household_size'] = round(numeric_sizes.mean(), 2) if not numeric_sizes.isna().all() else 0
    else:
        stats['avg_household_size'] = 0

    # Demographics
    stats.update(_calculate_obs_age_distribution(df))
    stats.update(_calculate_obs_sex_distribution(df))

    if region == 'New England':
        stats.update(_calculate_obs_gender_distribution(df))

    stats.update(_calculate_obs_race_distribution(df))

    # Household analysis
    stats.update(_calculate_obs_household_distribution(df))

    return stats


def _safe_numeric_sum(df: pd.DataFrame, column: str) -> int:
    """Safely sum a numeric column, handling missing columns and non-numeric values."""
    if column not in df.columns:
        return 0
    numeric_values = pd.to_numeric(df[column], errors='coerce')
    return int(numeric_values.sum()) if not numeric_values.isna().all() else 0


def _count_total_persons(df: pd.DataFrame) -> int:
    """Count total persons across all person slots in observation data."""
    total = 0

    # Check each person slot for data (person has data if age range or sex is present)
    for i in range(1, 6):
        age_col = f'Person #{i}: Age Range'
        sex_col = f'Person #{i}: Sex'

        if age_col in df.columns:
            total += df[age_col].notna().sum()
            # Exclude empty strings
            total -= (df[age_col].fillna('').astype(str).str.strip() == '').sum()
        elif sex_col in df.columns:
            total += df[sex_col].notna().sum()
            total -= (df[sex_col].fillna('').astype(str).str.strip() == '').sum()

    return max(0, total)


def _calculate_obs_age_distribution(df: pd.DataFrame) -> Dict[str, int]:
    """Calculate age range distribution across all persons."""
    from config import OBSERVATION_VALID_AGE_RANGES

    age_columns = [f'Person #{i}: Age Range' for i in range(1, 6)]
    age_counts = {}

    for age in OBSERVATION_VALID_AGE_RANGES:
        count = 0
        for col in age_columns:
            if col in df.columns:
                col_values = df[col].fillna('').astype(str).str.strip()
                count += (col_values == age).sum()
        age_counts[f'age_{age}'] = count

    # Unknown/Not Reported - count persons with empty or null age
    unknown_count = 0
    for col in age_columns:
        if col in df.columns:
            col_values = df[col].fillna('').astype(str).str.strip()
            # Only count as unknown if person exists (has some data) but age is missing
            # We check if any other person column has data for that slot
            person_num = col.split('#')[1].split(':')[0]
            sex_col = f'Person #{person_num}: Sex'
            if sex_col in df.columns:
                has_person = df[sex_col].notna() & (df[sex_col].fillna('').astype(str).str.strip() != '')
                age_missing = (col_values == '') | df[col].isna()
                unknown_count += (has_person & age_missing).sum()

    age_counts['age_unknown'] = unknown_count

    return age_counts


def _calculate_obs_sex_distribution(df: pd.DataFrame) -> Dict[str, int]:
    """Calculate sex distribution across all persons."""
    from config import OBSERVATION_VALID_SEX

    sex_columns = [f'Person #{i}: Sex' for i in range(1, 6)]
    sex_counts = {}

    for sex in OBSERVATION_VALID_SEX:
        count = 0
        for col in sex_columns:
            if col in df.columns:
                col_values = df[col].fillna('').astype(str).str.strip()
                count += (col_values == sex).sum()
        sex_counts[f'sex_{sex.lower()}'] = count

    # Unknown/Not Reported
    unknown_count = 0
    for col in sex_columns:
        if col in df.columns:
            col_values = df[col].fillna('').astype(str).str.strip()
            # Count if column exists but value is empty/null and person has age data
            person_num = col.split('#')[1].split(':')[0]
            age_col = f'Person #{person_num}: Age Range'
            if age_col in df.columns:
                has_person = df[age_col].notna() & (df[age_col].fillna('').astype(str).str.strip() != '')
                sex_missing = (col_values == '') | df[col].isna()
                unknown_count += (has_person & sex_missing).sum()

    sex_counts['sex_unknown'] = unknown_count

    return sex_counts


def _calculate_obs_gender_distribution(df: pd.DataFrame) -> Dict[str, int]:
    """Calculate gender distribution across all persons (New England only)."""
    gender_columns = [f'Person #{i}: Gender' for i in range(1, 6)]

    # Gender categories mapping (simplify long names)
    gender_mapping = {
        'Woman (Girl if child)': 'gender_woman',
        'Man (Boy if child)': 'gender_man',
        'Culturally Specific Identity (e.g., Two-Spirit)': 'gender_culturally_specific',
        'Transgender': 'gender_transgender',
        'Non-Binary': 'gender_nonbinary',
        'Questioning': 'gender_questioning',
        'Different Identity': 'gender_different'
    }

    gender_counts = {v: 0 for v in gender_mapping.values()}
    gender_counts['gender_unknown'] = 0

    for col in gender_columns:
        if col not in df.columns:
            continue

        col_values = df[col].fillna('').astype(str).str.strip()

        for gender_value, key in gender_mapping.items():
            gender_counts[key] += (col_values == gender_value).sum()

        # Unknown - check if person exists but gender missing
        person_num = col.split('#')[1].split(':')[0]
        age_col = f'Person #{person_num}: Age Range'
        if age_col in df.columns:
            has_person = df[age_col].notna() & (df[age_col].fillna('').astype(str).str.strip() != '')
            gender_missing = (col_values == '') | df[col].isna()
            gender_counts['gender_unknown'] += (has_person & gender_missing).sum()

    return gender_counts


def _calculate_obs_race_distribution(df: pd.DataFrame) -> Dict[str, int]:
    """Calculate race/ethnicity distribution across all persons."""
    race_columns = [f'Person #{i}: Race/Ethnicity' for i in range(1, 6)]

    # Race categories (partial match for flexibility)
    race_patterns = {
        'race_white': ['white'],
        'race_black': ['black', 'african american', 'african'],
        'race_asian': ['asian'],
        'race_indigenous': ['american indian', 'alaska native', 'indigenous'],
        'race_pacific': ['native hawaiian', 'pacific islander'],
        'race_middle_eastern': ['middle eastern', 'north african'],
        'race_hispanic': ['hispanic', 'latina', 'latino', 'latine'],
        'race_multiracial': ['multi', 'two or more', 'multiple']
    }

    race_counts = {k: 0 for k in race_patterns.keys()}
    race_counts['race_unknown'] = 0

    for col in race_columns:
        if col not in df.columns:
            continue

        col_values = df[col].fillna('').astype(str).str.strip().str.lower()

        for key, patterns in race_patterns.items():
            for pattern in patterns:
                race_counts[key] += col_values.str.contains(pattern, case=False, na=False).sum()

        # Unknown - person exists but race missing
        person_num = col.split('#')[1].split(':')[0]
        age_col = f'Person #{person_num}: Age Range'
        if age_col in df.columns:
            has_person = df[age_col].notna() & (df[age_col].fillna('').astype(str).str.strip() != '')
            race_missing = (df[col].fillna('').astype(str).str.strip() == '') | df[col].isna()
            race_counts['race_unknown'] += (has_person & race_missing).sum()

    return race_counts


def _calculate_obs_household_distribution(df: pd.DataFrame) -> Dict[str, int]:
    """Calculate household size and composition distribution."""
    hh_counts = {}

    # Household size distribution
    hh_size_col = 'Total persons staying together as a household'
    if hh_size_col in df.columns:
        sizes = pd.to_numeric(df[hh_size_col], errors='coerce')
        hh_counts['hh_size_1'] = (sizes == 1).sum()
        hh_counts['hh_size_2'] = (sizes == 2).sum()
        hh_counts['hh_size_3'] = (sizes == 3).sum()
        hh_counts['hh_size_4'] = (sizes == 4).sum()
        hh_counts['hh_size_5plus'] = (sizes >= 5).sum()
    else:
        for i in range(1, 5):
            hh_counts[f'hh_size_{i}'] = 0
        hh_counts['hh_size_5plus'] = 0

    # Household composition
    adults_col = 'Number of adults'
    children_col = 'Number of children'

    if adults_col in df.columns and children_col in df.columns:
        adults = pd.to_numeric(df[adults_col], errors='coerce').fillna(0)
        children = pd.to_numeric(df[children_col], errors='coerce').fillna(0)

        hh_counts['hh_comp_adults_only'] = ((adults > 0) & (children == 0)).sum()
        hh_counts['hh_comp_with_children'] = ((adults > 0) & (children > 0)).sum()
        hh_counts['hh_comp_children_only'] = ((adults == 0) & (children > 0)).sum()
        hh_counts['hh_comp_unknown'] = ((adults == 0) & (children == 0)).sum()
    else:
        hh_counts['hh_comp_adults_only'] = 0
        hh_counts['hh_comp_with_children'] = 0
        hh_counts['hh_comp_children_only'] = 0
        hh_counts['hh_comp_unknown'] = len(df)

    return hh_counts


def _calculate_obs_location_stats(observation_data: Dict[str, pd.DataFrame], region: str) -> Dict[str, Dict[str, Dict[str, int]]]:
    """
    Calculate location-based statistics for all sources.

    Returns:
        Dict with structure: {location_type: {location_value: {source: count}}}
    """
    from config import OBSERVATION_LOCATION_COLUMNS

    location_config = OBSERVATION_LOCATION_COLUMNS.get(region, {})
    location_stats = {
        'by_county': {},
        'by_project': {},
        'by_welfare_office': {}
    }

    for source_name, df in observation_data.items():
        # County/Location
        county_col = location_config.get('county')
        if county_col and county_col in df.columns:
            for value in df[county_col].dropna().unique():
                value_str = str(value).strip()
                if value_str:
                    if value_str not in location_stats['by_county']:
                        location_stats['by_county'][value_str] = {'Sheltered_ES': 0, 'Sheltered_TH': 0, 'Unsheltered': 0}
                    location_stats['by_county'][value_str][source_name] = (df[county_col].fillna('').astype(str).str.strip() == value_str).sum()

        # Project
        project_col = location_config.get('project')
        if project_col and project_col in df.columns:
            for value in df[project_col].dropna().unique():
                value_str = str(value).strip()
                if value_str:
                    if value_str not in location_stats['by_project']:
                        location_stats['by_project'][value_str] = {'Sheltered_ES': 0, 'Sheltered_TH': 0, 'Unsheltered': 0}
                    location_stats['by_project'][value_str][source_name] = (df[project_col].fillna('').astype(str).str.strip() == value_str).sum()

        # Welfare Office (New England only)
        welfare_col = location_config.get('welfare_office')
        if welfare_col and welfare_col in df.columns:
            for value in df[welfare_col].dropna().unique():
                value_str = str(value).strip()
                if value_str:
                    if value_str not in location_stats['by_welfare_office']:
                        location_stats['by_welfare_office'][value_str] = {'Sheltered_ES': 0, 'Sheltered_TH': 0, 'Unsheltered': 0}
                    location_stats['by_welfare_office'][value_str][source_name] = (df[welfare_col].fillna('').astype(str).str.strip() == value_str).sum()

    return location_stats


# ============================================================================
# OBSERVATION SUMMARY SHEET CREATION FUNCTIONS
# ============================================================================

def _apply_obs_header_style(cell, font_module, fill_module, alignment_module):
    """Apply consistent header styling for observation summary sheets."""
    cell.font = font_module(bold=True, color="FFFFFF")
    cell.fill = fill_module(start_color="00629b", end_color="00629b", fill_type="solid")
    cell.alignment = alignment_module(horizontal="center", vertical="center", wrap_text=True)


def _apply_obs_category_style(cell, font_module, fill_module, alignment_module):
    """Apply category row styling (section headers within data)."""
    cell.font = font_module(bold=True)
    cell.fill = fill_module(start_color="e2efe8", end_color="e2efe8", fill_type="solid")
    cell.alignment = alignment_module(horizontal="left", vertical="center")


def _adjust_obs_column_widths(worksheet, get_column_letter_func):
    """Auto-adjust column widths for better readability."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 12), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _create_obs_summary_overview_sheet(workbook, source_stats: Dict, region: str):
    """Create the Summary Overview sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    worksheet = workbook.create_sheet("Summary Overview", 0)

    # Define rows with (label, stat_key, is_average)
    rows = [
        ("Total Observations", "total_observations", False),
        ("Total Persons Counted", "total_persons", False),
        ("Total Adults", "total_adults", False),
        ("Total Children", "total_children", False),
        ("Persons Unknown Age", "total_unknown_age", False),
        ("Confirmed Homeless", "homeless_yes", False),
        ("Homeless Not Determined", "homeless_other", False),
        ("Average Household Size", "avg_household_size", True),
    ]

    # Write header row
    headers = ["Metric", "Sheltered_ES", "Sheltered_TH", "Unsheltered", "Total"]
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        _apply_obs_header_style(cell, Font, PatternFill, Alignment)

    # Write data rows
    for row_idx, (label, key, is_average) in enumerate(rows, 2):
        worksheet.cell(row=row_idx, column=1, value=label)

        total = 0
        count = 0
        for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
            value = source_stats.get(source, {}).get(key, 0)
            worksheet.cell(row=row_idx, column=col_idx, value=value if value else 0)
            if isinstance(value, (int, float)) and value:
                total += value
                count += 1

        # Total column - for averages, calculate weighted average or just average
        if is_average:
            avg_total = round(total / count, 2) if count > 0 else 0
            worksheet.cell(row=row_idx, column=5, value=avg_total)
        else:
            worksheet.cell(row=row_idx, column=5, value=total)

    _adjust_obs_column_widths(worksheet, get_column_letter)


def _create_obs_demographics_sheet(workbook, source_stats: Dict, region: str):
    """Create the Demographics sheet with age, sex, gender, and race distributions."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from config import OBSERVATION_VALID_AGE_RANGES

    worksheet = workbook.create_sheet("Demographics", 1)

    # Write header row
    headers = ["Category", "Sheltered_ES", "Sheltered_TH", "Unsheltered", "Total"]
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        _apply_obs_header_style(cell, Font, PatternFill, Alignment)

    current_row = 2

    # Age Range Distribution section
    cell = worksheet.cell(row=current_row, column=1, value="Age Range Distribution")
    _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    for col_idx in range(2, 6):
        cell = worksheet.cell(row=current_row, column=col_idx, value="")
        _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    current_row += 1

    for age in OBSERVATION_VALID_AGE_RANGES:
        key = f'age_{age}'
        worksheet.cell(row=current_row, column=1, value=f"  {age}")
        total = 0
        for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
            value = source_stats.get(source, {}).get(key, 0)
            worksheet.cell(row=current_row, column=col_idx, value=value)
            total += value
        worksheet.cell(row=current_row, column=5, value=total)
        current_row += 1

    # Unknown age
    worksheet.cell(row=current_row, column=1, value="  Unknown/Not Reported")
    total = 0
    for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
        value = source_stats.get(source, {}).get('age_unknown', 0)
        worksheet.cell(row=current_row, column=col_idx, value=value)
        total += value
    worksheet.cell(row=current_row, column=5, value=total)
    current_row += 1

    # Sex Distribution section
    cell = worksheet.cell(row=current_row, column=1, value="Sex Distribution")
    _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    for col_idx in range(2, 6):
        cell = worksheet.cell(row=current_row, column=col_idx, value="")
        _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    current_row += 1

    for sex, key in [("Male", "sex_male"), ("Female", "sex_female"), ("Unknown/Not Reported", "sex_unknown")]:
        worksheet.cell(row=current_row, column=1, value=f"  {sex}")
        total = 0
        for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
            value = source_stats.get(source, {}).get(key, 0)
            worksheet.cell(row=current_row, column=col_idx, value=value)
            total += value
        worksheet.cell(row=current_row, column=5, value=total)
        current_row += 1

    # Gender Distribution section (New England only)
    if region == 'New England':
        cell = worksheet.cell(row=current_row, column=1, value="Gender Distribution")
        _apply_obs_category_style(cell, Font, PatternFill, Alignment)
        for col_idx in range(2, 6):
            cell = worksheet.cell(row=current_row, column=col_idx, value="")
            _apply_obs_category_style(cell, Font, PatternFill, Alignment)
        current_row += 1

        gender_items = [
            ("Woman (Girl if child)", "gender_woman"),
            ("Man (Boy if child)", "gender_man"),
            ("Culturally Specific Identity", "gender_culturally_specific"),
            ("Transgender", "gender_transgender"),
            ("Non-Binary", "gender_nonbinary"),
            ("Questioning", "gender_questioning"),
            ("Different Identity", "gender_different"),
            ("Unknown/Not Reported", "gender_unknown"),
        ]

        for label, key in gender_items:
            worksheet.cell(row=current_row, column=1, value=f"  {label}")
            total = 0
            for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
                value = source_stats.get(source, {}).get(key, 0)
                worksheet.cell(row=current_row, column=col_idx, value=value)
                total += value
            worksheet.cell(row=current_row, column=5, value=total)
            current_row += 1

    # Race/Ethnicity Distribution section
    cell = worksheet.cell(row=current_row, column=1, value="Race/Ethnicity Distribution")
    _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    for col_idx in range(2, 6):
        cell = worksheet.cell(row=current_row, column=col_idx, value="")
        _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    current_row += 1

    race_items = [
        ("White", "race_white"),
        ("Black, African American, or African", "race_black"),
        ("Asian or Asian American", "race_asian"),
        ("American Indian, Alaska Native, or Indigenous", "race_indigenous"),
        ("Native Hawaiian or Pacific Islander", "race_pacific"),
        ("Middle Eastern or North African", "race_middle_eastern"),
        ("Hispanic/Latina/e/o", "race_hispanic"),
        ("Multi-Racial", "race_multiracial"),
        ("Unknown/Not Reported", "race_unknown"),
    ]

    for label, key in race_items:
        worksheet.cell(row=current_row, column=1, value=f"  {label}")
        total = 0
        for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
            value = source_stats.get(source, {}).get(key, 0)
            worksheet.cell(row=current_row, column=col_idx, value=value)
            total += value
        worksheet.cell(row=current_row, column=5, value=total)
        current_row += 1

    _adjust_obs_column_widths(worksheet, get_column_letter)


def _create_obs_location_summary_sheet(workbook, observation_data: Dict, region: str):
    """Create Location Summary sheet with dynamic location categories."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    worksheet = workbook.create_sheet("Location Summary", 2)

    # Calculate location stats
    location_stats = _calculate_obs_location_stats(observation_data, region)

    # Write header row
    headers = ["Location", "Sheltered_ES", "Sheltered_TH", "Unsheltered", "Total"]
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        _apply_obs_header_style(cell, Font, PatternFill, Alignment)

    current_row = 2

    # By County section
    county_label = "By County" if region == 'New England' else "By Location"
    cell = worksheet.cell(row=current_row, column=1, value=county_label)
    _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    for col_idx in range(2, 6):
        cell = worksheet.cell(row=current_row, column=col_idx, value="")
        _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    current_row += 1

    for location, source_counts in sorted(location_stats['by_county'].items()):
        worksheet.cell(row=current_row, column=1, value=f"  {location}")
        total = 0
        for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
            value = source_counts.get(source, 0)
            worksheet.cell(row=current_row, column=col_idx, value=value)
            total += value
        worksheet.cell(row=current_row, column=5, value=total)
        current_row += 1

    if not location_stats['by_county']:
        worksheet.cell(row=current_row, column=1, value="  (No data available)")
        current_row += 1

    # By Project section
    cell = worksheet.cell(row=current_row, column=1, value="By Project Name")
    _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    for col_idx in range(2, 6):
        cell = worksheet.cell(row=current_row, column=col_idx, value="")
        _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    current_row += 1

    for project, source_counts in sorted(location_stats['by_project'].items()):
        worksheet.cell(row=current_row, column=1, value=f"  {project}")
        total = 0
        for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
            value = source_counts.get(source, 0)
            worksheet.cell(row=current_row, column=col_idx, value=value)
            total += value
        worksheet.cell(row=current_row, column=5, value=total)
        current_row += 1

    if not location_stats['by_project']:
        worksheet.cell(row=current_row, column=1, value="  (No data available)")
        current_row += 1

    # By Welfare Office section (New England only)
    if region == 'New England':
        cell = worksheet.cell(row=current_row, column=1, value="By Welfare Office Town")
        _apply_obs_category_style(cell, Font, PatternFill, Alignment)
        for col_idx in range(2, 6):
            cell = worksheet.cell(row=current_row, column=col_idx, value="")
            _apply_obs_category_style(cell, Font, PatternFill, Alignment)
        current_row += 1

        for office, source_counts in sorted(location_stats['by_welfare_office'].items()):
            worksheet.cell(row=current_row, column=1, value=f"  {office}")
            total = 0
            for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
                value = source_counts.get(source, 0)
                worksheet.cell(row=current_row, column=col_idx, value=value)
                total += value
            worksheet.cell(row=current_row, column=5, value=total)
            current_row += 1

        if not location_stats['by_welfare_office']:
            worksheet.cell(row=current_row, column=1, value="  (No data available)")
            current_row += 1

    _adjust_obs_column_widths(worksheet, get_column_letter)


def _create_obs_household_analysis_sheet(workbook, source_stats: Dict, region: str):
    """Create Household Analysis sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    worksheet = workbook.create_sheet("Household Analysis", 3)

    # Write header row
    headers = ["Metric", "Sheltered_ES", "Sheltered_TH", "Unsheltered", "Total"]
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        _apply_obs_header_style(cell, Font, PatternFill, Alignment)

    current_row = 2

    # Household Size Distribution section
    cell = worksheet.cell(row=current_row, column=1, value="Household Size Distribution")
    _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    for col_idx in range(2, 6):
        cell = worksheet.cell(row=current_row, column=col_idx, value="")
        _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    current_row += 1

    size_items = [
        ("1 person", "hh_size_1"),
        ("2 persons", "hh_size_2"),
        ("3 persons", "hh_size_3"),
        ("4 persons", "hh_size_4"),
        ("5+ persons", "hh_size_5plus"),
    ]

    for label, key in size_items:
        worksheet.cell(row=current_row, column=1, value=f"  {label}")
        total = 0
        for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
            value = source_stats.get(source, {}).get(key, 0)
            worksheet.cell(row=current_row, column=col_idx, value=value)
            total += value
        worksheet.cell(row=current_row, column=5, value=total)
        current_row += 1

    # Household Composition section
    cell = worksheet.cell(row=current_row, column=1, value="Household Composition")
    _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    for col_idx in range(2, 6):
        cell = worksheet.cell(row=current_row, column=col_idx, value="")
        _apply_obs_category_style(cell, Font, PatternFill, Alignment)
    current_row += 1

    comp_items = [
        ("Adults only", "hh_comp_adults_only"),
        ("Adults with children", "hh_comp_with_children"),
        ("Children only", "hh_comp_children_only"),
        ("Unknown composition", "hh_comp_unknown"),
    ]

    for label, key in comp_items:
        worksheet.cell(row=current_row, column=1, value=f"  {label}")
        total = 0
        for col_idx, source in enumerate(['Sheltered_ES', 'Sheltered_TH', 'Unsheltered'], 2):
            value = source_stats.get(source, {}).get(key, 0)
            worksheet.cell(row=current_row, column=col_idx, value=value)
            total += value
        worksheet.cell(row=current_row, column=5, value=total)
        current_row += 1

    _adjust_obs_column_widths(worksheet, get_column_letter)


def generate_observation_data_export(uploaded_data: Dict[str, pd.DataFrame], region: str) -> BytesIO:
    """
    Generate Excel file with observation data and summary sheets from all sources.

    Creates summary sheets followed by raw data sheets:
    - Summary Overview
    - Demographics
    - Location Summary
    - Household Analysis
    - Emergency Shelter (ES) - raw data
    - Transitional Housing (TH) - raw data
    - Unsheltered - raw data

    Args:
        uploaded_data: Dict mapping source names to raw DataFrames
        region: Region identifier for column selection

    Returns:
        BytesIO buffer containing Excel file

    Raises:
        ValueError: If no observation data found in any source
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Source labels for Excel sheet names
    SOURCE_LABELS = {
        'Sheltered_ES': 'Emergency Shelter (ES)',
        'Sheltered_TH': 'Transitional Housing (TH)',
        'Unsheltered': 'Unsheltered'
    }

    # Process each source
    observation_data = {}
    total_records = 0

    for source_name in ['Sheltered_ES', 'Sheltered_TH', 'Unsheltered']:
        raw_df = uploaded_data.get(source_name)
        if raw_df is None or raw_df.empty:
            continue

        # Filter and select columns
        filtered_df = filter_observation_data(raw_df, source_name)
        if filtered_df.empty:
            continue

        selected_df = select_observation_columns(filtered_df, region)
        if not selected_df.empty:
            observation_data[source_name] = selected_df
            total_records += len(selected_df)

    # Check if any data found
    if not observation_data:
        raise ValueError("No observation data found matching filter criteria in any source.")

    # Calculate statistics for each source
    source_stats = {}
    for source_name, df in observation_data.items():
        source_stats[source_name] = calculate_observation_stats(df, region)

    # Create Excel workbook
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        workbook = writer.book

        # ====================================================================
        # Create Summary Sheets (first 4 sheets)
        # ====================================================================

        # 1. Summary Overview sheet
        _create_obs_summary_overview_sheet(workbook, source_stats, region)

        # 2. Demographics sheet
        _create_obs_demographics_sheet(workbook, source_stats, region)

        # 3. Location Summary sheet
        _create_obs_location_summary_sheet(workbook, observation_data, region)

        # 4. Household Analysis sheet
        _create_obs_household_analysis_sheet(workbook, source_stats, region)

        # ====================================================================
        # Create Raw Data Sheets (after summary sheets)
        # ====================================================================

        for source_name in ['Sheltered_ES', 'Sheltered_TH', 'Unsheltered']:
            if source_name not in observation_data:
                continue

            df = observation_data[source_name]
            sheet_name = SOURCE_LABELS[source_name]

            # Create sheet for this source (index 4+ to place after summary sheets)
            worksheet = workbook.create_sheet(sheet_name)

            # Write headers (row 1)
            for col_idx, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=column)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="00629b", end_color="00629b", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Write data rows (starting row 2)
            for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

            # Auto-adjust column widths
            for col_idx, column in enumerate(df.columns, 1):
                max_length = max(
                    len(str(column)),
                    df[column].astype(str).str.len().max() if not df.empty else 0
                )
                adjusted_width = min(max(max_length + 2, 10), 60)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Remove default sheet if exists
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

    excel_buffer.seek(0)
    return excel_buffer

def create_comprehensive_excel_export(reports_data, raw_data_with_ids, processed_data_with_ids, 
                                    include_raw, include_processed):
    """Create Excel file with all reports and optional data - EXACT ORIGINAL FORMATTING"""
    excel_buffer = BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        workbook = writer.book
        
        # Create styles
        cell_style = NamedStyle(
            name="cell_style",
            font=Font(size=11),
            alignment=Alignment(horizontal="left", vertical="center")
        )
        
        title_style = NamedStyle(
            name="title_style",
            font=Font(size=12, bold=True, color="000000"),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill(start_color="e2efe8", end_color="e2efe8", fill_type="solid")
        )
        
        if 'cell_style' not in workbook.named_styles:
            workbook.add_named_style(cell_style)
        if 'title_style' not in workbook.named_styles:
            workbook.add_named_style(title_style)
        
        # Process each report type
        for report_type, reports in reports_data.items():
            if not reports:
                continue
            
            # Create worksheet for this report type
            if report_type not in workbook.sheetnames:
                workbook.create_sheet(report_type)
            
            worksheet = workbook[report_type]
            worksheet.sheet_state = 'visible'
            
            current_row = 1
            
            # Add each report to the worksheet
            for report_name, report_df in reports.items():
                if report_df.empty:
                    continue
                
                # Write the DataFrame to Excel
                report_df.to_excel(
                    writer,
                    sheet_name=report_type,
                    index=True,
                    startrow=current_row
                )
                
                # Format the section
                format_worksheet_section(worksheet, report_df, report_name, current_row)
                
                # Move to next section
                current_row += len(report_df) + 8
        
        # Add raw data sheets if requested
        if include_raw and raw_data_with_ids:
            for source_name, raw_df in raw_data_with_ids.items():
                # Shorten source name for Excel 31-char sheet name limit
                short_source = source_name.replace('Sheltered_', '').replace('Unsheltered', 'Unshelt')
                sheet_name = f"Raw_{short_source}"[:31]
                # Clean data before export to avoid Excel corruption
                cleaned_raw = clean_dataframe_for_export(raw_df)
                cleaned_raw.to_excel(writer, sheet_name=sheet_name, index=False)

                # Format the raw data sheet
                worksheet = workbook[sheet_name]
                worksheet.sheet_state = 'visible'

                # Auto-adjust column widths
                for col_num, column in enumerate(raw_df.columns, 1):
                    max_length = max(
                        len(str(column)),
                        raw_df[column].astype(str).str.len().max() if not raw_df.empty else 0
                    )
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        # Add processed data sheets if requested
        if include_processed and processed_data_with_ids:
            for source_name, data_dict in processed_data_with_ids.items():
                # Shorten source name for Excel 31-char sheet name limit
                short_source = source_name.replace('Sheltered_', '').replace('Unsheltered', 'Unshelt')

                # Add persons sheet
                if 'persons' in data_dict and not data_dict['persons'].empty:
                    sheet_name = f"Persons_{short_source}"[:31]
                    # Clean data before export to avoid Excel corruption
                    cleaned_persons = clean_dataframe_for_export(data_dict['persons'])
                    cleaned_persons.to_excel(writer, sheet_name=sheet_name, index=False)

                    worksheet = workbook[sheet_name]
                    worksheet.sheet_state = 'visible'

                    # Auto-adjust column widths
                    for col_num, column in enumerate(data_dict['persons'].columns, 1):
                        max_length = max(
                            len(str(column)),
                            data_dict['persons'][column].astype(str).str.len().max() if not data_dict['persons'].empty else 0
                        )
                        adjusted_width = min(max(max_length + 2, 10), 50)
                        worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

                # Add households sheet
                if 'households' in data_dict and not data_dict['households'].empty:
                    sheet_name = f"Households_{short_source}"[:31]
                    # Clean data before export to avoid Excel corruption
                    cleaned_households = clean_dataframe_for_export(data_dict['households'])
                    cleaned_households.to_excel(writer, sheet_name=sheet_name, index=False)

                    worksheet = workbook[sheet_name]
                    worksheet.sheet_state = 'visible'

                    # Auto-adjust column widths
                    for col_num, column in enumerate(data_dict['households'].columns, 1):
                        max_length = max(
                            len(str(column)),
                            data_dict['households'][column].astype(str).str.len().max() if not data_dict['households'].empty else 0
                        )
                        adjusted_width = min(max(max_length + 2, 10), 50)
                        worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        
        # Remove default sheet if it exists and we have other sheets
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook["Sheet"]
    
    excel_buffer.seek(0)
    return excel_buffer

def format_worksheet_section(worksheet, df, title, start_row):
    """Apply formatting to a worksheet section"""
    from openpyxl.utils import get_column_letter
    
    # Add title
    title_cell = worksheet.cell(row=start_row, column=1, value=title)
    title_cell.style = 'title_style'
    
    # Merge title cells
    end_col = df.shape[1] + 2 if not df.empty else 3
    worksheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=end_col
    )
    
    # Format data cells
    data_start_row = start_row + 1
    data_end_row = data_start_row + df.shape[0] + 1
    
    for row in worksheet.iter_rows(
        min_row=data_start_row,
        max_row=data_end_row,
        min_col=1,
        max_col=end_col
    ):
        for cell in row:
            cell.style = 'cell_style'
    
    # Auto-adjust column widths
    for col_num in range(1, end_col + 1):
        column = worksheet[get_column_letter(col_num)]
        max_length = 0
        
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set minimum width and maximum width
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

def show_individual_downloads(calculated_reports, region):
    """Show individual report downloads section"""
    st.subheader("Individual Report Downloads")
    
    if not calculated_reports:
        st.info("No reports available.")
        return
    
    # Create report options
    report_options = []
    for report_type, reports in calculated_reports.items():
        for report_name in reports.keys():
            report_options.append(f"{report_type} - {report_name}")
    
    if not report_options:
        st.info("No individual reports available.")
        return
    
    selected_report = st.selectbox(
        "Select Report for Individual Download",
        report_options
    )
    
    if selected_report:
        # Parse selection
        report_type, report_name = selected_report.split(" - ", 1)
        report_df = calculated_reports[report_type][report_name]
        
        # Show preview
        with st.expander(f"Preview: {selected_report}"):
            st.dataframe(report_df, width='stretch')
        
        col1, col2 = st.columns(2)
        
        with col1:
            # CSV download
            csv_data = report_df.to_csv()
            csv_filename = f"{region}_{report_type}_{report_name.replace(' ', '_')}.csv"
            
            st.download_button(
                label="📄 Download as CSV",
                data=csv_data,
                file_name=csv_filename,
                mime='text/csv'
            )
        
        with col2:
            # Excel download (single sheet)
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                report_df.to_excel(writer, sheet_name=report_name[:30], index=True)
            
            excel_filename = f"{region}_{report_type}_{report_name.replace(' ', '_')}.xlsx"
            
            st.download_button(
                label="📊 Download as Excel",
                data=excel_buffer.getvalue(),
                file_name=excel_filename,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

def show_raw_data_downloads(uploaded_data, processed_data, region):
    """Show raw data downloads section"""
    st.subheader("Enhanced Raw Data Downloads")
    
    if not uploaded_data:
        st.info("No raw data available.")
        return
    
    st.info("Raw data downloads include Household ID and Person IDs for verification purposes.")
    
    # Prepare raw data with IDs
    raw_data_with_ids = prepare_raw_data_with_ids(uploaded_data, processed_data)
    
    if not raw_data_with_ids:
        st.warning("No processed data available to generate IDs.")
        return
    
    # Show available sources
    for source_name, enhanced_df in raw_data_with_ids.items():
        st.write(f"### {source_name}")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Households", len(enhanced_df))
        
        with col2:
            # Count total persons from Person_IDs column
            total_persons = sum(len(pid.split(', ')) for pid in enhanced_df['Person_IDs'])
            st.metric("Total Persons", total_persons)
        
        with col3:
            # Download button
            csv_data = enhanced_df.to_csv(index=False)
            filename = f"{region}_{source_name}_Raw_with_IDs.csv"
            
            st.download_button(
                label="📥 Download CSV",
                data=csv_data,
                file_name=filename,
                mime='text/csv',
                key=f"raw_{source_name}"
            )
        
        # Preview
        with st.expander(f"Preview first 10 rows"):
            # Show only key columns in preview
            preview_cols = ['Household_ID', 'Person_IDs', 'Timestamp', 'Gender', 'Age Range', 'Race/Ethnicity']
            display_cols = [col for col in preview_cols if col in enhanced_df.columns]
            from utils import safe_dataframe_display
            st.dataframe(safe_dataframe_display(enhanced_df[display_cols].head(10)), width='stretch')

def show_processed_data_downloads(processed_data, region):
    """Show processed data downloads section"""
    st.subheader("Processed Data Downloads")
    
    if not processed_data:
        st.info("No processed data available.")
        return
    
    st.info("Download the fully processed persons and households data with all calculated fields.")
    
    # Prepare processed data
    processed_data_with_ids = prepare_processed_data_with_ids(processed_data)
    
    if not processed_data_with_ids:
        st.warning("No processed data available.")
        return
    
    # Show available sources
    for source_name, data_dict in processed_data_with_ids.items():
        st.write(f"### {source_name}")
        
        # Create tabs for persons and households
        tab1, tab2 = st.tabs(["👥 Persons Data", "🏠 Households Data"])
        
        with tab1:
            if 'persons' in data_dict and not data_dict['persons'].empty:
                persons_df = data_dict['persons']
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Total Persons", len(persons_df))
                
                with col2:
                    st.metric("Unique Households", persons_df['Household_ID'].nunique())
                
                with col3:
                    # Download button
                    csv_data = persons_df.to_csv(index=False)
                    filename = f"{region}_{source_name}_Processed_Persons.csv"
                    
                    st.download_button(
                        label="📥 Download Persons CSV",
                        data=csv_data,
                        file_name=filename,
                        mime='text/csv',
                        key=f"processed_persons_{source_name}"
                    )
                
                # Preview
                with st.expander("Preview first 20 rows"):
                    st.dataframe(persons_df.head(20), width='stretch')
            else:
                st.info("No persons data available for this source.")
        
        with tab2:
            if 'households' in data_dict and not data_dict['households'].empty:
                households_df = data_dict['households']
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Total Households", len(households_df))
                
                with col2:
                    avg_size = households_df['total_persons'].mean()
                    st.metric("Avg Household Size", f"{avg_size:.1f}")
                
                with col3:
                    # Download button
                    csv_data = households_df.to_csv(index=False)
                    filename = f"{region}_{source_name}_Processed_Households.csv"
                    
                    st.download_button(
                        label="📥 Download Households CSV",
                        data=csv_data,
                        file_name=filename,
                        mime='text/csv',
                        key=f"processed_households_{source_name}"
                    )
                
                # Preview
                with st.expander("Preview first 20 rows"):
                    st.dataframe(households_df.head(20), width='stretch')
            else:
                st.info("No households data available for this source.")

def show_additional_resources():
    """Show additional resources section"""
    st.subheader("Additional Resources")

    st.write("**Related Tools:**")
    st.markdown("- **Combine HMIS & Non-HMIS Data**: Use the Combine Data step in the main navigation")

    st.write("**Support:**")
    st.write("For technical support or questions about the PIT Count application, please contact your system administrator.")


def show_combine_interface():
    """Show the PIT Combiner interface for combining HMIS and Non-HMIS data."""
    from config import (
        CombinerConfig,
        COMBINER_RANGE_SPECIFICATIONS,
        COMBINER_TERMS_TO_DELETE,
        COMBINER_VALIDATION_RULES
    )
    from processor import CombinerDataProcessor
    from pathlib import Path

    st.markdown("""
    Use this tool to combine your HMIS (HUDX 230) data with Non-HMIS (HDX) data
    into a unified HUD submission template.
    """)

    # Initialize config
    config = CombinerConfig()

    # Check template exists
    template_path = Path(config.template_file)
    if not template_path.exists():
        st.error(f"Template file '{config.template_file}' not found. Please ensure the template is in the application directory.")
        return

    st.info(f"Template file: {config.template_file} (found)")

    st.markdown("---")
    st.markdown("### Upload Files for Processing")

    # Create two columns for file uploaders
    col1, col2 = st.columns(2)

    with col1:
        st.info("**HMIS Data File (HUDX 230)**")
        hmis_file = st.file_uploader(
            "Upload HMIS Data",
            type=['xlsx'],
            key="combiner_hmis_uploader",
            help="Upload your HUDX 230 AD report from your HMIS system"
        )
        if hmis_file:
            st.caption(f"{hmis_file.name} ({hmis_file.size:,} bytes)")

    with col2:
        st.info("**Non-HMIS Data File (HDX)**")
        non_hmis_file = st.file_uploader(
            "Upload Non-HMIS Data",
            type=['xlsx'],
            key="combiner_non_hmis_uploader",
            help="Upload your HDX data file containing unsheltered counts and additional data"
        )
        if non_hmis_file:
            st.caption(f"{non_hmis_file.name} ({non_hmis_file.size:,} bytes)")

    # Process button
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        process_button = st.button(
            "Process and Combine Files",
            type="primary",
            width='stretch',
            disabled=(hmis_file is None or non_hmis_file is None)
        )

    if process_button:
        # Validate files
        is_valid, error_msg = _validate_combiner_files(hmis_file, non_hmis_file)

        if not is_valid:
            st.error(f"{error_msg}")
        else:
            # Process files with progress indicator
            with st.spinner("Processing files... This may take a moment."):
                processor = CombinerDataProcessor(config.template_file)

                try:
                    output = processor.process_and_combine(
                        hmis_stream=hmis_file,
                        non_hmis_stream=non_hmis_file,
                        range_specs=COMBINER_RANGE_SPECIFICATIONS,
                        terms_to_delete=COMBINER_TERMS_TO_DELETE,
                        validation_rules=COMBINER_VALIDATION_RULES
                    )

                    st.success("Files processed successfully!")

                    # Get timestamp for filename
                    region = st.session_state.get('region', 'NewEngland')
                    timezone = get_timezone_for_region(region) if region else 'America/New_York'
                    timestamp = get_current_timestamp(timezone)

                    output_bytes = output.getvalue()
                    output_filename = f"combined_pit_data_{timestamp}.xlsx"

                    # Store in session state for potential re-download
                    st.session_state['combiner_output'] = output_bytes
                    st.session_state['combiner_filename'] = output_filename

                    # Download button
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        st.download_button(
                            label="Download Combined Data",
                            data=output_bytes,
                            file_name=output_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            width='stretch'
                        )

                    # Display summary
                    with st.expander("Processing Summary", expanded=True):
                        st.markdown(f"""
                        **Input Files:**
                        - HMIS File: `{hmis_file.name}`
                        - Non-HMIS File: `{non_hmis_file.name}`

                        **Output:**
                        - File: `{output_filename}`
                        - Size: {len(output_bytes):,} bytes

                        **Processed:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
                        """)

                except FileNotFoundError:
                    st.error("Template file not found. Please ensure template.xlsx exists in the application directory.")
                except Exception as e:
                    st.error(f"An error occurred while processing: {str(e)}")

    # Show previous download if available (check that value is not None)
    if st.session_state.get('combiner_output') is not None and not process_button:
        st.markdown("---")
        st.info("Previous combined file is available for download:")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.download_button(
                label=f"Re-download: {st.session_state.get('combiner_filename', 'combined_data.xlsx')}",
                data=st.session_state['combiner_output'],
                file_name=st.session_state.get('combiner_filename', 'combined_data.xlsx'),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="secondary",
                width='stretch'
            )

    # Instructions section
    with st.expander("Instructions & Data Requirements", expanded=False):
        st.markdown("""
        ### How to Use the PIT Combiner

        1. **Upload HMIS Data**: Select your HUDX 230 AD report from your HMIS system
        2. **Upload Non-HMIS Data**: Select your HDX data file with unsheltered counts
        3. **Process Files**: Click "Process and Combine Files" to merge the data
        4. **Download Result**: Download the combined Excel file

        ### File Requirements

        **HMIS File (HUDX 230) must contain these sheets:**
        - Adult-Child
        - Without Children
        - Only Children
        - Veteran Adult-Child
        - Veteran Without Children
        - Unaccompanied Youth
        - Parenting Youth
        - Homeless Subpopulations

        **Non-HMIS File must contain these sheets:**
        - HDX_Totals
        - HDX_Veterans
        - HDX_Youth
        - HDX_Subpopulations

        ### Data Processing Notes

        - Rows with "Client Doesn't Know", "Missing Information", etc. are automatically cleaned
        - Template formulas (SUM for totals) are preserved
        - Data from both sources is combined cell-by-cell
        - N/A values are treated as zero
        """)

    # Technical details section
    with st.expander("Technical Details", expanded=False):
        st.markdown("""
        ### Data Mapping Overview

        The combiner maps data from two sources into the template:

        | Template Column | Data Source |
        |----------------|-------------|
        | B (Sheltered ES) | HMIS + Non-HMIS |
        | C (Sheltered TH) | HMIS + Non-HMIS |
        | D (Unsheltered) | Non-HMIS only |
        | E (Total) | Formula (preserved) |

        ### Template Sheets

        - **All Households**: Combined data for all household types
        - **Veteran Households Only**: Veteran-specific data
        - **Youth Households**: Unaccompanied and parenting youth
        - **Additional Homeless Populations**: Subpopulation data (mental illness, substance use, etc.)
        """)


def _validate_combiner_files(hmis_file, non_hmis_file) -> Tuple[bool, Optional[str]]:
    """Validate that both required files are uploaded and valid."""
    # Check HMIS file
    if hmis_file is None:
        return False, "HMIS file is required"
    if not hmis_file.name.endswith('.xlsx'):
        return False, "HMIS file must be an Excel file (.xlsx format)"
    if hmis_file.size == 0:
        return False, "HMIS file is empty"

    # Check Non-HMIS file
    if non_hmis_file is None:
        return False, "Non-HMIS file is required"
    if not non_hmis_file.name.endswith('.xlsx'):
        return False, "Non-HMIS file must be an Excel file (.xlsx format)"
    if non_hmis_file.size == 0:
        return False, "Non-HMIS file is empty"

    # Check file size (max 50MB)
    max_size = 50 * 1024 * 1024
    if hmis_file.size > max_size:
        return False, "HMIS file too large. Maximum size is 50MB"
    if non_hmis_file.size > max_size:
        return False, "Non-HMIS file too large. Maximum size is 50MB"

    return True, None


def show_dv_summation_interface():
    """Show the DV Excel Summation interface for combining multiple DV Excel files."""
    import io
    from pathlib import Path

    st.markdown("""
    Use this tool to sum numeric values across multiple Excel files that follow
    the DV template structure. All matching cells with numeric values will be summed.
    """)

    # Check template exists
    template_path = Path("DV_temp.xlsx")
    if not template_path.exists():
        st.error("Template file 'DV_temp.xlsx' not found. Please ensure the template is in the application directory.")
        return

    st.info("Template file: DV_temp.xlsx found")
    st.markdown("---")

    # File uploader - multiple files
    uploaded_files = st.file_uploader(
        "Upload Excel files to sum",
        type=["xlsx"],
        accept_multiple_files=True,
        key="dv_sum_uploader",
        help="Upload multiple Excel files with the same structure as DV_temp.xlsx"
    )

    if uploaded_files:
        st.caption(f"{len(uploaded_files)} file(s) selected:")
        for f in uploaded_files:
            st.caption(f"  - {f.name} ({f.size:,} bytes)")

    # Process button
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        process_button = st.button(
            "Sum Excel Files",
            type="primary",
            disabled=len(uploaded_files) < 2 if uploaded_files else True
        )

    if process_button and uploaded_files:
        # Validate files
        is_valid, error_msg = _validate_dv_files(uploaded_files)

        if not is_valid:
            st.error(error_msg)
        else:
            with st.spinner("Processing files..."):
                try:
                    output = sum_excel_files(uploaded_files, "DV_temp.xlsx")

                    st.success("Summation completed!")

                    # Generate filename with timestamp
                    region = st.session_state.get('region', 'NewEngland')
                    timezone = get_timezone_for_region(region) if region else 'America/New_York'
                    timestamp = get_current_timestamp(timezone)
                    output_filename = f"DV_summed_{timestamp}.xlsx"

                    output_bytes = output.getvalue()

                    # Store for re-download
                    st.session_state['dv_sum_output'] = output_bytes
                    st.session_state['dv_sum_filename'] = output_filename

                    # Download button
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        st.download_button(
                            label="Download Summed Data",
                            data=output_bytes,
                            file_name=output_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )

                    # Summary
                    with st.expander("Processing Summary", expanded=True):
                        file_list = "\n".join([f"  - {f.name}" for f in uploaded_files])
                        st.markdown(f"""
**Input Files ({len(uploaded_files)}):**
{file_list}

**Output:** `{output_filename}`
                        """)

                except Exception as e:
                    st.error(f"An error occurred: {str(e)}")

    # Show previous download if available
    if st.session_state.get('dv_sum_output') is not None and not (process_button if uploaded_files else False):
        st.markdown("---")
        st.info("Previous summed file is available for download:")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.download_button(
                label=f"Re-download: {st.session_state.get('dv_sum_filename', 'dv_summed.xlsx')}",
                data=st.session_state['dv_sum_output'],
                file_name=st.session_state.get('dv_sum_filename', 'dv_summed.xlsx'),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="secondary"
            )

    # Instructions
    with st.expander("Instructions", expanded=False):
        st.markdown("""
### How to Use
1. Upload 2 or more Excel files with the same structure as DV_temp.xlsx
2. Click "Sum Excel Files" to process
3. Download the resulting file with summed values

### Processing Rules
- Only numeric cells are summed
- Non-numeric cells (text, headers) are preserved from the template
- Empty/null cells are treated as 0 for summation
- All sheets present in both template and source files are processed
        """)


def _validate_dv_files(files) -> Tuple[bool, Optional[str]]:
    """Validate uploaded DV files."""
    if not files or len(files) < 2:
        return False, "Please upload at least 2 Excel files to sum"

    max_size = 50 * 1024 * 1024  # 50MB
    for f in files:
        if not f.name.endswith('.xlsx'):
            return False, f"File '{f.name}' must be .xlsx format"
        if f.size == 0:
            return False, f"File '{f.name}' is empty"
        if f.size > max_size:
            return False, f"File '{f.name}' exceeds 50MB limit"

    return True, None


def sum_excel_files(uploaded_files, template_path: str) -> BytesIO:
    """Sum numeric values across multiple Excel files using template structure.

    - Sums all numeric values from source files
    - Treats null/empty cells as 0 for numeric positions
    - Preserves text cells (headers, labels) from template
    - Calculates Total column as sum of data columns
    """
    import openpyxl

    # Load template
    wb_dest = openpyxl.load_workbook(template_path)

    # Load all source workbooks
    workbooks = []
    for f in uploaded_files:
        f.seek(0)  # Reset file pointer
        wb = openpyxl.load_workbook(f)
        workbooks.append(wb)

    # Process each sheet in template
    for sheet_name in wb_dest.sheetnames:
        # Get sheets from sources that have this sheet
        source_sheets = [wb[sheet_name] for wb in workbooks if sheet_name in wb.sheetnames]

        if not source_sheets:
            continue

        ws_dest = wb_dest[sheet_name]
        max_row = ws_dest.max_row
        max_col = ws_dest.max_column

        # First pass: identify which cells should be numeric based on template
        # Check template cell - if it's numeric or 0, treat that position as numeric
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                template_val = ws_dest.cell(row=row, column=col).value

                # Check if this cell position should be numeric
                # (template has a number, or sources have numbers here)
                is_numeric_cell = isinstance(template_val, (int, float))

                if not is_numeric_cell:
                    # Check if any source has a numeric value here
                    for ws in source_sheets:
                        try:
                            val = ws.cell(row=row, column=col).value
                            if isinstance(val, (int, float)):
                                is_numeric_cell = True
                                break
                            elif val is not None:
                                try:
                                    float(val)
                                    is_numeric_cell = True
                                    break
                                except (ValueError, TypeError):
                                    pass
                        except Exception:
                            pass

                if is_numeric_cell:
                    # Sum values from all source files, treating null as 0
                    total = 0
                    for ws in source_sheets:
                        try:
                            val = ws.cell(row=row, column=col).value
                            if val is not None and isinstance(val, (int, float)):
                                total += val
                            elif val is not None:
                                # Try to convert string numbers
                                try:
                                    total += float(val)
                                except (ValueError, TypeError):
                                    pass
                            # null/None treated as 0 (no addition needed)
                        except Exception:
                            pass

                    ws_dest.cell(row=row, column=col).value = total

        # Second pass: Calculate Total column (usually last column or column with "Total" header)
        # Find the Total column by checking header row
        total_col = None
        for col in range(1, max_col + 1):
            header_val = ws_dest.cell(row=1, column=col).value
            if header_val and str(header_val).strip().lower() == 'total':
                total_col = col
                break

        # If Total column found, recalculate it as sum of other data columns
        if total_col:
            # Find data columns (numeric columns before Total, excluding row labels)
            data_cols = []
            for col in range(1, total_col):
                # Check if this column has numeric data (check a few rows)
                for check_row in range(2, min(10, max_row + 1)):
                    val = ws_dest.cell(row=check_row, column=col).value
                    if isinstance(val, (int, float)):
                        data_cols.append(col)
                        break

            # Calculate Total for each row
            for row in range(2, max_row + 1):  # Skip header row
                row_sum = 0
                has_data = False
                for col in data_cols:
                    val = ws_dest.cell(row=row, column=col).value
                    if isinstance(val, (int, float)):
                        row_sum += val
                        has_data = True

                if has_data:
                    ws_dest.cell(row=row, column=total_col).value = row_sum

    # Save to BytesIO
    output = BytesIO()
    wb_dest.save(output)
    output.seek(0)

    # Close workbooks
    for wb in workbooks:
        wb.close()
    wb_dest.close()

    return output
