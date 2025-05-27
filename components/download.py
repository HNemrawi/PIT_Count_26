"""
Download component for PIT Count application.
Handles export of reports to Excel format with formatting.
Enhanced to include raw data with Household ID and Person IDs.
Now includes processed households and persons data downloads.
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict

from utils.session import SessionManager
from utils.helpers import create_download_filename, get_current_timestamp, get_timezone_for_region

class ExcelExporter:
    """Handles Excel export functionality with formatting."""
    
    def __init__(self):
        """Initialize Excel exporter."""
        self.cell_style = None
        self.title_style = None
    
    def create_styles(self, workbook):
        """Create named styles for Excel formatting."""
        if 'cell_style' not in workbook.named_styles:
            self.cell_style = NamedStyle(
                name="cell_style",
                font=Font(size=11),
                alignment=Alignment(horizontal="left", vertical="center")
            )
            workbook.add_named_style(self.cell_style)
        
        if 'title_style' not in workbook.named_styles:
            self.title_style = NamedStyle(
                name="title_style",
                font=Font(size=12, bold=True, color="000000"),
                alignment=Alignment(horizontal="center", vertical="center"),
                fill=PatternFill(start_color="e2efe8", end_color="e2efe8", fill_type="solid")
            )
            workbook.add_named_style(self.title_style)
    
    def format_worksheet(self, worksheet, df: pd.DataFrame, title: str, start_row: int):
        """Apply formatting to a worksheet."""
        # Add title
        title_cell = worksheet.cell(row=start_row, column=1, value=title)
        title_cell.style = 'title_style'
        
        # Merge title cells
        end_col = df.shape[1] + 2 if not df.empty else 3
        worksheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=end_col
        )
        
        # Format data cells
        data_start_row = start_row + 1
        data_end_row = data_start_row + df.shape[0] + 1
        
        for row in worksheet.iter_rows(
            min_row=data_start_row,
            max_row=data_end_row,
            min_col=1,
            max_col=end_col
        ):
            for cell in row:
                cell.style = 'cell_style'
        
        # Auto-adjust column widths
        for col_num in range(1, end_col + 1):
            column = worksheet[get_column_letter(col_num)]
            max_length = 0
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set minimum width and maximum width
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
    
    def create_excel_file(self, reports_data: Dict[str, Dict[str, pd.DataFrame]], 
                         include_raw_data: bool = False,
                         raw_data_with_ids: Dict[str, pd.DataFrame] = None,
                         include_processed_data: bool = False,
                         processed_data_with_ids: Dict[str, Dict[str, pd.DataFrame]] = None) -> BytesIO:
        """Create Excel file with all reports and optionally raw/processed data with IDs."""
        excel_buffer = BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            workbook = writer.book
            
            # Create styles
            self.create_styles(workbook)
            
            # Process each report type
            for report_type, reports in reports_data.items():
                if not reports:
                    continue
                
                # Create worksheet for this report type
                if report_type not in workbook.sheetnames:
                    workbook.create_sheet(report_type)
                
                worksheet = workbook[report_type]
                worksheet.sheet_state = 'visible'
                
                current_row = 1
                
                # Add each report to the worksheet
                for report_name, report_df in reports.items():
                    if report_df.empty:
                        continue
                    
                    # Write the DataFrame to Excel
                    report_df.to_excel(
                        writer,
                        sheet_name=report_type,
                        index=True,
                        startrow=current_row
                    )
                    
                    # Format the section
                    self.format_worksheet(worksheet, report_df, report_name, current_row)
                    
                    # Move to next section
                    current_row += len(report_df) + 8
            
            # Add raw data sheets with IDs if requested
            if include_raw_data and raw_data_with_ids:
                for source_name, raw_df in raw_data_with_ids.items():
                    sheet_name = f"Raw_{source_name}"
                    raw_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Format the raw data sheet
                    worksheet = workbook[sheet_name]
                    worksheet.sheet_state = 'visible'
                    
                    # Auto-adjust column widths
                    for col_num, column in enumerate(raw_df.columns, 1):
                        max_length = max(
                            len(str(column)),
                            raw_df[column].astype(str).str.len().max() if not raw_df.empty else 0
                        )
                        adjusted_width = min(max(max_length + 2, 10), 50)
                        worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
            
            # Add processed data sheets if requested
            if include_processed_data and processed_data_with_ids:
                for source_name, data_dict in processed_data_with_ids.items():
                    # Add persons sheet
                    if 'persons' in data_dict:
                        sheet_name = f"Processed_Persons_{source_name}"
                        data_dict['persons'].to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        worksheet = workbook[sheet_name]
                        worksheet.sheet_state = 'visible'
                        
                        # Auto-adjust column widths
                        for col_num, column in enumerate(data_dict['persons'].columns, 1):
                            max_length = max(
                                len(str(column)),
                                data_dict['persons'][column].astype(str).str.len().max() if not data_dict['persons'].empty else 0
                            )
                            adjusted_width = min(max(max_length + 2, 10), 50)
                            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
                    
                    # Add households sheet
                    if 'households' in data_dict:
                        sheet_name = f"Processed_Households_{source_name}"
                        data_dict['households'].to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        worksheet = workbook[sheet_name]
                        worksheet.sheet_state = 'visible'
                        
                        # Auto-adjust column widths
                        for col_num, column in enumerate(data_dict['households'].columns, 1):
                            max_length = max(
                                len(str(column)),
                                data_dict['households'][column].astype(str).str.len().max() if not data_dict['households'].empty else 0
                            )
                            adjusted_width = min(max(max_length + 2, 10), 50)
                            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
            
            # Remove default sheet if it exists and we have other sheets
            if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
                del workbook["Sheet"]
        
        excel_buffer.seek(0)
        return excel_buffer

class DownloadManager:
    """Manages download functionality for PIT Count reports."""
    
    def __init__(self):
        """Initialize download manager with cached session data."""
        self.calculated_reports = SessionManager.get_calculated_reports()
        self.processed_data = SessionManager.get_processed_data()
        self.uploaded_data = SessionManager.get_uploaded_data()
        self.region = SessionManager.get_session_value('region', 'Unknown')
        self.exporter = ExcelExporter()
    
    def prepare_raw_data_with_ids(self) -> Dict[str, pd.DataFrame]:
        """Prepare raw data with Household ID and Person IDs added."""
        raw_data_with_ids = {}
        
        for source_name in ['Sheltered_ES', 'Sheltered_TH', 'Unsheltered']:
            # Get raw data
            raw_df = self.uploaded_data.get(source_name, pd.DataFrame())
            if raw_df.empty:
                continue
            
            # Get processed data for this source
            source_data = self.processed_data.get(source_name, {})
            persons_df = source_data.get('persons_df', pd.DataFrame())
            
            if persons_df.empty:
                continue
            
            # Create a copy of raw data
            enhanced_raw = raw_df.copy()
            
            # Add Household ID (1-based index matching the processing logic)
            enhanced_raw.insert(0, 'Household_ID', range(1, len(enhanced_raw) + 1))
            
            # Create Person IDs for each household
            person_ids_list = []
            
            for idx in range(len(enhanced_raw)):
                household_id = idx + 1
                
                # Get all persons in this household from processed data
                household_persons = persons_df[persons_df['Household_ID'] == household_id]
                
                if not household_persons.empty:
                    # Create person IDs in format: HH{household_id}_P{person_number}
                    person_ids = []
                    for _, person in household_persons.iterrows():
                        member_type = person.get('Member_Type', 'Unknown')
                        member_number = person.get('Member_Number', 1)
                        person_id = f"HH{household_id}_{member_type[0]}{member_number}"
                        person_ids.append(person_id)
                    
                    person_ids_str = ', '.join(sorted(person_ids))
                else:
                    # If no persons found in processed data, at least show primary
                    person_ids_str = f"HH{household_id}_A1"
                
                person_ids_list.append(person_ids_str)
            
            # Add Person IDs column
            enhanced_raw.insert(1, 'Person_IDs', person_ids_list)
            
            # Store enhanced raw data
            raw_data_with_ids[source_name] = enhanced_raw
        
        return raw_data_with_ids
    
    def prepare_processed_data_with_ids(self) -> Dict[str, Dict[str, pd.DataFrame]]:
        """Prepare processed persons and households data."""
        processed_data_with_ids = {}
        
        for source_name in ['Sheltered_ES', 'Sheltered_TH', 'Unsheltered']:
            source_data = self.processed_data.get(source_name, {})
            persons_df = source_data.get('persons_df', pd.DataFrame())
            households_df = source_data.get('households_df', pd.DataFrame())
            
            if persons_df.empty and households_df.empty:
                continue
            
            # Prepare persons data with proper ID format
            if not persons_df.empty:
                persons_enhanced = persons_df.copy()
                
                # Add Person_ID column
                person_ids = []
                for _, person in persons_enhanced.iterrows():
                    household_id = person.get('Household_ID', 0)
                    member_type = person.get('Member_Type', 'Unknown')
                    member_number = person.get('Member_Number', 1)
                    person_id = f"HH{household_id}_{member_type[0]}{member_number}"
                    person_ids.append(person_id)
                
                persons_enhanced.insert(0, 'Person_ID', person_ids)
                
                # Reorder columns for clarity
                important_cols = ['Person_ID', 'Household_ID', 'Member_Type', 'Member_Number', 
                                'Gender', 'race', 'age_range', 'age_group', 'DV', 'vet', 'CH']
                other_cols = [col for col in persons_enhanced.columns if col not in important_cols]
                persons_enhanced = persons_enhanced[important_cols + other_cols]
            else:
                persons_enhanced = pd.DataFrame()
            
            # Prepare households data
            if not households_df.empty:
                households_enhanced = households_df.copy()
                
                # Add person count and IDs for each household
                person_counts = []
                person_ids_lists = []
                
                for _, household in households_enhanced.iterrows():
                    household_id = household.get('household_id', 0)
                    
                    # Get all persons in this household
                    household_persons = persons_df[persons_df['Household_ID'] == household_id]
                    
                    person_counts.append(len(household_persons))
                    
                    if not household_persons.empty:
                        person_ids = []
                        for _, person in household_persons.iterrows():
                            member_type = person.get('Member_Type', 'Unknown')
                            member_number = person.get('Member_Number', 1)
                            person_id = f"HH{household_id}_{member_type[0]}{member_number}"
                            person_ids.append(person_id)
                        person_ids_lists.append(', '.join(sorted(person_ids)))
                    else:
                        person_ids_lists.append('')
                
                households_enhanced.insert(1, 'Person_Count', person_counts)
                households_enhanced.insert(2, 'Person_IDs', person_ids_lists)
                
                # Reorder columns
                important_cols = ['household_id', 'Person_Count', 'Person_IDs', 'household_type', 
                                'total_persons', 'count_adult', 'count_youth', 'count_child_hoh', 
                                'count_child_hh', 'youth']
                other_cols = [col for col in households_enhanced.columns if col not in important_cols]
                households_enhanced = households_enhanced[important_cols + other_cols]
            else:
                households_enhanced = pd.DataFrame()
            
            processed_data_with_ids[source_name] = {
                'persons': persons_enhanced,
                'households': households_enhanced
            }
        
        return processed_data_with_ids
    
    def create_download_interface(self):
        """Create the download interface."""
        st.header("Download Reports")
        
        if not self.calculated_reports:
            st.info("No reports available for download. Please complete the previous steps first.")
            return
        
        # Display available reports
        st.subheader("Available Reports")
        
        total_reports = sum(len(reports) for reports in self.calculated_reports.values())
        st.write(f"**Total Reports:** {total_reports}")
        
        # Show report summary
        summary_data = []
        for report_type, reports in self.calculated_reports.items():
            for report_name, report_df in reports.items():
                summary_data.append({
                    'Report Type': report_type,
                    'Report Name': report_name,
                    'Rows': len(report_df),
                    'Columns': len(report_df.columns)
                })
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            st.dataframe(summary_df, use_container_width=True)
        
        # Download options
        st.subheader("Download Options")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            # Download type selection
            download_type = st.radio(
                "Select Download Type",
                ["Reports Only", 
                 "Reports + Raw Data with IDs", 
                 "Reports + Processed Data with IDs",
                 "All Data (Reports + Raw + Processed)"],
                help="Choose what data to include in the download"
            )
        
        with col2:
            st.write("**File Information:**")
            timezone = get_timezone_for_region(self.region)
            timestamp = get_current_timestamp(timezone)
            filename = create_download_filename(self.region, "Reports", timestamp)
            st.write(f"Filename: `{filename}`")
            st.write(f"Timestamp: {timestamp}")
        
        # Generate and download
        if st.button("Generate Download File", type="primary"):
            self.generate_download_file(
                filename, 
                include_raw_data=(download_type in ["Reports + Raw Data with IDs", "All Data (Reports + Raw + Processed)"]),
                include_processed_data=(download_type in ["Reports + Processed Data with IDs", "All Data (Reports + Raw + Processed)"])
            )
    
    def generate_download_file(self, filename: str, include_raw_data: bool = False, 
                             include_processed_data: bool = False):
        """Generate and provide download file."""
        with st.spinner("Generating Excel file..."):
            try:
                # Prepare raw data with IDs if requested
                raw_data_with_ids = None
                if include_raw_data:
                    raw_data_with_ids = self.prepare_raw_data_with_ids()
                
                # Prepare processed data if requested
                processed_data_with_ids = None
                if include_processed_data:
                    processed_data_with_ids = self.prepare_processed_data_with_ids()
                
                # Create Excel file
                excel_buffer = self.exporter.create_excel_file(
                    self.calculated_reports,
                    include_raw_data=include_raw_data,
                    raw_data_with_ids=raw_data_with_ids,
                    include_processed_data=include_processed_data,
                    processed_data_with_ids=processed_data_with_ids
                )
                
                # Calculate file size
                file_size_mb = len(excel_buffer.getvalue()) / 1024 / 1024
                
                # Provide download
                st.download_button(
                    label=f"📥 Download Excel File ({file_size_mb:.1f} MB)",
                    data=excel_buffer.getvalue(),
                    file_name=filename,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    type="primary",
                    use_container_width=True
                )
                
                st.success("Excel file generated successfully!")
                
                # Show file details
                with st.expander("File Details"):
                    st.write(f"**Filename:** {filename}")
                    st.write(f"**File Size:** {file_size_mb:.2f} MB")
                    st.write(f"**Report Sheets:** {len(self.calculated_reports)}")
                    
                    for sheet_name, sheet_data in self.calculated_reports.items():
                        st.write(f"- {sheet_name}: {len(sheet_data)} reports")
                    
                    if include_raw_data and raw_data_with_ids:
                        st.write(f"**Raw Data Sheets:** {len(raw_data_with_ids)}")
                        for source_name, raw_df in raw_data_with_ids.items():
                            st.write(f"- Raw_{source_name}: {len(raw_df)} households")
                    
                    if include_processed_data and processed_data_with_ids:
                        st.write(f"**Processed Data Sheets:** {len(processed_data_with_ids) * 2}")
                        for source_name, data_dict in processed_data_with_ids.items():
                            if 'persons' in data_dict:
                                st.write(f"- Processed_Persons_{source_name}: {len(data_dict['persons'])} persons")
                            if 'households' in data_dict:
                                st.write(f"- Processed_Households_{source_name}: {len(data_dict['households'])} households")
                
            except Exception as e:
                st.error(f"Error generating Excel file: {str(e)}")
    
    def create_individual_downloads(self):
        """Create interface for downloading individual reports."""
        st.subheader("Individual Report Downloads")
        
        if not self.calculated_reports:
            st.info("No reports available.")
            return
        
        # Select report for individual download
        report_options = []
        for report_type, reports in self.calculated_reports.items():
            for report_name in reports.keys():
                report_options.append(f"{report_type} - {report_name}")
        
        if not report_options:
            st.info("No individual reports available.")
            return
        
        selected_report = st.selectbox(
            "Select Report for Individual Download",
            report_options
        )
        
        if selected_report:
            # Parse selection
            report_type, report_name = selected_report.split(" - ", 1)
            report_df = self.calculated_reports[report_type][report_name]
            
            # Show preview
            with st.expander(f"Preview: {selected_report}"):
                st.dataframe(report_df, use_container_width=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                # CSV download
                csv_data = report_df.to_csv()
                csv_filename = f"{self.region}_{report_type}_{report_name.replace(' ', '_')}.csv"
                
                st.download_button(
                    label="📄 Download as CSV",
                    data=csv_data,
                    file_name=csv_filename,
                    mime='text/csv'
                )
            
            with col2:
                # Excel download (single sheet)
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    report_df.to_excel(writer, sheet_name=report_name[:30], index=True)
                
                excel_filename = f"{self.region}_{report_type}_{report_name.replace(' ', '_')}.xlsx"
                
                st.download_button(
                    label="📊 Download as Excel",
                    data=excel_buffer.getvalue(),
                    file_name=excel_filename,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
    
    def create_raw_data_downloads(self):
        """Create interface for downloading enhanced raw data with IDs."""
        st.subheader("Enhanced Raw Data Downloads")
        
        if not self.uploaded_data:
            st.info("No raw data available.")
            return
        
        st.info("Raw data downloads include Household ID and Person IDs for verification purposes.")
        
        # Prepare raw data with IDs
        raw_data_with_ids = self.prepare_raw_data_with_ids()
        
        if not raw_data_with_ids:
            st.warning("No processed data available to generate IDs.")
            return
        
        # Show available sources
        for source_name, enhanced_df in raw_data_with_ids.items():
            st.write(f"### {source_name}")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Households", len(enhanced_df))
            
            with col2:
                # Count total persons from Person_IDs column
                total_persons = sum(len(pid.split(', ')) for pid in enhanced_df['Person_IDs'])
                st.metric("Total Persons", total_persons)
            
            with col3:
                # Download button
                csv_data = enhanced_df.to_csv(index=False)
                filename = f"{self.region}_{source_name}_Raw_with_IDs.csv"
                
                st.download_button(
                    label="📥 Download CSV",
                    data=csv_data,
                    file_name=filename,
                    mime='text/csv',
                    key=f"raw_{source_name}"
                )
            
            # Preview
            with st.expander(f"Preview first 10 rows"):
                # Show only key columns in preview
                preview_cols = ['Household_ID', 'Person_IDs', 'Timestamp', 'Gender', 'Age Range', 'Race/Ethnicity']
                display_cols = [col for col in preview_cols if col in enhanced_df.columns]
                st.dataframe(enhanced_df[display_cols].head(10), use_container_width=True)
    
    def create_processed_data_downloads(self):
        """Create interface for downloading processed persons and households data."""
        st.subheader("Processed Data Downloads")
        
        if not self.processed_data:
            st.info("No processed data available.")
            return
        
        st.info("Download the fully processed persons and households data with all calculated fields.")
        
        # Prepare processed data
        processed_data_with_ids = self.prepare_processed_data_with_ids()
        
        if not processed_data_with_ids:
            st.warning("No processed data available.")
            return
        
        # Show available sources
        for source_name, data_dict in processed_data_with_ids.items():
            st.write(f"### {source_name}")
            
            # Create tabs for persons and households
            tab1, tab2 = st.tabs(["👥 Persons Data", "🏠 Households Data"])
            
            with tab1:
                if 'persons' in data_dict and not data_dict['persons'].empty:
                    persons_df = data_dict['persons']
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Total Persons", len(persons_df))
                    
                    with col2:
                        st.metric("Unique Households", persons_df['Household_ID'].nunique())
                    
                    with col3:
                        # Download button
                        csv_data = persons_df.to_csv(index=False)
                        filename = f"{self.region}_{source_name}_Processed_Persons.csv"
                        
                        st.download_button(
                            label="📥 Download Persons CSV",
                            data=csv_data,
                            file_name=filename,
                            mime='text/csv',
                            key=f"processed_persons_{source_name}"
                        )
                    
                    # Preview
                    with st.expander("Preview first 20 rows"):
                        st.dataframe(persons_df.head(20), use_container_width=True)
                else:
                    st.info("No persons data available for this source.")
            
            with tab2:
                if 'households' in data_dict and not data_dict['households'].empty:
                    households_df = data_dict['households']
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Total Households", len(households_df))
                    
                    with col2:
                        avg_size = households_df['total_persons'].mean()
                        st.metric("Avg Household Size", f"{avg_size:.1f}")
                    
                    with col3:
                        # Download button
                        csv_data = households_df.to_csv(index=False)
                        filename = f"{self.region}_{source_name}_Processed_Households.csv"
                        
                        st.download_button(
                            label="📥 Download Households CSV",
                            data=csv_data,
                            file_name=filename,
                            mime='text/csv',
                            key=f"processed_households_{source_name}"
                        )
                    
                    # Preview
                    with st.expander("Preview first 20 rows"):
                        st.dataframe(households_df.head(20), use_container_width=True)
                else:
                    st.info("No households data available for this source.")
    
    def create_additional_resources(self):
        """Create section with additional resources and links."""
        st.subheader("Additional Resources")
        
        st.write("**Related Tools:**")
        st.markdown("- [HMIS and Non-HMIS Combiner](https://combinepit.streamlit.app/)")
        st.markdown("- [Breakdown by Project Name on HIC](https://hic-project.streamlit.app/)")
        
        st.write("**Support:**")
        st.write("For technical support or questions about the PIT Count application, please contact your system administrator.")

def create_download_interface():
    """Create the complete download interface."""
    
    if not SessionManager.has_calculated_reports():
        st.info("Please complete the previous steps to access downloads.")
        return
    
    # Update current step
    SessionManager.update_current_step('download')
    
    # Initialize download manager
    download_manager = DownloadManager()
    
    # Create main download interface
    download_manager.create_download_interface()
    
    st.write("---")
    
    # Individual downloads
    download_manager.create_individual_downloads()
    
    st.write("---")
    
    # Raw data downloads
    download_manager.create_raw_data_downloads()
    
    st.write("---")
    
    # Processed data downloads
    download_manager.create_processed_data_downloads()
    
    st.write("---")
    
    # Additional resources
    download_manager.create_additional_resources()