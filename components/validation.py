"""
Data validation functionality for PIT Count application.
Validates age ranges, gender, and race/ethnicity selections.
"""

import streamlit as st
import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Set
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from utils.session import SessionManager

class DataValidator:
    """Validates data entries for correct age, gender, and race values."""
    
    # Valid values for validation
    VALID_AGE_RANGES = ['Under 18', '18-24', '25-34', '35-44', '45-54', '55-64', '65+']
    
    # Updated gender values as per requirements
    VALID_GENDERS = [
        'Woman (Girl if child)',
        'Man (Boy if child)',
        'Culturally Specific Identity',
        'Non-Binary',
        'Transgender',
        'Questioning',
        'Different Identity'
    ]
    
    # Updated race values as per requirements
    VALID_RACES = [
        'White',
        'Black/African American/African',
        'Asian/Asian American',
        'Indigenous (American Indian/Alaska Native/Indigenous)',
        'Native Hawaiian/Pacific Islander',
        'Middle Eastern/North African',
        'Hispanic/Latina/e/o'
    ]
    
    def __init__(self, data: pd.DataFrame, source_name: str, region: str):
        """Initialize validator with data."""
        self.data = data.copy()
        self.source_name = source_name
        self.region = region
        self.validation_results = []
    
    def validate_column(self, column_name: str, valid_values: List[str], 
                       allow_multiple: bool = False) -> Tuple[pd.DataFrame, int]:
        """Validate a single column against valid values."""
        
        if column_name not in self.data.columns:
            return pd.DataFrame(), 0
        
        invalid_rows = []
        
        for idx, value in self.data[column_name].items():
            # Skip null/empty values
            if pd.isna(value) or str(value).strip() == '':
                continue
            
            value_str = str(value).strip()
            
            if allow_multiple:
                # For multi-select columns (like Race/Ethnicity)
                # Split by comma and check each value
                selected_values = [v.strip() for v in value_str.split(',')]
                invalid_selections = [v for v in selected_values if v and v not in valid_values]
                
                if invalid_selections:
                    invalid_rows.append({
                        'Row': idx + 2,  # Excel row number
                        'Column': column_name,
                        'Value': value_str,
                        'Invalid_Parts': ', '.join(invalid_selections),
                        'Valid_Options': ', '.join(valid_values[:5]) + '...' if len(valid_values) > 5 else ', '.join(valid_values)
                    })
            else:
                # For single-select columns
                if value_str not in valid_values:
                    invalid_rows.append({
                        'Row': idx + 2,  # Excel row number
                        'Column': column_name,
                        'Value': value_str,
                        'Invalid_Parts': value_str,
                        'Valid_Options': ', '.join(valid_values[:5]) + '...' if len(valid_values) > 5 else ', '.join(valid_values)
                    })
        
        return pd.DataFrame(invalid_rows), len(invalid_rows)
    
    def validate_all_columns(self) -> Dict[str, pd.DataFrame]:
        """Validate all age, gender, and race columns in the dataset."""
        
        validation_results = {}
        
        # Columns to validate
        columns_to_validate = []
        
        # Age Range columns
        age_columns = [
            'Age Range',
            'Adult/Parent #2: Age Range',
            'Adult/Parent #3: Age Range',
            'Adult/Parent #4: Age Range'
        ]
        
        # Gender columns
        gender_columns = [
            'Gender',
            'Adult/Parent #2: Gender',
            'Adult/Parent #3: Gender',
            'Adult/Parent #4: Gender',
            'Child #1: Gender',
            'Child #2: Gender',
            'Child #3: Gender',
            'Child #4: Gender',
            'Child #5: Gender',
            'Child #6: Gender'
        ]
        
        # Race/Ethnicity columns
        race_columns = [
            'Race/Ethnicity',
            'Adult/Parent #2: Race/Ethnicity',
            'Adult/Parent #3: Race/Ethnicity',
            'Adult/Parent #4: Race/Ethnicity',
            'Child #1: Race/Ethnicity',
            'Child #2: Race/Ethnicity',
            'Child #3: Race/Ethnicity',
            'Child #4: Race/Ethnicity',
            'Child #5: Race/Ethnicity',
            'Child #6: Race/Ethnicity'
        ]
        
        # Validate Age Range columns
        for col in age_columns:
            if col in self.data.columns:
                invalid_df, count = self.validate_column(col, self.VALID_AGE_RANGES, allow_multiple=False)
                if count > 0:
                    validation_results[f"age_{col}"] = invalid_df
        
        # Validate Gender columns
        for col in gender_columns:
            if col in self.data.columns:
                invalid_df, count = self.validate_column(col, self.VALID_GENDERS, allow_multiple=True)
                if count > 0:
                    validation_results[f"gender_{col}"] = invalid_df
        
        # Validate Race/Ethnicity columns
        for col in race_columns:
            if col in self.data.columns:
                invalid_df, count = self.validate_column(col, self.VALID_RACES, allow_multiple=True)
                if count > 0:
                    validation_results[f"race_{col}"] = invalid_df
        
        return validation_results
    
    def create_validation_excel(self, validation_results: Dict[str, pd.DataFrame]) -> BytesIO:
        """Create Excel file with validation issues highlighted."""
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Validation Summary"
        
        # Add headers
        headers = ["Issue Type", "Column", "Invalid Count", "Sample Invalid Values"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add summary data
        row = 2
        for key, df in validation_results.items():
            if not df.empty:
                issue_type = "Age Range" if key.startswith("age_") else "Gender" if key.startswith("gender_") else "Race/Ethnicity"
                column_name = df['Column'].iloc[0] if not df.empty else key
                invalid_count = len(df)
                sample_values = ', '.join(df['Value'].head(3).astype(str).tolist())
                
                ws_summary.cell(row=row, column=1, value=issue_type)
                ws_summary.cell(row=row, column=2, value=column_name)
                ws_summary.cell(row=row, column=3, value=invalid_count)
                ws_summary.cell(row=row, column=4, value=sample_values)
                row += 1
        
        # Auto-adjust column widths
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # Raw data sheet with highlights
        ws_data = wb.create_sheet("Raw Data with Validation")
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        invalid_fills = {
            'age': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Light red
            'gender': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Light yellow
            'race': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        }
        
        # Write headers
        headers = list(self.data.columns)
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Track invalid cells
        invalid_cells = {}
        for key, df in validation_results.items():
            for _, invalid_row in df.iterrows():
                row_num = invalid_row['Row']
                col_name = invalid_row['Column']
                
                if row_num not in invalid_cells:
                    invalid_cells[row_num] = {}
                
                if key.startswith('age_'):
                    invalid_cells[row_num][col_name] = 'age'
                elif key.startswith('gender_'):
                    invalid_cells[row_num][col_name] = 'gender'
                else:
                    invalid_cells[row_num][col_name] = 'race'
        
        # Write data with highlights
        for idx, row in self.data.iterrows():
            excel_row = idx + 2
            
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws_data.cell(row=excel_row, column=col_idx, value=value)
                
                # Apply highlight if this cell has validation issues
                if excel_row in invalid_cells and col_name in invalid_cells[excel_row]:
                    issue_type = invalid_cells[excel_row][col_name]
                    cell.fill = invalid_fills[issue_type]
        
        # Add legend
        legend_row = len(self.data) + 4
        ws_data.cell(row=legend_row, column=1, value="LEGEND:").font = Font(bold=True)
        
        legend_items = [
            ("Age Range Issues", invalid_fills['age']),
            ("Gender Issues", invalid_fills['gender']),
            ("Race/Ethnicity Issues", invalid_fills['race'])
        ]
        
        for i, (label, fill) in enumerate(legend_items):
            row = legend_row + i + 1
            cell = ws_data.cell(row=row, column=1, value=label)
            cell.fill = fill
            ws_data.cell(row=row, column=2, value=f"Invalid {label.split()[0]} selection")
        
        # Detailed issues sheet
        ws_details = wb.create_sheet("Validation Details")
        
        # Combine all validation results
        all_issues = []
        for key, df in validation_results.items():
            if not df.empty:
                issue_type = "Age Range" if key.startswith("age_") else "Gender" if key.startswith("gender_") else "Race/Ethnicity"
                df['Issue_Type'] = issue_type
                all_issues.append(df)
        
        if all_issues:
            combined_df = pd.concat(all_issues, ignore_index=True)
            
            # Write headers
            headers = ['Row', 'Issue_Type', 'Column', 'Value', 'Invalid_Parts', 'Valid_Options']
            for col, header in enumerate(headers, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Write data
            for idx, row in combined_df.iterrows():
                for col_idx, value in enumerate(row[headers], 1):
                    ws_details.cell(row=idx + 2, column=col_idx, value=value)
            
            # Auto-adjust columns
            for column in ws_details.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_details.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer


def create_validation_interface(uploaded_data: Dict[str, pd.DataFrame], region: str):
    """Create the data validation interface."""
    
    st.subheader("✅ Data Validation")
    
    st.info("""
    **Validation Checks:**
    - **Age Range**: Must be one of: Under 18, 18-24, 25-34, 35-44, 45-54, 55-64, 65+
    - **Gender**: Must match predefined gender categories (multiple selections allowed)
    - **Race/Ethnicity**: Must match predefined race categories (multiple selections allowed)
    
    Invalid entries will be highlighted in the Excel export.
    """)
    
    if st.button("🔍 Run Validation", type="primary"):
        all_validation_results = {}
        
        for source_name, df in uploaded_data.items():
            with st.spinner(f"Validating {source_name}..."):
                validator = DataValidator(df, source_name, region)
                validation_results = validator.validate_all_columns()
                
                if validation_results:
                    all_validation_results[source_name] = {
                        'validator': validator,
                        'results': validation_results
                    }
        
        # Store results in session
        st.session_state['validation_results'] = all_validation_results
        
        if all_validation_results:
            st.warning("⚠️ Validation issues found!")
        else:
            st.success("✅ All data entries are valid!")
    
    # Display results if available
    if 'validation_results' in st.session_state:
        results = st.session_state['validation_results']
        
        if results:
            st.write("---")
            st.subheader("📊 Validation Results")
            
            # Summary statistics
            total_issues = sum(
                sum(len(df) for df in r['results'].values())
                for r in results.values()
            )
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Issues", f"{total_issues:,}")
            
            with col2:
                age_issues = sum(
                    sum(len(df) for k, df in r['results'].items() if k.startswith('age_'))
                    for r in results.values()
                )
                st.metric("Age Range Issues", f"{age_issues:,}")
            
            with col3:
                gender_issues = sum(
                    sum(len(df) for k, df in r['results'].items() if k.startswith('gender_'))
                    for r in results.values()
                )
                st.metric("Gender Issues", f"{gender_issues:,}")
            
            # Results by source
            for source_name, result in results.items():
                validator = result['validator']
                validation_results = result['results']
                
                source_issues = sum(len(df) for df in validation_results.values())
                
                with st.expander(f"📋 {source_name} - {source_issues} validation issues found", expanded=True):
                    
                    if validation_results:
                        # Group by issue type
                        age_issues = {k: v for k, v in validation_results.items() if k.startswith('age_')}
                        gender_issues = {k: v for k, v in validation_results.items() if k.startswith('gender_')}
                        race_issues = {k: v for k, v in validation_results.items() if k.startswith('race_')}
                        
                        # Display summaries
                        if age_issues:
                            st.write("**Age Range Issues:**")
                            for key, df in age_issues.items():
                                st.write(f"- {df['Column'].iloc[0]}: {len(df)} invalid entries")
                        
                        if gender_issues:
                            st.write("**Gender Issues:**")
                            for key, df in gender_issues.items():
                                st.write(f"- {df['Column'].iloc[0]}: {len(df)} invalid entries")
                        
                        if race_issues:
                            st.write("**Race/Ethnicity Issues:**")
                            for key, df in race_issues.items():
                                st.write(f"- {df['Column'].iloc[0]}: {len(df)} invalid entries")
                        
                        # Show sample issues
                        st.write("**Sample Issues:**")
                        all_issues = []
                        for df in validation_results.values():
                            if not df.empty:
                                all_issues.append(df.head(5))
                        
                        if all_issues:
                            sample_df = pd.concat(all_issues, ignore_index=True).head(20)
                            st.dataframe(
                                sample_df[['Row', 'Column', 'Value', 'Invalid_Parts']],
                                use_container_width=True
                            )
                        
                        # Download option
                        try:
                            excel_buffer = validator.create_validation_excel(validation_results)
                            st.download_button(
                                label="📥 Download Validation Report (Excel)",
                                data=excel_buffer.getvalue(),
                                file_name=f"{source_name}_validation_report.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"val_excel_{source_name}"
                            )
                        except Exception as e:
                            st.error(f"Error creating Excel file: {str(e)}")
                    else:
                        st.success("✅ No validation issues found!")
        else:
            st.success("✅ All data entries are valid across all sources!")