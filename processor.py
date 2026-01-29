"""
Data processing module for PIT Count Application
Contains all data transformation and processing logic
"""

import pandas as pd
import numpy as np
import streamlit as st
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Set, Any, Union
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from config import (
    UNIFIED_COLUMN_MAPPINGS, REGION_SIGNATURES, CONDITION_MAPPING, AGE_RANGES,
    SEX_CATEGORIES, GENDER_CATEGORIES, RACE_CATEGORIES, HOUSEHOLD_CATEGORIES,
    VALID_AGE_RANGES, VALID_SEX, VALID_GENDERS, VALID_RACES
)

def process_pit_data(df: pd.DataFrame, source_name: str, region: str = None) -> Dict[str, pd.DataFrame]:
    """
    Main processing function that applies all transformations with auto-detection.

    Args:
        df: Input DataFrame with raw column names
        source_name: Name of the data source (ES, TH, Unsheltered)
        region: Optional region name. If None, will auto-detect from column presence.

    Returns:
        dict with 'persons_df', 'households_df', and 'raw_df'
    """
    import streamlit as st
    from utils import detect_region_and_format, validate_name_fields_completeness, handle_mixed_format_data

    # Copy original for reference
    raw_df = df.copy()

    # Auto-detect region if not provided
    if region is None:
        detection_result = detect_region_and_format(df)

        if detection_result['region'] is None:
            st.error(f"❌ Could not automatically detect data format. {detection_result.get('error', '')}")
            raise ValueError(f"Unable to detect data format: {detection_result.get('error', '')}")

        region = detection_result['region']

        # Display detection results
        st.success(f"✅ Auto-detected: {detection_result['detected_format']}")
        st.info(f"📍 Timezone: {detection_result['timezone']}")
        st.info(f"👥 Max adults: {detection_result['max_adults']}")
        st.info(f"🔤 Name format: {detection_result['name_format']}")

        if detection_result['missing_optional']:
            st.info(f"ℹ️ Optional data not found: {', '.join(detection_result['missing_optional'])}")

        # Store detected region in session state for later use
        st.session_state['region'] = region
        st.session_state['timezone'] = detection_result['timezone']

    # Step 1: Apply unified column mappings
    df_mapped, mapping_log = apply_column_mapping(df, UNIFIED_COLUMN_MAPPINGS, log_mapping=True)

    # Step 2: Handle mixed format data (synthesize missing fields if possible)
    df_mapped = handle_mixed_format_data(df_mapped, mapping_log)

    # Step 3: Count age groups
    df_mapped = count_age_groups(df_mapped)

    # Step 4: Classify household types
    df_mapped = classify_household_type(df_mapped)

    # Step 5: Flatten dataset (create person-level records)
    flattened_df = flatten_entire_dataset(df_mapped)

    # Step 6: Flag chronically homeless
    flattened_df = flag_chronically_homeless(flattened_df)

    # Step 7: Add age group column
    flattened_df = add_age_group_column(flattened_df)

    # Step 8: Process race
    flattened_df = process_race(flattened_df)

    # Step 9: Process sex (required field)
    flattened_df = process_sex(flattened_df)

    # Step 10: Process gender (optional field)
    flattened_df = process_gender(flattened_df)

    # Step 11: Standardize conditions
    flattened_df = standardize_conditions(flattened_df)

    # Add source column
    flattened_df['source'] = source_name

    # Create households summary
    households_df = create_households_summary(flattened_df)

    return {
        'persons_df': flattened_df,
        'households_df': households_df,
        'raw_df': raw_df
    }

def apply_column_mapping(df: pd.DataFrame, unified_mapping: Dict[str, List[Tuple[str, int]]],
                        log_mapping: bool = True) -> Tuple[pd.DataFrame, Dict[str, Dict]]:
    """
    Apply column mappings with priority-based source selection.

    Args:
        df: Input DataFrame with raw column names
        unified_mapping: Unified mapping structure with priorities
            Format: {'target_col': [('source_col_1', priority_100), ('source_col_2', priority_90), ...]}
        log_mapping: Whether to log which source columns were selected

    Returns:
        Tuple of (mapped_df, mapping_log)
        mapping_log shows which source column was used for each target
    """
    import streamlit as st

    # Strip whitespace and remove duplicates
    df.columns = df.columns.str.strip()
    df = df.loc[:, ~df.columns.duplicated(keep='first')]

    # Track which mappings we actually used
    mapping_log = {}
    selected_columns = {}  # source -> target

    # For each target column, select the highest-priority source that exists
    for target_col, source_options in unified_mapping.items():
        # Sort by priority (descending)
        sorted_options = sorted(source_options, key=lambda x: x[1], reverse=True)

        # Find first available source column
        for source_col, priority in sorted_options:
            if source_col in df.columns:
                selected_columns[source_col] = target_col
                mapping_log[target_col] = {
                    'source': source_col,
                    'priority': priority,
                    'alternatives_available': len([s for s, p in source_options if s in df.columns]) - 1
                }
                break
        else:
            # No source column found - this is OK for optional fields
            mapping_log[target_col] = {
                'source': None,
                'priority': None,
                'alternatives_available': 0
            }

    # Check if we have any valid mappings
    if not selected_columns:
        # Get a sample of expected and found columns for better debugging
        expected_sample = list(unified_mapping.keys())[:10]
        found_sample = list(df.columns)[:10]
        st.error(f"No valid columns found after applying unified mapping.\n\n"
                f"**Expected columns (sample):** {', '.join(expected_sample)}\n\n"
                f"**Found columns (sample):** {', '.join(found_sample)}\n\n"
                f"Please verify your data format matches New England or Great Lakes format.")
        return pd.DataFrame(), mapping_log

    # Select and rename columns
    df_mapped = df[list(selected_columns.keys())].copy()
    df_mapped.rename(columns=selected_columns, inplace=True)

    # Log mapping decisions if requested
    if log_mapping:
        st.info(f"📋 Mapped {len(selected_columns)} columns from source data")

        # Show detailed mapping in expander
        with st.expander("🔍 View Column Mapping Details"):
            log_data = []
            for target, info in mapping_log.items():
                if info['source'] is not None:
                    log_data.append({
                        'Target Column': target,
                        'Source Column': info['source'],
                        'Priority': info['priority'],
                        'Alternatives': info['alternatives_available']
                    })

            if log_data:
                log_df = pd.DataFrame(log_data)
                st.dataframe(log_df, width='stretch')

    return df_mapped, mapping_log

def count_age_groups(df: pd.DataFrame) -> pd.DataFrame:
    """Count the number of adults, youth, and children in each household"""
    
    # Initialize count columns
    age_group_columns = ['count_adult', 'count_youth', 'count_child_hoh', 'count_child_hh']
    for column in age_group_columns:
        df[column] = 0
    
    # Define age categories
    adult_ages = ['25-34', '35-44', '45-54', '55-64', '65+']
    youth_ages = ['18-24']
    child_age = ['Under 18']
    
    # Check age columns for adults
    age_related_cols = ['age_range', 'adult_2_age_range', 'adult_3_age_range', 'adult_4_age_range']
    
    for col in age_related_cols:
        if col in df.columns:
            df[col] = df[col].fillna('')
            df['count_adult'] += df[col].isin(adult_ages).astype(int)
            df['count_youth'] += df[col].isin(youth_ages).astype(int)
            df['count_child_hoh'] += df[col].isin(child_age).astype(int)
    
    # Check child indicators
    child_related_cols = [f'child_{i}' for i in range(1, 7)]
    
    for col in child_related_cols:
        if col in df.columns:
            df[col] = df[col].fillna('No')
            df['count_child_hh'] += (df[col] == 'Yes').astype(int)
    
    # Calculate total persons and youth flag
    df['total_person_in_household'] = df['count_adult'] + df['count_youth'] + df['count_child_hoh'] + df['count_child_hh']
    df['youth'] = df['count_adult'].apply(lambda x: 'Yes' if x == 0 else 'No')
    
    return df

def classify_household_type(df: pd.DataFrame) -> pd.DataFrame:
    """Classify households based on composition"""
    
    # Check required columns
    required_columns = ['count_adult', 'count_youth', 'count_child_hh', 'count_child_hoh']
    if not all(column in df.columns for column in required_columns):
        raise ValueError(f"Missing required columns for household classification")
    
    # Define conditions
    has_adults_or_youth = df['count_adult'] + df['count_youth'] > 0
    has_children = df['count_child_hh'] > 0
    only_children = df['count_child_hoh'] > 0
    
    # Set up conditions and choices
    conditions = [
        has_adults_or_youth & has_children,
        has_adults_or_youth & ~has_children,
        only_children
    ]
    choices = ['Household with Children', 'Household without Children', 'Household with Only Children']
    
    # Apply classification
    df['household_type'] = np.select(conditions, choices, default='Unknown')
    
    return df

def flatten_entire_dataset(df: pd.DataFrame) -> pd.DataFrame:
    """Transform household-based data into person-based data"""

    # Reset index and create Household_ID
    df.reset_index(drop=True, inplace=True)
    df['Household_ID'] = df.index + 1
    # Store original row index for traceability (Excel row = index + 2 for header)
    df['_Source_Row_Number'] = df.index + 2

    def create_member(row, member_type, member_number):
        """Create a member record"""
        # Set prefix based on member type
        if member_type == 'Child':
            prefix = f'child_{member_number}_'
        elif member_number != 1:
            prefix = f'adult_{member_number}_'
        else:
            prefix = ''
        
        # Member attributes to extract
        member_attrs = [
            # Demographics
            'Sex', 'Gender', 'Race/Ethnicity', 'age_range', 'dob', 'age',
            # Name fields (region-specific, some may be None)
            'first_initial', 'last_initial', 'last_third',  # NE format
            'first_name', 'first_letter_last',  # GL format
            # Status
            'DV', 'vet', 'chronic_condition', 'disability',
            # Homelessness history
            'first_time', 'homeless_long', 'homeless_long_this_time',
            'homeless_times', 'homeless_total',
            'specific_homeless_long_this_time', 'specific_homeless_long'
        ]
        
        # Household attributes to include
        household_attrs = [
            'count_adult', 'count_youth', 'count_child_hoh',
            'count_child_hh', 'total_person_in_household',
            'household_type', 'youth',
            # Filter columns (preserve for report filtering)
            'Location: General', 'Project Name on HIC', 'County', 'AHS District'
        ]
        
        # Initialize member dictionary
        member = {
            'Household_ID': row['Household_ID'],
            'Source_Row_Number': row.get('_Source_Row_Number', row['Household_ID'] + 1),
            'Member_Type': member_type,
            'Member_Number': member_number
        }
        
        # Check if member exists FIRST (has Sex or Race data - required fields)
        sex_col = f'{prefix}Sex'
        race_col = f'{prefix}Race/Ethnicity'
        sex_val = row.get(sex_col)
        race_val = row.get(race_col)

        # Early return if member doesn't exist (performance optimization)
        if not (pd.notna(sex_val) or pd.notna(race_val)):
            return None

        # Add member attributes (only if member exists)
        for attr in member_attrs:
            col_name = f'{prefix}{attr}'
            member[attr] = row.get(col_name, None)

        # Add household attributes
        for attr in household_attrs:
            member[attr] = row.get(attr)

        return member
    
    # Determine which adult slots exist in the data (for optional adult_3, adult_4 support)
    adult_slots = [1]  # Adult 1 always exists
    for i in range(2, 5):  # Check adults 2-4
        # Check if any key columns for this adult exist
        if f'adult_{i}_age_range' in df.columns or f'adult_{i}_Sex' in df.columns:
            adult_slots.append(i)

    # Create flattened list using list comprehension (faster than append in loop)
    members = []

    # Use itertuples() for better performance (faster than to_dict('records'))
    # We need to map tuple values back to column names for compatibility
    columns = df.columns.tolist()

    for row_tuple in df.itertuples(index=False, name=None):
        # Create dict from tuple values and column names
        row = dict(zip(columns, row_tuple))

        # Process adults (only those that exist in the data)
        for i in adult_slots:
            member = create_member(row, 'Adult', i)
            if member:
                members.append(member)

        # Process children (up to 6 - checked dynamically by create_member)
        for i in range(1, 7):
            member = create_member(row, 'Child', i)
            if member:
                members.append(member)

    # Create DataFrame once from list (faster than incremental building)
    flattened_df = pd.DataFrame(members)

    # Assign unique Person_ID to each person for traceability
    if not flattened_df.empty:
        flattened_df.insert(0, 'Person_ID', [f'P{i}' for i in range(1, len(flattened_df) + 1)])

    return flattened_df

def flag_chronically_homeless(df: pd.DataFrame) -> pd.DataFrame:
    """Flag chronically homeless individuals based on criteria with reason tracking"""

    required_columns = ['homeless_long', 'first_time', 'homeless_long_this_time',
                       'homeless_times', 'homeless_total', 'disability']

    # Check for required columns
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        st.warning(f"Missing columns for chronic homelessness calculation: {', '.join(missing_cols)}. Setting all CH to 'No'.")
        df['CH'] = 'No'
        df['CH_Reason'] = ''
        return df

    # Fill NaN values to prevent silent failures in conditions
    df['homeless_long'] = df['homeless_long'].fillna('')
    df['first_time'] = df['first_time'].fillna('')
    df['homeless_long_this_time'] = df['homeless_long_this_time'].fillna('')
    df['homeless_times'] = df['homeless_times'].fillna('')
    df['homeless_total'] = df['homeless_total'].fillna('')
    df['disability'] = df['disability'].fillna('No')

    # Initialize CH_Reason column for traceability
    df['CH_Reason'] = ''

    # Define chronic homelessness conditions
    has_disability = df['disability'] == 'Yes'
    cond1 = (df['homeless_long'] == 'One year or more') & (df['first_time'] == 'Yes')
    cond2 = (df['homeless_long_this_time'] == 'One year or more') & (df['first_time'] == 'No')
    cond3 = ((df['first_time'] == 'No') &
             (df['homeless_long_this_time'] == 'Less than one year') &
             (df['homeless_times'] == '4 or more times') &
             (df['homeless_total'] == '12 months or more'))

    # Track reasons for CH flag (applied in priority order)
    df.loc[cond1 & has_disability, 'CH_Reason'] = 'First time homeless + 1yr+ duration + disability'
    df.loc[cond2 & has_disability & (df['CH_Reason'] == ''), 'CH_Reason'] = 'Returning + 1yr+ this time + disability'
    df.loc[cond3 & has_disability & (df['CH_Reason'] == ''), 'CH_Reason'] = 'Returning + <1yr + 4+ episodes + 12mo+ total + disability'

    # Apply conditions with disability requirement
    chronic_homeless_condition = cond1 | cond2 | cond3
    df['CH'] = np.where(chronic_homeless_condition & has_disability, 'Yes', 'No')

    return df

def add_age_group_column(df: pd.DataFrame) -> pd.DataFrame:
    """Add age_group column categorizing individuals"""
    
    if 'age_range' not in df.columns:
        raise ValueError("'age_range' column is missing")
    
    # Define age range mappings
    age_ranges = {
        'adult': ['25-34', '35-44', '45-54', '55-64', '65+', '25-59', '60+'],
        'youth': ['18-24'],
        'child': ['Under 18']
    }
    
    # Create mapping dictionary
    age_range_to_group = {}
    for group, ranges in age_ranges.items():
        for age_range in ranges:
            age_range_to_group[age_range] = group
    
    # Apply mapping
    df['age_group'] = df['age_range'].map(age_range_to_group).fillna('unknown')
    
    return df

def process_race(df: pd.DataFrame) -> pd.DataFrame:
    """Process race/ethnicity column"""
    
    if 'Race/Ethnicity' not in df.columns:
        raise ValueError("'Race/Ethnicity' column is missing")
    
    def categorize_race(race_ethnicity):
        """Categorize race/ethnicity based on rules"""
        if pd.isna(race_ethnicity):
            return 'Unknown'

        # Ensure race_ethnicity is string before split
        race_ethnicity_str = str(race_ethnicity) if not isinstance(race_ethnicity, str) else race_ethnicity
        selected_races = race_ethnicity_str.split(', ')
        hispanic_selected = "Hispanic/Latina/e/o" in selected_races
        
        # Only Hispanic selected
        if hispanic_selected and len(selected_races) == 1:
            return "Hispanic/Latina/e/o"

        if hispanic_selected:
            # Use list comprehension to remove all instances (more robust than .remove())
            selected_races = [r for r in selected_races if r != "Hispanic/Latina/e/o"]
        
        if len(selected_races) > 1:
            return "Multi-Racial & Hispanic/Latina/e/o" if hispanic_selected else "Multi-Racial (not Hispanic/Latina/e/o)"
        elif selected_races:
            return f"{selected_races[0]} & Hispanic/Latina/e/o" if hispanic_selected else selected_races[0]
        else:
            return "Unknown"
    
    # Apply categorization
    df['race'] = df['Race/Ethnicity'].apply(categorize_race)
    df.drop('Race/Ethnicity', axis=1, inplace=True)
    
    return df

def process_sex(df: pd.DataFrame) -> pd.DataFrame:
    """Process sex column - required binary field"""
    if 'Sex' not in df.columns:
        df['Sex'] = 'Not Reported'
    # Standardize values
    df['Sex'] = df['Sex'].fillna('Not Reported')
    return df

def process_gender(df: pd.DataFrame) -> pd.DataFrame:
    """Process gender column to add gender_count (Gender is now optional)"""

    if 'Gender' not in df.columns:
        df['Gender'] = 'Not Reported'
        df['gender_count'] = 'unknown'
        return df

    def count_gender(gender):
        """Count number of gender selections"""
        if pd.isna(gender):
            return 'unknown'
        # Ensure gender is string before split
        gender_str = str(gender) if not isinstance(gender, str) else gender

        # Check if it's "Not Reported" or empty
        if not gender_str or gender_str.strip() == '' or gender_str == 'Not Reported':
            return 'unknown'

        # Split by comma and count non-empty parts
        parts = [p.strip() for p in gender_str.split(',') if p.strip()]
        return 'one' if len(parts) == 1 else 'more'

    df['gender_count'] = df['Gender'].apply(count_gender)

    return df

def standardize_conditions(df: pd.DataFrame) -> pd.DataFrame:
    """Standardize chronic condition values"""

    def map_conditions(conditions, mapping):
        """Map conditions to standardized forms"""
        try:
            if isinstance(conditions, str):
                return ', '.join(mapping.get(condition.strip(), condition.strip())
                               for condition in conditions.split(','))
            return conditions
        except (AttributeError, TypeError):
            return conditions

    if 'chronic_condition' in df.columns:
        try:
            df['chronic_condition'] = df['chronic_condition'].apply(
                lambda x: map_conditions(x, CONDITION_MAPPING)
            )
        except Exception as e:
            import streamlit as st
            st.warning(f"Warning: Could not standardize all conditions: {str(e)}")

    return df

def create_households_summary(persons_df: pd.DataFrame) -> pd.DataFrame:
    """Create household-level summary from person-level data"""

    # Validate critical columns exist
    critical_columns = ['Household_ID', 'household_type']
    missing_critical = [col for col in critical_columns if col not in persons_df.columns]
    if missing_critical:
        raise ValueError(f"Missing critical columns for household summary: {missing_critical}")

    # Get unique households
    unique_households = persons_df.drop_duplicates(subset='Household_ID')

    # Select household-level columns
    household_columns = [
        'Household_ID', 'household_type', 'total_person_in_household',
        'count_adult', 'count_youth', 'count_child_hoh', 'count_child_hh',
        'youth', 'source'
    ]

    # Filter columns that exist
    existing_columns = [col for col in household_columns if col in unique_households.columns]

    # Ensure critical columns are in existing columns (defensive check)
    required = ['Household_ID', 'household_type']
    missing_required = [col for col in required if col not in existing_columns]
    if missing_required:
        import streamlit as st
        st.error(f"Critical columns missing for household summary: {', '.join(missing_required)}")
        return pd.DataFrame()

    households_df = unique_households[existing_columns].copy()
    
    # Rename Household_ID to household_id for consistency
    households_df.rename(columns={'Household_ID': 'household_id'}, inplace=True)
    
    # Add total_persons column (same as total_person_in_household)
    if 'total_person_in_household' in households_df.columns:
        households_df['total_persons'] = households_df['total_person_in_household']
    
    return households_df

# Duplication Detection Functions
def map_name_columns_for_duplication(df: pd.DataFrame, region: str = None) -> pd.DataFrame:
    """
    Lightweight column mapping that only maps name fields for duplication detection.

    This function applies column mapping to convert raw column names to standardized
    names expected by the DuplicationDetector, without performing full data processing.

    Args:
        df: Input DataFrame with raw column names
        region: Region name (optional, will auto-detect if not provided)

    Returns:
        DataFrame with name columns mapped to standardized names
    """
    from utils import detect_region_and_format

    # Auto-detect region if not provided
    if region is None:
        detection_result = detect_region_and_format(df)
        region = detection_result.get('region')

    # Create a subset mapping with only name-related fields
    name_field_keys = [
        'first_initial', 'last_initial', 'last_third',  # New England
        'first_name', 'first_letter_last', 'last_name',  # Great Lakes
        'dob', 'age', 'age_range'  # Age-related for duplication logic
    ]

    # Filter UNIFIED_COLUMN_MAPPINGS to only include name fields
    name_mappings = {k: v for k, v in UNIFIED_COLUMN_MAPPINGS.items() if k in name_field_keys}

    # Apply column mapping (without logging to avoid cluttering UI)
    df_mapped, _ = apply_column_mapping(df, name_mappings, log_mapping=False)

    # Preserve all unmapped columns from the original dataframe
    # This ensures we don't lose any data that might be needed
    for col in df.columns:
        if col not in df_mapped.columns:
            df_mapped[col] = df[col]

    return df_mapped

def detect_duplicates(df: pd.DataFrame, source_name: str, region: str) -> pd.DataFrame:
    """
    Detect potential duplicates in the data using hierarchical matching.
    
    The detection uses a comprehensive priority-based system:
    1. Likely (High confidence): Full name + DOB, Initials + DOB, Full name + exact age
    2. Somewhat Likely (Medium): Initials + exact age, Full name + age range
    3. Possible (Low): Initials + age range
    
    Args:
        df: Input DataFrame with PIT count data
        source_name: Name of the data source (e.g., 'Sheltered_ES')
        region: Region name for determining name field format
        
    Returns:
        DataFrame with added columns:
        - Duplication_Score: Category of duplication likelihood
        - Duplication_Reason: Explanation of the match
        - Duplicates_With: Comma-separated list of matching row indices
    """
    detector = DuplicationDetector(df, source_name, region)
    return detector.annotate()

class DuplicationDetector:
    """
    Handles duplication detection logic with hierarchical matching.
    
    Features:
    - Region-specific name field handling
    - Multiple confidence levels for matches
    - Prevents downgrading of high-confidence matches
    - Supports various date formats
    - Creates Excel exports with color-coding
    """
    
    def __init__(self, data: pd.DataFrame, source_name: str, region: str):
        self.data = data.copy().reset_index(drop=True)
        self.source_name = source_name
        self.region = region
        self._prepare_name_fields()

        # Track matches
        self.high_confidence_matches = set()
        self.medium_confidence_matches = set()

        # Validate region matches detected format
        self._validate_region_format()
    
    def _safe_str(self, value: Any) -> str:
        """Convert to uppercase string safely"""
        if pd.isna(value) or value is None:
            return ""
        s = str(value).strip()
        return s.upper() if s else ""
    
    def _safe_int(self, value: Any) -> Optional[int]:
        """Convert to integer safely"""
        try:
            if pd.isna(value) or value is None:
                return None
            return int(value)
        except (ValueError, TypeError):
            return None
    
    def parse_dob(self, dob: Any) -> Optional[datetime]:
        """
        Parse date of birth from various formats.
        
        Supports formats:
        - YYYY-MM-DD, MM/DD/YYYY, MM-DD-YYYY
        - YYYY/MM/DD, DD/MM/YYYY, DD-MM-YYYY
        - MM/DD/YY, DD/MM/YY, YYYYMMDD
        """
        if pd.isna(dob) or dob is None or str(dob).strip() == "":
            return None
        
        s = str(dob).strip()
        date_formats = [
            "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y",
            "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y",
            "%m/%d/%y", "%d/%m/%y", "%Y%m%d"
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None
    
    def _prepare_name_fields(self):
        """
        Prepare name fields based on available columns (auto-detected format).

        Supports five formats:
        1. New England Complete: 1st letter first name + 1st & 3rd letters last name (3-letter code)
        2. New England Partial: 1st letter first name + 1st letter last name (2-letter code)
        3. Great Lakes Option A: First Name + First Letter of Last Name (hybrid)
        4. Great Lakes Option B: First Name + Last Name (full names)
        5. Fallback: Use whatever is available, warn if incomplete
        """
        df = self.data

        # Detect which name format is present in the data
        has_ne_complete = all(col in df.columns for col in ['first_initial', 'last_initial', 'last_third'])
        has_ne_partial = ('first_initial' in df.columns and 'last_initial' in df.columns and not has_ne_complete)
        # GL can use either 'first_letter_last' (new mapping) or 'last_initial' (legacy)
        has_gl_option_a = ('first_name' in df.columns and
                          ('first_letter_last' in df.columns or 'last_initial' in df.columns) and
                          'last_third' not in df.columns)  # GL uses first_letter_last without last_third
        has_gl_option_b = 'first_name' in df.columns and 'last_name' in df.columns
        has_first_name = 'first_name' in df.columns
        has_last_initial = 'last_initial' in df.columns or 'first_letter_last' in df.columns

        if has_ne_complete:
            # Format 1: New England Complete - 3-letter code (e.g., "ABC")
            df["_p1"] = df["first_initial"].apply(self._safe_str)
            df["_p2"] = df["last_initial"].apply(self._safe_str)
            df["_p3"] = df["last_third"].apply(self._safe_str)
            df["_full_name"] = df.apply(
                lambda r: "".join([r["_p1"], r["_p2"], r["_p3"]]),
                axis=1
            )
            df["_initials"] = df.apply(
                lambda r: "".join([r["_p1"][:1], r["_p2"][:1]]),
                axis=1
            )

        elif has_ne_partial:
            # Format 2: New England Partial - 2-letter code (e.g., "AB")
            # 3rd letter of last name not collected for all clients
            df["_p1"] = df["first_initial"].apply(self._safe_str)
            df["_p2"] = df["last_initial"].apply(self._safe_str)
            df["_p3"] = ""  # Not available
            df["_full_name"] = df.apply(
                lambda r: "".join([r["_p1"], r["_p2"]]),
                axis=1
            )
            df["_initials"] = df.apply(
                lambda r: "".join([r["_p1"][:1], r["_p2"][:1]]),
                axis=1
            )

        elif has_gl_option_b:
            # Format 3: Great Lakes Option B - "FirstName LastName" (full names)
            # Check this before Option A since last_name is more complete
            df["_p1"] = df["first_name"].apply(self._safe_str)
            df["_p2"] = df["last_name"].apply(self._safe_str)
            df["_full_name"] = df.apply(
                lambda r: " ".join(p for p in (r["_p1"], r["_p2"]) if p),
                axis=1
            )
            df["_initials"] = df.apply(
                lambda r: "".join([r["_p1"][:1] if r["_p1"] else "", r["_p2"][:1] if r["_p2"] else ""]),
                axis=1
            )

        elif has_gl_option_a:
            # Format 4: Great Lakes Option A - "FirstName L" (first name + last initial)
            df["_p1"] = df["first_name"].apply(self._safe_str)
            # Use first_letter_last if available, otherwise fall back to last_initial
            last_col = 'first_letter_last' if 'first_letter_last' in df.columns else 'last_initial'
            df["_p2"] = df[last_col].apply(self._safe_str)
            df["_full_name"] = df.apply(
                lambda r: " ".join(p for p in (r["_p1"], r["_p2"]) if p),
                axis=1
            )
            df["_initials"] = df.apply(
                lambda r: "".join([r["_p1"][:1] if r["_p1"] else "", r["_p2"][:1] if r["_p2"] else ""]),
                axis=1
            )

        elif has_first_name and has_last_initial:
            # Format 5: Partial - First name + last initial (synthesized from mixed format)
            df["_p1"] = df["first_name"].apply(self._safe_str)
            # Use first_letter_last if available, otherwise fall back to last_initial
            last_col = 'first_letter_last' if 'first_letter_last' in df.columns else 'last_initial'
            df["_p2"] = df[last_col].apply(self._safe_str)
            df["_full_name"] = df.apply(
                lambda r: " ".join(p for p in (r["_p1"], r["_p2"]) if p),
                axis=1
            )
            df["_initials"] = df.apply(
                lambda r: "".join([r["_p1"][:1] if r["_p1"] else "", r["_p2"][:1] if r["_p2"] else ""]),
                axis=1
            )

        else:
            # Format 6: Fallback - No name data
            # Records will be marked as "No Name" and highlighted in purple
            # No warning needed - this is handled gracefully
            df["_p1"] = ""
            df["_p2"] = ""
            df["_p3"] = ""
            df["_full_name"] = ""
            df["_initials"] = ""

        df["_is_no_name"] = df["_full_name"] == ""

    def _validate_region_format(self):
        """
        Validate that the selected region matches the detected name format.

        Displays a warning if mismatch is detected, which could indicate:
        - Wrong region selected by user
        - Data from different region uploaded
        - Missing expected columns

        This helps catch configuration errors early.

        Note: Does not warn if all records have no name data (handled gracefully with purple highlighting).
        """
        import streamlit as st

        # Check if all records have no name data - if so, skip validation
        # This is okay and will be handled with purple highlighting
        if (self.data["_full_name"] == "").all():
            return  # All records have no name - this is fine, skip validation

        # Detect which name formats are present
        ne_columns = ['first_initial', 'last_initial', 'last_third']
        has_any_ne_column = any(col in self.data.columns for col in ne_columns)

        has_first_name = 'first_name' in self.data.columns
        has_last_letter = 'first_letter_last' in self.data.columns
        has_last_name = 'last_name' in self.data.columns

        # Validate New England region - need at least ONE of the 3 columns
        if self.region == "New England":
            if not has_any_ne_column:
                st.warning(
                    f"⚠️ Region set to **New England** but name format doesn't match.\n\n"
                    f"Expected at least one of: `first_initial`, `last_initial`, `last_third`\n\n"
                    f"Duplication detection may not work as expected."
                )

        # Validate Great Lakes region - need first_name + (first_letter_last OR last_name)
        elif self.region == "Great Lakes":
            if not (has_first_name and (has_last_letter or has_last_name)):
                st.warning(
                    f"⚠️ Region set to **Great Lakes** but name format doesn't match.\n\n"
                    f"Expected columns: `first_name` + `first_letter_last` OR `first_name` + `last_name`\n\n"
                    f"Duplication detection may not work as expected."
                )

    def _compare_pair(self, i: int, j: int) -> Tuple[str, str]:
        """
        Compare two records for duplication using region-specific rules.

        Routes to region-specific comparison methods based on self.region:
        - New England: Simplified 3-rule hierarchy (no initials matching)
        - Great Lakes: Expanded 6-rule hierarchy (includes initials)
        - Other: Universal hierarchy (fallback)
        """
        r1, r2 = self.data.iloc[i], self.data.iloc[j]
        pair = (min(i, j), max(i, j))

        # Check for no name information - flag but don't match
        if not r1["_full_name"] or not r2["_full_name"]:
            return "Not Duplicate", ""

        # Route to region-specific comparison
        if self.region == "New England":
            return self._compare_pair_new_england(r1, r2, pair)
        elif self.region == "Great Lakes":
            return self._compare_pair_great_lakes(r1, r2, pair)
        else:
            # Fallback to universal hierarchy for unknown regions
            return self._compare_pair_universal(r1, r2, pair)

    def _compare_pair_new_england(self, r1, r2, pair: tuple) -> Tuple[str, str]:
        """
        New England duplication rules with contradictory evidence checking.

        Data Precision Hierarchy (most to least precise):
            DOB > Exact Age > Age Range

        Matching Rules (uses 3-letter codes: first_initial + last_initial + last_third):
        - Only falls back to less precise data when more precise data is unavailable
        - If DOB or Age exists but doesn't match, stop (contradictory evidence)
        - No separate initials matching (only full 3-letter code)

        Full Name (3-letter code) Matching:
            - Full Name + DOB match         → LIKELY
            - Full Name + DOB differs       → NOT DUPLICATE (contradicts)
            - Full Name + Age match         → SOMEWHAT LIKELY
            - Full Name + Age differs       → NOT DUPLICATE (contradicts)
            - Full Name + Age Range match   → POSSIBLE (only if no DOB/Age)
        """
        # Extract fields
        f1, f2 = r1["_full_name"], r2["_full_name"]
        dob1 = self.parse_dob(r1.get("dob") or r1.get("Date of Birth"))
        dob2 = self.parse_dob(r2.get("dob") or r2.get("Date of Birth"))
        age1 = self._safe_int(r1.get("age") or r1.get("Age"))
        age2 = self._safe_int(r2.get("age") or r2.get("Age"))
        rng1 = r1.get("age_range") or r1.get("Age Range")
        rng2 = r2.get("age_range") or r2.get("Age Range")

        # Determine what data is available
        has_dob = bool(dob1) and bool(dob2)
        has_age = (age1 is not None) and (age2 is not None)

        # Full name must match for any duplicate detection in New England
        if f1 != f2:
            return "Not Duplicate", ""

        # LIKELY: Full Name + DOB
        if has_dob and dob1 == dob2:
            self.high_confidence_matches.add(pair)
            return "Likely Duplicate 🔴", "Full name and DOB match"

        # Contradictory: DOBs exist but don't match
        if has_dob and dob1 != dob2:
            return "Not Duplicate", ""

        # SOMEWHAT LIKELY: Full Name + Exact Age
        if pair not in self.high_confidence_matches:
            if has_age and age1 == age2:
                self.medium_confidence_matches.add(pair)
                return "Somewhat Likely Duplicate 🟠", "Full name and exact age match"

        # Contradictory: Ages exist but don't match
        if has_age and age1 != age2:
            return "Not Duplicate", ""

        # POSSIBLE: Full Name + Age Range (only if no DOB/Age available)
        if pair not in self.high_confidence_matches and pair not in self.medium_confidence_matches:
            if rng1 and rng2 and rng1 == rng2:
                return "Possible Duplicate 🟡", "Full name and age range match"

        return "Not Duplicate", ""

    def _compare_pair_great_lakes(self, r1, r2, pair: tuple) -> Tuple[str, str]:
        """
        Great Lakes duplication rules with contradictory evidence checking.

        Data Precision Hierarchy (most to least precise):
            DOB > Exact Age > Age Range

        Matching Rules:
        - Only falls back to less precise data when more precise data is unavailable
        - If DOB or Age exists but doesn't match, stop (contradictory evidence)

        Full Name Matching:
            - Full Name + DOB match         → LIKELY
            - Full Name + DOB differs       → NOT DUPLICATE (contradicts)
            - Full Name + Age match         → LIKELY
            - Full Name + Age differs       → NOT DUPLICATE (contradicts)
            - Full Name + Age Range match   → SOMEWHAT LIKELY (only if no DOB/Age)

        Initials Matching (when full names don't match):
            - Initials + DOB match          → LIKELY
            - Initials + DOB differs        → NOT DUPLICATE (contradicts)
            - Initials + Age match          → SOMEWHAT LIKELY
            - Initials + Age differs        → NOT DUPLICATE (contradicts)
            - Initials + Age Range match    → POSSIBLE (only if no DOB/Age)
        """
        # Extract fields
        f1, f2 = r1["_full_name"], r2["_full_name"]
        init1, init2 = r1["_initials"], r2["_initials"]
        dob1 = self.parse_dob(r1.get("dob") or r1.get("Date of Birth"))
        dob2 = self.parse_dob(r2.get("dob") or r2.get("Date of Birth"))
        age1 = self._safe_int(r1.get("age") or r1.get("Age"))
        age2 = self._safe_int(r2.get("age") or r2.get("Age"))
        rng1 = r1.get("age_range") or r1.get("Age Range")
        rng2 = r2.get("age_range") or r2.get("Age Range")

        # Determine what data is available for comparison
        has_dob = bool(dob1) and bool(dob2)
        has_age = (age1 is not None) and (age2 is not None)

        # === FULL NAME MATCHING ===
        if f1 == f2:
            # LIKELY: Full Name + DOB
            if has_dob and dob1 == dob2:
                self.high_confidence_matches.add(pair)
                return "Likely Duplicate 🔴", "Full name and DOB match"

            # Contradictory: DOBs exist but don't match
            if has_dob and dob1 != dob2:
                return "Not Duplicate", ""

            # LIKELY: Full Name + Exact Age
            if has_age and age1 == age2:
                self.high_confidence_matches.add(pair)
                return "Likely Duplicate 🔴", "Full name and exact age match"

            # Contradictory: Ages exist but don't match
            if has_age and age1 != age2:
                return "Not Duplicate", ""

            # SOMEWHAT LIKELY: Full Name + Age Range (only if no DOB/Age available)
            if pair not in self.high_confidence_matches:
                if rng1 and rng2 and rng1 == rng2:
                    self.medium_confidence_matches.add(pair)
                    return "Somewhat Likely Duplicate 🟠", "Full name and age range match"

            return "Not Duplicate", ""

        # === INITIALS MATCHING (when full names don't match) ===
        if init1 == init2:
            # LIKELY: Initials + DOB
            if has_dob and dob1 == dob2:
                self.high_confidence_matches.add(pair)
                return "Likely Duplicate 🔴", "Initials and DOB match"

            # Contradictory: DOBs exist but don't match
            if has_dob and dob1 != dob2:
                return "Not Duplicate", ""

            # SOMEWHAT LIKELY: Initials + Exact Age
            if pair not in self.high_confidence_matches:
                if has_age and age1 == age2:
                    self.medium_confidence_matches.add(pair)
                    return "Somewhat Likely Duplicate 🟠", "Initials and exact age match"

            # Contradictory: Ages exist but don't match
            if has_age and age1 != age2:
                return "Not Duplicate", ""

            # POSSIBLE: Initials + Age Range (only if no DOB/Age available)
            if pair not in self.high_confidence_matches and pair not in self.medium_confidence_matches:
                if rng1 and rng2 and rng1 == rng2:
                    return "Possible Duplicate 🟡", "Initials and age range match"

        return "Not Duplicate", ""

    def _compare_pair_universal(self, r1, r2, pair: tuple) -> Tuple[str, str]:
        """
        Universal duplication rules with contradictory evidence checking (fallback).

        Data Precision Hierarchy (most to least precise):
            DOB > Exact Age > Age Range

        Matching Rules:
        - Only falls back to less precise data when more precise data is unavailable
        - If DOB or Age exists but doesn't match, stop (contradictory evidence)

        Full Name Matching:
            - Full Name + DOB match         → LIKELY
            - Full Name + DOB differs       → NOT DUPLICATE (contradicts)
            - Full Name + Age match         → LIKELY
            - Full Name + Age differs       → NOT DUPLICATE (contradicts)
            - Full Name + Age Range match   → SOMEWHAT LIKELY (only if no DOB/Age)

        Initials Matching (when full names don't match):
            - Initials + DOB match          → LIKELY
            - Initials + DOB differs        → NOT DUPLICATE (contradicts)
            - Initials + Age match          → SOMEWHAT LIKELY
            - Initials + Age differs        → NOT DUPLICATE (contradicts)
            - Initials + Age Range match    → POSSIBLE (only if no DOB/Age)
        """
        # Extract fields
        f1, f2 = r1["_full_name"], r2["_full_name"]
        init1, init2 = r1["_initials"], r2["_initials"]
        dob1 = self.parse_dob(r1.get("dob") or r1.get("Date of Birth"))
        dob2 = self.parse_dob(r2.get("dob") or r2.get("Date of Birth"))
        age1 = self._safe_int(r1.get("age") or r1.get("Age"))
        age2 = self._safe_int(r2.get("age") or r2.get("Age"))
        rng1 = r1.get("age_range") or r1.get("Age Range")
        rng2 = r2.get("age_range") or r2.get("Age Range")

        # Determine what data is available for comparison
        has_dob = bool(dob1) and bool(dob2)
        has_age = (age1 is not None) and (age2 is not None)

        # === FULL NAME MATCHING ===
        if f1 == f2:
            # LIKELY: Full Name + DOB
            if has_dob and dob1 == dob2:
                self.high_confidence_matches.add(pair)
                return "Likely Duplicate 🔴", "Full name and DOB match"

            # Contradictory: DOBs exist but don't match
            if has_dob and dob1 != dob2:
                return "Not Duplicate", ""

            # LIKELY: Full Name + Exact Age
            if has_age and age1 == age2:
                self.high_confidence_matches.add(pair)
                return "Likely Duplicate 🔴", "Full name and exact age match"

            # Contradictory: Ages exist but don't match
            if has_age and age1 != age2:
                return "Not Duplicate", ""

            # SOMEWHAT LIKELY: Full Name + Age Range (only if no DOB/Age available)
            if pair not in self.high_confidence_matches:
                if rng1 and rng2 and rng1 == rng2:
                    self.medium_confidence_matches.add(pair)
                    return "Somewhat Likely Duplicate 🟠", "Full name and age range match"

            return "Not Duplicate", ""

        # === INITIALS MATCHING (when full names don't match) ===
        if init1 == init2:
            # LIKELY: Initials + DOB
            if has_dob and dob1 == dob2:
                self.high_confidence_matches.add(pair)
                return "Likely Duplicate 🔴", "Initials and DOB match"

            # Contradictory: DOBs exist but don't match
            if has_dob and dob1 != dob2:
                return "Not Duplicate", ""

            # SOMEWHAT LIKELY: Initials + Exact Age
            if pair not in self.high_confidence_matches:
                if has_age and age1 == age2:
                    self.medium_confidence_matches.add(pair)
                    return "Somewhat Likely Duplicate 🟠", "Initials and exact age match"

            # Contradictory: Ages exist but don't match
            if has_age and age1 != age2:
                return "Not Duplicate", ""

            # POSSIBLE: Initials + Age Range (only if no DOB/Age available)
            if pair not in self.high_confidence_matches and pair not in self.medium_confidence_matches:
                if rng1 and rng2 and rng1 == rng2:
                    return "Possible Duplicate 🟡", "Initials and age range match"

        return "Not Duplicate", ""

    def annotate(self) -> pd.DataFrame:
        """
        Annotate dataframe with duplication scores and validation columns.

        Returns DataFrame with columns:
        - Duplication_Score: Category of duplication likelihood
        - Duplication_Reason: Explanation of the match
        - Duplicates_With: Comma-separated list of matching row indices
        - Dup_Sex: Sex value of first duplicate partner (for manual verification)
        - Dup_Race/Ethnicity: Race/Ethnicity value of first duplicate partner (for manual verification)
        """
        n = len(self.data)
        best_match = {}
        partners = {}

        # Clear tracking sets
        self.high_confidence_matches.clear()
        self.medium_confidence_matches.clear()

        # Compare all pairs
        for i in range(n):
            if self.data.at[i, "_is_no_name"]:
                continue

            for j in range(i + 1, n):
                if self.data.at[j, "_is_no_name"]:
                    continue

                score, reason = self._compare_pair(i, j)

                if score != "Not Duplicate":
                    # Track partners
                    partners.setdefault(i, set()).add(j)
                    partners.setdefault(j, set()).add(i)

                    # Update best match
                    for idx in [i, j]:
                        prev_score, _ = best_match.get(idx, ("Not Duplicate", ""))
                        if self._score_priority(score) > self._score_priority(prev_score):
                            best_match[idx] = (score, reason)

        # Create output
        out = self.data.copy()
        out["Duplication_Score"] = "Not Duplicate"
        out["Duplication_Reason"] = ""
        out["Duplicates_With"] = ""
        out["Dup_Sex"] = ""
        out["Dup_Race/Ethnicity"] = ""

        for idx in range(n):
            if self.data.at[idx, "_is_no_name"]:
                out.at[idx, "Duplication_Score"] = "No name information provided 🟣"
                out.at[idx, "Duplication_Reason"] = "No name information provided"
            elif idx in best_match:
                score, reason = best_match[idx]
                out.at[idx, "Duplication_Score"] = score
                out.at[idx, "Duplication_Reason"] = reason
                out.at[idx, "Duplicates_With"] = ",".join(
                    str(p) for p in sorted(partners[idx])
                )

                # Add partner's Sex and Race for validation
                if idx in partners and partners[idx]:
                    # Get first partner (lowest row number) for comparison
                    partner_idx = min(partners[idx])

                    # Get partner's Sex
                    partner_sex = self.data.iloc[partner_idx].get("Sex", "")
                    out.at[idx, "Dup_Sex"] = partner_sex if partner_sex and str(partner_sex) != 'nan' else ""

                    # Get partner's Race/Ethnicity
                    partner_race = (self.data.iloc[partner_idx].get("Race/Ethnicity", "")
                                   or self.data.iloc[partner_idx].get("race", ""))
                    out.at[idx, "Dup_Race/Ethnicity"] = partner_race if partner_race and str(partner_race) != 'nan' else ""

        return out
    
    def _score_priority(self, score: str) -> int:
        """Get priority for score comparison"""
        priorities = {
            "Likely Duplicate 🔴": 4,
            "Somewhat Likely Duplicate 🟠": 3,
            "Possible Duplicate 🟡": 2,
            "No name information provided 🟣": 1,
            "Not Duplicate": 0,
        }
        return priorities.get(score, 0)
    
    def create_excel_with_highlights(self, annotated_df: pd.DataFrame) -> BytesIO:
        """
        Export to Excel with visual formatting.

        Features:
        - Preserves original data columns in original order
        - Adds duplication columns at the end
        - Color-coded rows based on duplication score
        - Formatted header row
        - Auto-adjusted column widths
        - Duplicates_With indices adjusted for Excel (zero-based → 1-based starting at row 2)
        """
        # Create merged dataframe: original data + duplication columns + validation columns
        # Get original columns (stored in self.data before name field preparation)
        original_cols = [col for col in self.data.columns if not col.startswith('_')]
        dup_cols = ['Duplication_Score', 'Duplication_Reason', 'Duplicates_With', 'Dup_Sex', 'Dup_Race/Ethnicity']

        # Reorder: original columns first, then duplication columns, then validation columns
        export_df = annotated_df[original_cols + dup_cols].copy()

        # Reset index to ensure continuous 0-based indexing for row number calculations
        export_df = export_df.reset_index(drop=True)

        wb = Workbook()
        ws = wb.active
        ws.title = self.source_name

        # Define colors for each duplication score
        score_colors = {
            "Likely Duplicate 🔴": "FF9999",  # Light red
            "Somewhat Likely Duplicate 🟠": "FFCC99",  # Light orange
            "Possible Duplicate 🟡": "FFFF99",  # Light yellow
            "No name information provided 🟣": "D8BFD8",  # Light purple
            "Not Duplicate": "FFFFFF",  # White
        }

        # Style header row
        for ci, col in enumerate(export_df.columns, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = PatternFill(start_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Add data with row highlighting
        # Get Duplication_Score column index
        score_col_idx = export_df.columns.get_loc("Duplication_Score") if "Duplication_Score" in export_df.columns else None

        for r_idx, record in enumerate(export_df.itertuples(index=False), start=2):
            row_score = None
            for ci, val in enumerate(record, start=1):
                col_name = export_df.columns[ci - 1]

                # Track duplication score
                if score_col_idx is not None and ci == score_col_idx + 1:
                    row_score = val

                # Adjust Duplicates_With indices for Excel
                if col_name == "Duplicates_With" and isinstance(val, str) and val:
                    shifted_indices = []
                    for part in val.split(","):
                        try:
                            idx = int(part)
                            # Bounds check: ensure index is valid
                            if 0 <= idx < len(export_df):
                                # Convert 0-based to Excel row (starting at 2)
                                shifted_indices.append(str(idx + 2))
                        except ValueError:
                            continue
                    ws.cell(row=r_idx, column=ci, value=",".join(shifted_indices))
                else:
                    ws.cell(row=r_idx, column=ci, value=val)

            # Apply row color based on duplication score
            color = score_colors.get(row_score, "FFFFFF") if row_score else "FFFFFF"
            fill = PatternFill(start_color=color, fill_type="solid")
            for ci in range(1, len(export_df.columns) + 1):
                ws.cell(row=r_idx, column=ci).fill = fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_length + 2, 50)

        # Save to buffer
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    
# Data Validation Functions
def validate_data(df: pd.DataFrame, source_name: str, region: str) -> Dict[str, pd.DataFrame]:
    """Validate data for age, gender, and race values"""
    
    validator = DataValidator(df, source_name, region)
    return validator.validate_all_columns()

class DataValidator:
    """Handles data validation logic"""
    
    def __init__(self, data: pd.DataFrame, source_name: str, region: str):
        self.data = data.copy()
        self.source_name = source_name
        self.region = region
    
    def validate_column(self, column_name: str, valid_values: List[str], 
                       allow_multiple: bool = False) -> Tuple[pd.DataFrame, int]:
        """Validate a single column against valid values"""
        
        if column_name not in self.data.columns:
            return pd.DataFrame(), 0
        
        invalid_rows = []
        
        for idx, value in self.data[column_name].items():
            # Skip null/empty values
            if pd.isna(value) or str(value).strip() == '':
                continue
            
            value_str = str(value).strip()
            
            if allow_multiple:
                # For multi-select columns
                selected_values = [v.strip() for v in value_str.split(',')]
                invalid_selections = [v for v in selected_values if v and v not in valid_values]
                
                if invalid_selections:
                    invalid_rows.append({
                        'Row': idx + 2,  # Excel row number
                        'Column': column_name,
                        'Value': value_str,
                        'Invalid_Parts': ', '.join(invalid_selections),
'Valid_Options': ', '.join(valid_values[:5]) + '...' if len(valid_values) > 5 else ', '.join(valid_values)
                    })
            else:
                # For single-select columns
                if value_str not in valid_values:
                    invalid_rows.append({
                        'Row': idx + 2,  # Excel row number
                        'Column': column_name,
                        'Value': value_str,
                        'Invalid_Parts': value_str,
                        'Valid_Options': ', '.join(valid_values[:5]) + '...' if len(valid_values) > 5 else ', '.join(valid_values)
                    })
        
        return pd.DataFrame(invalid_rows), len(invalid_rows)
    
    def validate_all_columns(self) -> Dict[str, pd.DataFrame]:
        """Validate all age, gender, and race columns"""
        
        validation_results = {}
        
        # Age Range columns
        age_columns = [
            'Age Range',
            'Adult/Parent #2: Age Range',
            'Adult/Parent #3: Age Range',
            'Adult/Parent #4: Age Range'
        ]
        
        # Sex columns (REQUIRED)
        sex_columns = [
            'Sex',
            'Adult/Parent #2: Sex',
            'Adult/Parent #3: Sex',
            'Adult/Parent #4: Sex',
            'Child #1: Sex',
            'Child #2: Sex',
            'Child #3: Sex',
            'Child #4: Sex',
            'Child #5: Sex',
            'Child #6: Sex'
        ]

        # Gender columns (OPTIONAL)
        gender_columns = [
            'Gender',
            'Adult/Parent #2: Gender',
            'Adult/Parent #3: Gender',
            'Adult/Parent #4: Gender',
            'Child #1: Gender',
            'Child #2: Gender',
            'Child #3: Gender',
            'Child #4: Gender',
            'Child #5: Gender',
            'Child #6: Gender'
        ]

        # Race/Ethnicity columns
        race_columns = [
            'Race/Ethnicity',
            'Adult/Parent #2: Race/Ethnicity',
            'Adult/Parent #3: Race/Ethnicity',
            'Adult/Parent #4: Race/Ethnicity',
            'Child #1: Race/Ethnicity',
            'Child #2: Race/Ethnicity',
            'Child #3: Race/Ethnicity',
            'Child #4: Race/Ethnicity',
            'Child #5: Race/Ethnicity',
            'Child #6: Race/Ethnicity'
        ]
        
        # Validate Age Range columns
        for col in age_columns:
            if col in self.data.columns:
                invalid_df, count = self.validate_column(col, VALID_AGE_RANGES, allow_multiple=False)
                if count > 0:
                    validation_results[f"age_{col}"] = invalid_df

        # Validate Sex columns (REQUIRED, single-select)
        for col in sex_columns:
            if col in self.data.columns:
                invalid_df, count = self.validate_column(col, VALID_SEX, allow_multiple=False)
                if count > 0:
                    validation_results[f"sex_{col}"] = invalid_df

        # Validate Gender columns (OPTIONAL, multi-select when provided)
        for col in gender_columns:
            if col in self.data.columns:
                invalid_df, count = self.validate_column(col, VALID_GENDERS, allow_multiple=True)
                if count > 0:
                    validation_results[f"gender_{col}"] = invalid_df
        
        # Validate Race/Ethnicity columns
        for col in race_columns:
            if col in self.data.columns:
                invalid_df, count = self.validate_column(col, VALID_RACES, allow_multiple=True)
                if count > 0:
                    validation_results[f"race_{col}"] = invalid_df

        return validation_results


# ============================================================================
# PIT COMBINER DATA PROCESSOR
# For combining HMIS (HUDX 230) and Non-HMIS (HDX) data
# ============================================================================

import logging
from pathlib import Path

combiner_logger = logging.getLogger(__name__)

class CombinerDataProcessor:
    """Handles all data processing operations for PIT Combiner."""

    def __init__(self, template_path: str = "template.xlsx"):
        self.template_path = template_path
        self._hmis_wb: Optional[Workbook] = None
        self._non_hmis_wb: Optional[Workbook] = None
        self._template_wb: Optional[Workbook] = None

    def clean_cell_value(self, value: Any) -> Union[float, int]:
        """Convert cell value to numeric type, handling various edge cases."""
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            if str(value).upper() == 'N/A':
                return 0
            return value
        if isinstance(value, str):
            cleaned = value.strip().replace(',', '').upper()
            if not cleaned or cleaned == 'N/A':
                return 0
            try:
                return float(cleaned)
            except ValueError:
                return 0
        return 0

    def should_delete_row(self, row: tuple, terms_to_delete: List[str]) -> bool:
        """Check if a row should be deleted based on forbidden terms."""
        for cell in row:
            if cell.value is None:
                continue
            cell_str = str(cell.value)
            # Preserve HUD instruction rows
            if "HUD does not allow" in cell_str:
                return False
            for term in terms_to_delete:
                if term in cell_str:
                    return True
        return False

    def clean_workbook(self, workbook: Workbook, sheets_to_clean: List[str],
                       terms_to_delete: List[str]) -> Dict[str, int]:
        """Clean specified sheets by removing rows with forbidden terms."""
        stats = {}
        for sheet_name in sheets_to_clean:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            rows_to_delete = []
            for row in sheet.iter_rows():
                if self.should_delete_row(row, terms_to_delete):
                    rows_to_delete.append(row[0].row)

            deleted = 0
            for row_idx in reversed(rows_to_delete):
                try:
                    sheet.delete_rows(row_idx)
                    deleted += 1
                except Exception as e:
                    combiner_logger.error(f"Error deleting row {row_idx} from {sheet_name}: {e}")
            stats[sheet_name] = deleted
            if deleted > 0:
                combiner_logger.info(f"Cleaned {deleted} rows from '{sheet_name}'")
        return stats

    def get_cell_value(self, sheet, col: str, row: int) -> Union[float, int]:
        """Safely get a numeric value from a cell."""
        try:
            value = sheet[f"{col}{row}"].value
            return self.clean_cell_value(value)
        except Exception as e:
            combiner_logger.debug(f"Error getting cell {col}{row}: {e}")
            return 0

    def parse_source_key(self, source_key: str) -> Tuple[str, str]:
        """Parse source key like 'hmis:Adult-Child' into (type, sheet_name)."""
        parts = source_key.split(':', 1)
        if len(parts) != 2:
            raise ValueError(f"Invalid source key format: {source_key}")
        return parts[0], parts[1]

    def get_workbook_for_source(self, source_type: str) -> Optional[Workbook]:
        """Get the appropriate workbook based on source type."""
        if source_type == 'hmis':
            return self._hmis_wb
        elif source_type == 'nonhmis':
            return self._non_hmis_wb
        return None

    def calculate_combined_value(self, source_ranges: List[Tuple[str, str, int, int]],
                                  row_offset: int) -> float:
        """Calculate combined value from multiple source ranges for a specific row."""
        total = 0.0
        for source_key, col, start_row, end_row in source_ranges:
            source_type, sheet_name = self.parse_source_key(source_key)
            workbook = self.get_workbook_for_source(source_type)

            if workbook is None:
                combiner_logger.warning(f"No workbook found for source type: {source_type}")
                continue

            if sheet_name not in workbook.sheetnames:
                combiner_logger.warning(f"Sheet '{sheet_name}' not found in {source_type} workbook")
                continue

            sheet = workbook[sheet_name]
            actual_row = start_row + row_offset

            if actual_row > end_row:
                continue

            value = self.get_cell_value(sheet, col, actual_row)
            total += value

        return total

    def process_range_specifications(self, range_specs: List[Any]) -> int:
        """Process all range specifications and update the template."""
        cells_updated = 0

        for spec in range_specs:
            source_ranges = spec.source_ranges
            target_sheet = spec.target_sheet
            target_col = spec.target_column
            target_start = spec.target_start_row

            if target_sheet not in self._template_wb.sheetnames:
                combiner_logger.warning(f"Target sheet '{target_sheet}' not found in template")
                continue

            output_sheet = self._template_wb[target_sheet]

            # Determine number of rows from source ranges
            if not source_ranges:
                continue

            # Use the first source range to determine row count
            _, _, start_row, end_row = source_ranges[0]
            num_rows = end_row - start_row + 1

            for row_offset in range(num_rows):
                combined_value = self.calculate_combined_value(source_ranges, row_offset)
                target_row = target_start + row_offset
                cell_ref = f"{target_col}{target_row}"

                try:
                    cell = output_sheet[cell_ref]
                    # Only update if cell doesn't contain a formula
                    if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell.value = combined_value
                        cells_updated += 1
                except Exception as e:
                    combiner_logger.error(f"Error updating cell {cell_ref}: {e}")

        combiner_logger.info(f"Updated {cells_updated} cells in template")
        return cells_updated

    def validate_workbooks(self, validation_rules: Dict) -> Tuple[bool, List[str]]:
        """Validate that all required sheets exist in the workbooks."""
        errors = []

        # Check HMIS sheets
        if self._hmis_wb:
            hmis_sheets = set(self._hmis_wb.sheetnames)
            required = set(validation_rules.get('required_sheets_hmis', []))
            missing = required - hmis_sheets
            if missing:
                errors.append(f"HMIS file missing sheets: {missing}")

        # Check Non-HMIS sheets
        if self._non_hmis_wb:
            non_hmis_sheets = set(self._non_hmis_wb.sheetnames)
            required = set(validation_rules.get('required_sheets_non_hmis', []))
            missing = required - non_hmis_sheets
            if missing:
                errors.append(f"Non-HMIS file missing sheets: {missing}")

        # Check Template sheets
        if self._template_wb:
            template_sheets = set(self._template_wb.sheetnames)
            required = set(validation_rules.get('required_sheets_template', []))
            missing = required - template_sheets
            if missing:
                errors.append(f"Template file missing sheets: {missing}")

        return len(errors) == 0, errors

    def process_and_combine(self, hmis_stream, non_hmis_stream,
                            range_specs: List[Any],
                            terms_to_delete: List[str],
                            validation_rules: Optional[Dict] = None) -> BytesIO:
        """Main processing function to combine HMIS and Non-HMIS data."""
        try:
            # Load HMIS workbook
            combiner_logger.info("Loading HMIS workbook...")
            hmis_stream.seek(0)
            self._hmis_wb = load_workbook(
                filename=BytesIO(hmis_stream.read()),
                data_only=True,
                read_only=False
            )

            # Load Non-HMIS workbook
            combiner_logger.info("Loading Non-HMIS workbook...")
            non_hmis_stream.seek(0)
            self._non_hmis_wb = load_workbook(
                filename=BytesIO(non_hmis_stream.read()),
                data_only=True,
                read_only=False
            )

            # Load template (preserve formulas)
            combiner_logger.info("Loading template workbook...")
            if not Path(self.template_path).exists():
                raise FileNotFoundError(f"Template file '{self.template_path}' not found")

            with open(self.template_path, "rb") as f:
                self._template_wb = load_workbook(
                    filename=BytesIO(f.read()),
                    data_only=False,
                    read_only=False
                )

            # Validate workbooks if rules provided
            if validation_rules:
                is_valid, errors = self.validate_workbooks(validation_rules)
                if not is_valid:
                    combiner_logger.warning(f"Validation warnings: {errors}")

            # Clean HMIS data (contains "Client Doesn't Know" etc.)
            combiner_logger.info("Cleaning HMIS data...")
            hmis_sheets = list(self._hmis_wb.sheetnames)
            self.clean_workbook(self._hmis_wb, hmis_sheets, terms_to_delete)

            # Process range specifications
            combiner_logger.info("Processing range specifications...")
            cells_updated = self.process_range_specifications(range_specs)

            # Save to BytesIO
            combiner_logger.info("Saving combined workbook...")
            output = BytesIO()
            self._template_wb.save(output)
            output.seek(0)

            combiner_logger.info(f"Processing complete. Cells updated: {cells_updated}")
            return output

        except Exception as e:
            combiner_logger.error(f"Error in process_and_combine: {e}")
            raise
        finally:
            # Clean up
            self._hmis_wb = None
            self._non_hmis_wb = None
            self._template_wb = None