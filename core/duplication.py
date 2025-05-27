import pandas as pd
from datetime import datetime
from typing import Optional, Set, Tuple, Dict, Any, List
from io import BytesIO

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from utils.session import SessionManager
from utils.helpers import format_number, calculate_percentage


class DuplicationScore:
    """Duplication likelihood categories with visual indicators."""
    LIKELY = "Likely Duplicate 🔴"
    SOMEWHAT_LIKELY = "Somewhat Likely Duplicate 🟠"
    POSSIBLE = "Possible Duplicate 🟡"
    NO_NAME = "No name information provided 🟣"
    NOT_DUPLICATE = "Not Duplicate"

    @classmethod
    def get_color(cls, score: str) -> str:
        """Return HEX color for a given score."""
        return {
            cls.LIKELY:        "FF9999",  # Light red
            cls.SOMEWHAT_LIKELY: "FFCC99",  # Light orange
            cls.POSSIBLE:      "FFFF99",  # Light yellow
            cls.NO_NAME:       "D8BFD8",  # Light purple
            cls.NOT_DUPLICATE: "FFFFFF",  # White
        }.get(score, "FFFFFF")

    @classmethod
    def get_description(cls, score: str) -> str:
        """Return user-friendly description for each score."""
        descriptions = {
            cls.LIKELY: "High confidence match - Same name and DOB/age",
            cls.SOMEWHAT_LIKELY: "Medium confidence match - Partial name with age match",
            cls.POSSIBLE: "Low confidence match - Partial name with age range match",
            cls.NO_NAME: "Cannot evaluate - Missing name information",
            cls.NOT_DUPLICATE: "No duplication detected"
        }
        return descriptions.get(score, "Unknown status")


class DuplicationDetector:
    """
    Detects duplicates in PIT Count data using hierarchical matching logic.
    
    The detector uses a priority-based system:
    1. Highest priority: Full name + DOB match
    2. High priority: Initials + DOB match or Full name + exact age
    3. Medium priority: Initials + exact age or Full name + age range
    4. Low priority: Initials + age range match
    
    Each row is annotated with:
      - Duplication_Score: The likelihood category
      - Duplication_Reason: Explanation of why it's flagged
      - Duplicates_With: Indices of matching records
    """

    def __init__(
        self,
        data: pd.DataFrame,
        source_name: str,
        region: str
    ) -> None:
        """
        Initialize the detector with data and regional settings.
        
        :param data: Raw upload data
        :param source_name: Name for sheet/file naming
        :param region: 'New England', 'Dashgreatlake', or other
        """
        self.data = data.copy().reset_index(drop=True)
        self.source_name = source_name
        self.region = region
        self._prepare_name_fields()

    def _safe_str(self, value: Any) -> str:
        """Convert to uppercase string, handling nulls safely."""
        if pd.isna(value) or value is None:
            return ""
        s = str(value).strip()
        return s.upper() if s else ""

    def _safe_int(self, value: Any) -> Optional[int]:
        """Convert to integer, returning None if invalid."""
        try:
            if pd.isna(value) or value is None:
                return None
            return int(value)
        except (ValueError, TypeError):
            return None

    def _safe_value(self, value: Any) -> Optional[str]:
        """Trim string or return None if blank."""
        if pd.isna(value) or value is None:
            return None
        s = str(value).strip()
        return s if s else None

    def parse_dob(self, dob: Any) -> Optional[datetime]:
        """
        Parse date of birth from various formats.
        Supports: YYYY-MM-DD, MM/DD/YYYY, MM-DD-YYYY, and more.
        """
        if pd.isna(dob) or dob is None or str(dob).strip() == "":
            return None
        
        s = str(dob).strip()
        # Common date formats to try
        date_formats = [
            "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y",
            "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y",
            "%m/%d/%y", "%d/%m/%y", "%Y%m%d"
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    def _prepare_name_fields(self) -> None:
        """
        Prepare name fields based on regional configuration.
        Creates: _full_name, _initials, and _is_no_name columns.
        """
        df = self.data
        
        if self.region == "New England":
            # New England uses: 1st letter of first, 1st letter of last, 3rd letter of last
            df["_p1"] = df.get("1st Letter of First Name", "").apply(self._safe_str)
            df["_p2"] = df.get("1st Letter of Last Name", "").apply(self._safe_str)
            df["_p3"] = df.get("3rd Letter of Last Name", "").apply(self._safe_str)
            df["_full_name"] = df.apply(
                lambda r: "".join([r["_p1"], r["_p2"], r["_p3"]]),
                axis=1
            )
            df["_initials"] = df.apply(
                lambda r: "".join([r["_p1"][:1], r["_p2"][:1]]),
                axis=1
            )
            
        elif self.region == "Dashgreatlake":
            # Dashgreatlake uses: Full first name, 1st letter of last
            df["_p1"] = df.get("First Name", "").apply(self._safe_str)
            df["_p2"] = df.get("First Letter of Last Name", "").apply(self._safe_str)
            df["_full_name"] = df.apply(
                lambda r: "".join([r["_p1"], r["_p2"]]),
                axis=1
            )
            df["_initials"] = df.apply(
                lambda r: "".join([r["_p1"][:1], r["_p2"][:1]]),
                axis=1
            )
            
        else:
            # Other regions use: Full first and last names
            df["_p1"] = df.get("First Name", "").apply(self._safe_str)
            df["_p2"] = df.get("Last Name", "").apply(self._safe_str)
            df["_full_name"] = df.apply(
                lambda r: " ".join(p for p in (r["_p1"], r["_p2"]) if p),
                axis=1
            )
            df["_initials"] = df.apply(
                lambda r: "".join(p[:1] for p in (r["_p1"], r["_p2"]) if p),
                axis=1
            )
            
        df["_is_no_name"] = df["_full_name"] == ""

    def _score_priority(self, score: str) -> int:
        """Get numeric priority for score comparison (higher = more likely duplicate)."""
        return {
            DuplicationScore.LIKELY: 4,
            DuplicationScore.SOMEWHAT_LIKELY: 3,
            DuplicationScore.POSSIBLE: 2,
            DuplicationScore.NO_NAME: 1,
            DuplicationScore.NOT_DUPLICATE: 0,
        }.get(score, 0)

    def _compare_pair(
        self,
        i: int,
        j: int,
        use_gender: bool,
        use_race: bool
    ) -> Tuple[str, str]:
        """
        Compare two records using hierarchical matching logic.
        
        Matching hierarchy:
        1. Full name + DOB → Likely
        2. Initials + DOB → Likely
        3. Full name + exact age → Likely
        4. Initials + exact age → Somewhat Likely
        5. Full name + age range → Somewhat Likely (only if no DOB/age)
        6. Initials + age range → Possible (only if no DOB/age)
        
        Gender/race matching appends to reason but doesn't affect score.
        """
        r1, r2 = self.data.iloc[i], self.data.iloc[j]
        
        # Extract comparison fields
        f1, f2 = r1["_full_name"], r2["_full_name"]
        if not f1 or not f2:
            return DuplicationScore.NOT_DUPLICATE, ""
            
        init1, init2 = r1["_initials"], r2["_initials"]
        dob1 = self.parse_dob(r1.get("Date of Birth"))
        dob2 = self.parse_dob(r2.get("Date of Birth"))
        age1 = self._safe_int(r1.get("Age"))
        age2 = self._safe_int(r2.get("Age"))
        rng1 = self._safe_value(r1.get("Age Range"))
        rng2 = self._safe_value(r2.get("Age Range"))
        gender1 = self._safe_value(r1.get("Gender"))
        gender2 = self._safe_value(r2.get("Gender"))
        race1 = self._safe_value(r1.get("Race/Ethnicity"))
        race2 = self._safe_value(r2.get("Race/Ethnicity"))

        def annotate_demographics(base_reason: str) -> str:
            """Append demographic matches to reason if enabled."""
            annotations = []
            if use_gender and gender1 and gender1 == gender2:
                annotations.append("same gender")
            if use_race and race1 and race1 == race2:
                annotations.append("same race")
            
            if annotations:
                return f"{base_reason} ({', '.join(annotations)})"
            return base_reason

        # 1) DOB-based matching (highest priority)
        if dob1 and dob2:
            if f1 == f2 and dob1 == dob2:
                return DuplicationScore.LIKELY, annotate_demographics(
                    "Full name and DOB match"
                )
            if init1 == init2 and dob1 == dob2:
                return DuplicationScore.LIKELY, annotate_demographics(
                    "Initials and DOB match"
                )

        # 2) Exact age matching
        if age1 is not None and age2 is not None:
            if f1 == f2 and age1 == age2:
                return DuplicationScore.LIKELY, annotate_demographics(
                    "Full name and age match"
                )
            if init1 == init2 and age1 == age2:
                return DuplicationScore.SOMEWHAT_LIKELY, annotate_demographics(
                    "Initials and age match"
                )

        # 3) Age range matching (only if no DOB and no exact age)
        if (not dob1 and not dob2 and 
            age1 is None and age2 is None and 
            rng1 and rng2):
            if f1 == f2 and rng1 == rng2:
                return DuplicationScore.SOMEWHAT_LIKELY, annotate_demographics(
                    "Full name and age range match"
                )
            if init1 == init2 and rng1 == rng2:
                return DuplicationScore.POSSIBLE, annotate_demographics(
                    "Initials and age range match"
                )

        return DuplicationScore.NOT_DUPLICATE, ""

    def annotate(
        self,
        use_gender: bool = True,
        use_race: bool = True
    ) -> pd.DataFrame:
        """
        Annotate every row with duplication information.
        
        Process:
        1. Compare all pairs of records
        2. Track highest-priority match for each record
        3. Build list of all matching partners
        4. Annotate each row with score, reason, and partner indices
        
        Returns DataFrame with three new columns:
        - Duplication_Score: Category of duplication likelihood
        - Duplication_Reason: Explanation of the match
        - Duplicates_With: Comma-separated list of matching row indices
        """
        n = len(self.data)
        best_match: Dict[int, Tuple[str, str]] = {}
        partners: Dict[int, Set[int]] = {}

        # Compare all pairs
        for i in range(n):
            if self.data.at[i, "_is_no_name"]:
                continue
                
            for j in range(i + 1, n):
                if self.data.at[j, "_is_no_name"]:
                    continue
                    
                score, reason = self._compare_pair(i, j, use_gender, use_race)
                
                if score != DuplicationScore.NOT_DUPLICATE:
                    # Track partners
                    partners.setdefault(i, set()).add(j)
                    partners.setdefault(j, set()).add(i)
                    
                    # Update best match if this is higher priority
                    for idx in [i, j]:
                        prev_score, _ = best_match.get(idx, (DuplicationScore.NOT_DUPLICATE, ""))
                        if self._score_priority(score) > self._score_priority(prev_score):
                            best_match[idx] = (score, reason)

        # Create output with annotations
        out = self.data.copy()
        out["Duplication_Score"] = DuplicationScore.NOT_DUPLICATE
        out["Duplication_Reason"] = ""
        out["Duplicates_With"] = ""

        for idx in range(n):
            if self.data.at[idx, "_is_no_name"]:
                out.at[idx, "Duplication_Score"] = DuplicationScore.NO_NAME
                out.at[idx, "Duplication_Reason"] = "No name information provided"
            elif idx in best_match:
                score, reason = best_match[idx]
                out.at[idx, "Duplication_Score"] = score
                out.at[idx, "Duplication_Reason"] = reason
                out.at[idx, "Duplicates_With"] = ",".join(
                    str(p) for p in sorted(partners[idx])
                )

        return out

    def create_excel_with_highlights(
        self,
        annotated_df: pd.DataFrame
    ) -> BytesIO:
        """
        Export to Excel with visual formatting.
        
        Features:
        - Color-coded rows based on duplication score
        - Formatted header row
        - Auto-adjusted column widths
        - Duplicates_With indices adjusted for Excel (zero-based → 1-based starting at row 2)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = self.source_name

        # Style header row
        for ci, col in enumerate(annotated_df.columns, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = PatternFill(start_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Add data with row highlighting
        for r_idx, record in enumerate(annotated_df.itertuples(index=False), start=2):
            for ci, val in enumerate(record, start=1):
                col_name = annotated_df.columns[ci - 1]
                
                # Adjust Duplicates_With indices for Excel
                if col_name == "Duplicates_With" and isinstance(val, str) and val:
                    shifted_indices = []
                    for part in val.split(","):
                        try:
                            # Convert 0-based to Excel row (starting at 2)
                            shifted_indices.append(str(int(part) + 2))
                        except ValueError:
                            continue
                    ws.cell(row=r_idx, column=ci, value=",".join(shifted_indices))
                else:
                    ws.cell(row=r_idx, column=ci, value=val)
                    
            # Apply row color based on duplication score
            score = getattr(record, "Duplication_Score")
            color = DuplicationScore.get_color(score)
            fill = PatternFill(start_color=color, fill_type="solid")
            for ci in range(1, len(annotated_df.columns) + 1):
                ws.cell(row=r_idx, column=ci).fill = fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_length + 2, 50)

        # Save to buffer
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def get_summary_stats(self, annotated_df: pd.DataFrame) -> Dict[str, Any]:
        """Calculate summary statistics for the detection results."""
        total = len(annotated_df)
        stats = {
            "total_records": total,
            "no_name": (annotated_df["Duplication_Score"] == DuplicationScore.NO_NAME).sum(),
            "likely": (annotated_df["Duplication_Score"] == DuplicationScore.LIKELY).sum(),
            "somewhat": (annotated_df["Duplication_Score"] == DuplicationScore.SOMEWHAT_LIKELY).sum(),
            "possible": (annotated_df["Duplication_Score"] == DuplicationScore.POSSIBLE).sum(),
            "not_duplicate": (annotated_df["Duplication_Score"] == DuplicationScore.NOT_DUPLICATE).sum()
        }
        
        # Calculate percentages
        for key in ["no_name", "likely", "somewhat", "possible", "not_duplicate"]:
            stats[f"{key}_pct"] = calculate_percentage(stats[key], total)
            
        # Total flagged (excluding NOT_DUPLICATE)
        stats["total_flagged"] = total - stats["not_duplicate"]
        stats["total_flagged_pct"] = calculate_percentage(stats["total_flagged"], total)
        
        return stats


def create_duplication_interface(
    uploaded_data: Dict[str, pd.DataFrame]
) -> None:
    """
    Streamlit UI for duplication detection with enhanced user guidance.
    """
    st.header("🔍 Duplication Detection")
    
    # Display region info
    region = SessionManager.get_session_value("region", "Unknown")
    st.info(f"**Region:** {region}")
    
    # Add expandable explanation
    with st.expander("ℹ️ How Duplication Detection Works", expanded=False):
        st.markdown("""
        ### Detection Logic
        
        The system uses a **hierarchical matching approach** to identify potential duplicates:
        
        1. **🔴 Likely Duplicates (High Confidence)**
           - Full name + Date of Birth match
           - Initials + Date of Birth match  
           - Full name + exact age match
        
        2. **🟠 Somewhat Likely (Medium Confidence)**
           - Initials + exact age match
           - Full name + age range match (only when DOB/age unavailable)
        
        3. **🟡 Possible Duplicates (Low Confidence)**
           - Initials + age range match (only when DOB/age unavailable)
        
        4. **🟣 No Name Information**
           - Records missing name data cannot be evaluated
        
        ### Regional Differences
        
        Different regions use different name formats:
        - **New England**: 1st letter of first name + 1st & 3rd letters of last name
        - **Dashgreatlake**: Full first name + 1st letter of last name
        - **Other regions**: Full first and last names
        
        ### Gender & Race Annotations
        
        When enabled, the system will note when matching records also share:
        - Same gender (e.g., "Full name and DOB match (same gender)")
        - Same race/ethnicity (e.g., "Full name and age match (same race)")
        
        **Important**: These demographic matches are informational only and do NOT affect whether records are flagged as duplicates.
        """)

    st.subheader("Detection Options")
    
    col1, col2 = st.columns(2)
    with col1:
        use_gender = st.checkbox(
            "Include gender in match descriptions",
            value=True,
            help="When enabled, adds '(same gender)' to reasons when applicable"
        )
    with col2:
        use_race = st.checkbox(
            "Include race/ethnicity in match descriptions",
            value=True,
            help="When enabled, adds '(same race)' to reasons when applicable"
        )

    if st.button("🚀 Run Duplication Detection", type="primary"):
        with st.spinner("Analyzing records for duplicates..."):
            results: Dict[str, Tuple[DuplicationDetector, pd.DataFrame]] = {}
            
            progress_bar = st.progress(0)
            for idx, (name, df) in enumerate(uploaded_data.items()):
                # Update progress
                progress = (idx + 1) / len(uploaded_data)
                progress_bar.progress(progress, f"Processing {name}...")
                
                # Run detection
                detector = DuplicationDetector(df, name, region)
                annotated = detector.annotate(use_gender, use_race)
                results[name] = (detector, annotated)
            
            progress_bar.empty()
            st.session_state["dup_results"] = results
            st.success("✅ Detection complete!")

    # Display results if available
    if "dup_results" in st.session_state:
        st.divider()
        
        for name, (detector, annotated) in st.session_state["dup_results"].items():
            st.subheader(f"📊 Results: {name}")
            
            # Get summary statistics
            stats = detector.get_summary_stats(annotated)
            
            # Display key metrics
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                st.metric(
                    "Total Records",
                    f"{stats['total_records']:,}"
                )
            
            with col2:
                st.metric(
                    "🔴 Likely",
                    f"{stats['likely']}",
                    help=DuplicationScore.get_description(DuplicationScore.LIKELY)
                )
            
            with col3:
                st.metric(
                    "🟠 Somewhat",
                    f"{stats['somewhat']}",
                    help=DuplicationScore.get_description(DuplicationScore.SOMEWHAT_LIKELY)
                )
            
            with col4:
                st.metric(
                    "🟡 Possible",
                    f"{stats['possible']}",
                    help=DuplicationScore.get_description(DuplicationScore.POSSIBLE)
                )
            
            with col5:
                st.metric(
                    "🟣 No Name",
                    f"{stats['no_name']}",
                    help=DuplicationScore.get_description(DuplicationScore.NO_NAME)
                )

            # Summary message
            st.info(f"""
            **Summary**: {stats['total_flagged']} records ({stats['total_flagged_pct']}%) 
            flagged for review out of {stats['total_records']:,} total records.
            """)

            # Important note about indices
            st.warning("""
            **📌 Index Reference:**
            - **In this table**: `Duplicates_With` shows zero-based indices (0, 1, 2...)
            - **In Excel download**: Indices are adjusted to match Excel row numbers (2, 3, 4...)
            """)

            # Display the annotated dataframe
            st.dataframe(
                annotated,
                use_container_width=True,
                height=400,
                column_config={
                    "Duplication_Score": st.column_config.TextColumn(
                        "Score",
                        help="Duplication likelihood category",
                        width="medium"
                    ),
                    "Duplication_Reason": st.column_config.TextColumn(
                        "Reason",
                        help="Why this record was flagged",
                        width="large"
                    ),
                    "Duplicates_With": st.column_config.TextColumn(
                        "Matches With",
                        help="Row indices of matching records",
                        width="small"
                    )
                }
            )

            # Download buttons
            col1, col2 = st.columns(2)
            
            with col1:
                csv_data = annotated.to_csv(index=False)
                st.download_button(
                    "📥 Download CSV",
                    data=csv_data,
                    file_name=f"{name}_duplicates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    help="Download results as CSV file"
                )
            
            with col2:
                excel_buffer = detector.create_excel_with_highlights(annotated)
                st.download_button(
                    "📥 Download Excel (with highlights)",
                    data=excel_buffer.getvalue(),
                    file_name=f"{name}_duplicates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download color-coded Excel file with adjusted row numbers"
                )
            
            st.divider()

        # Add batch download option if multiple files
        if len(st.session_state["dup_results"]) > 1:
            if st.button("📦 Download All Results as ZIP"):
                # Implementation would go here
                st.info("ZIP download functionality would be implemented here")